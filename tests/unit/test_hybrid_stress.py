"""Focused semantic stress flag and content-free diagnostics workbook checks."""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path
from typing import Any

import openpyxl
import pytest


def _stress_module():
    path = Path("scripts/run_legal_chatbot_stress.py")
    spec = importlib.util.spec_from_file_location("hybrid_stress_runner", path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def test_hybrid_stress_parser_defaults_off_and_accepts_semantic_modes() -> None:
    runner = _stress_module()
    assert runner.build_parser().parse_args([]).semantic_mode == "off"
    assert runner.build_parser().parse_args(["--semantic-mode", "hybrid"]).semantic_mode == "hybrid"
    quality = runner.build_parser().parse_args(
        [
            "--semantic-mode",
            "hybrid",
            "--quality-strategy",
            "quality_retrieval_evidence_repair_v1",
            "--quality-selected-pool",
            "8",
        ]
    )
    assert quality.quality_strategy == "quality_retrieval_evidence_repair_v1"
    assert quality.quality_selected_pool == 8


def test_hybrid_stress_report_adds_content_free_semantic_diagnostics(tmp_path: Path) -> None:
    runner = _stress_module()
    output = tmp_path / "report.xlsx"
    runner.write_report(
        output,
        (),
        [],
        [],
        [],
        [],
        cleanup_status="NOT_NEEDED",
        infrastructure_status="COMPLETED",
        semantic_mode="hybrid",
        semantic_profile_id="e5-small-384-mean-l2-prefix-v1",
    )
    workbook = openpyxl.load_workbook(output, read_only=True)
    try:
        sheet = workbook["Semantic Diagnostics"]
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        assert headers == [
            "Case ID",
            "Mode",
            "Retrieval decision",
            "Retrieval strategy",
            "Citation count",
            "Source count",
            "Expected indexed-doc hit count",
            "Query embedding duration ms",
        ]
        assert "Question" not in headers
    finally:
        workbook.close()


def test_metadata_repair_stress_flag_and_sheet_are_content_free(tmp_path: Path) -> None:
    runner = _stress_module()
    assert runner.build_parser().parse_args(["--metadata-repair-enabled"]).metadata_repair_enabled
    output = tmp_path / "metadata.xlsx"
    runner.write_report(
        output,
        (),
        [],
        [],
        [],
        [],
        cleanup_status="NOT_NEEDED",
        infrastructure_status="COMPLETED",
        metadata_repair_enabled=True,
    )
    workbook = openpyxl.load_workbook(output, read_only=True)
    try:
        headers = [
            cell.value
            for cell in next(workbook["Metadata Repair Diagnostics"].iter_rows(max_row=1))
        ]
        assert headers[0:3] == ["Case ID", "Enabled", "Semantic candidates"]
        assert "Question" not in headers
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_hybrid_stress_timing_wrapper_retains_no_query_or_vector() -> None:
    runner = _stress_module()

    class Delegate:
        async def embed_query(self, text: str) -> object:
            del text
            return object()

        async def embed_documents(self, texts: object) -> object:
            return texts

    counters = runner.SemanticEmbeddingCounters()
    records: dict[str, Any] = {}
    wrapper = runner.LabelOnlySemanticEmbeddingPort(
        Delegate(), counters, case_id="Q01", records=records
    )
    await wrapper.embed_query("private question")
    assert counters.calls == counters.successes == 1
    assert counters.failures == 0
    assert records["Q01"].duration_ms >= 0
    assert "private question" not in repr(wrapper.__dict__)


@pytest.mark.asyncio
async def test_reranker_stress_wrapper_and_flag_are_bounded_and_content_free() -> None:
    runner = _stress_module()
    assert runner.build_parser().parse_args(["--rerank-enabled"]).rerank_enabled

    class Delegate:
        async def rerank(self, request: object) -> object:
            del request
            return object()

    counters = runner.RerankerCounters()
    records: dict[str, Any] = {}
    wrapper = runner.LabelOnlyRerankerPort(Delegate(), counters, "Q01", records)
    await wrapper.rerank(object())
    assert counters.calls == counters.successes == 1
    assert counters.failures == 0
    assert records["Q01"].outcome == "SUCCESS"
    assert "private" not in repr(wrapper.__dict__)
