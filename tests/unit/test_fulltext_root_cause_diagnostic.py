"""Unit checks for the privacy-preserving Prompt-01 diagnostic."""

from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import openpyxl
import pytest

from legal_chatbot.diagnostics.fulltext_root_cause import (
    CONTROLLED_INPUT_REFERENCE,
    NOT_EMITTED,
    NOT_IMPLEMENTED,
    Candidate,
    ControlledCase,
    FulltextRootCauseEvaluator,
    assert_oracle_not_in_model_inputs,
    build_queries,
    classify_root_cause,
    normalize_document_number,
    parse_controlled_workbook,
    write_reports,
)
from legal_chatbot.reranking.models import RerankCandidate, RerankRequest, RerankResult
from legal_chatbot.semantic.models import SemanticEmbeddingBatch


def _workbook(path: Path) -> None:
    book = openpyxl.Workbook()
    questions = book.active
    assert questions is not None
    questions.title = "Kết quả 10 câu"
    questions.append(["ID", "Câu hỏi"])
    scoring = book.create_sheet("Chấm điểm")
    scoring.append(["ID", "Chủ đề", "Văn bản kỳ vọng chính", "Nhận xét", "Hướng xử lý"])
    for number in range(1, 11):
        case_id = f"Q{number:02d}"
        questions.append([case_id, f"controlled-question-{case_id}"])
        scoring.append([case_id, "topic", f"{number}/QĐ-UEB", "review-only", "direction-only"])
    book.save(path)
    book.close()


def _actual_layout_workbook(path: Path) -> None:
    book = openpyxl.Workbook()
    questions = book.active
    assert questions is not None
    questions.title = "Kết quả 10 câu"
    questions.append(["Câu hỏi", "Văn bản lấy", "Nội dung chat trả lời"])
    scoring = book.create_sheet("Chấm điểm")
    scoring.append(
        [
            "ID",
            "Điểm /10",
            "PASS/FAIL",
            "Chủ đề",
            "Failure class",
            "Nhận xét toàn văn",
            "Văn bản đúng/kỳ vọng",
            "Hướng sửa",
            "Ưu tiên",
        ]
    )
    for number in range(1, 11):
        case_id = f"Q{number:02d}"
        questions.append([f"actual-question-{case_id}", "private-source", "private-answer"])
        scoring.append(
            [
                case_id,
                "8",
                "FAIL",
                "actual topic",
                "INSUFFICIENT_EVIDENCE" if number == 2 else "NO_DECOMPOSITION",
                "không đủ bằng chứng" if number == 2 else "review-comment-private",
                f"{number}/QĐ-UEB",
                "review-direction-private",
                "P1",
            ]
        )
    book.save(path)
    book.close()


def test_parser_separates_oracle_and_requires_exact_ten_cases(tmp_path: Path) -> None:
    path = tmp_path / "controlled.xlsx"
    _workbook(path)

    cases = parse_controlled_workbook(path)

    assert tuple(case.case_id for case in cases) == tuple(f"Q{item:02d}" for item in range(1, 11))
    assert cases[0].question == "controlled-question-Q01"
    assert cases[0].expected_documents == ("1/QĐ-UEB",)
    assert "review-only" not in cases[0].question


def test_parser_supports_actual_header_layout_and_ordered_question_rows(tmp_path: Path) -> None:
    path = tmp_path / "actual-layout.xlsx"
    _actual_layout_workbook(path)

    cases = parse_controlled_workbook(path)

    assert [case.case_id for case in cases] == [f"Q{item:02d}" for item in range(1, 11)]
    assert cases[0].question == "actual-question-Q01"
    assert cases[0].expected_documents == ("1/QĐ-UEB",)
    assert cases[0].review_failure_class == "NO_DECOMPOSITION"
    assert cases[0].review_priority == "P1"
    assert cases[0].review_score == "8"
    assert cases[0].review_pass == "FAIL"
    assert "review-comment-private" not in repr(cases[0])
    assert FulltextRootCauseEvaluator._review_requests_insufficiency(cases[1]) is True


def test_number_normalization_and_oracle_never_enters_abc_model_inputs() -> None:
    assert normalize_document_number("2725 /QĐ- ĐHKT") == normalize_document_number("2725/QĐ-ĐHKT")
    case = ControlledCase(
        "Q01", "What is the applicable rule?", "student policy", ("ORACLE-777/QĐ",)
    )
    queries = build_queries(case, ("UEB",))
    assert all("ORACLE-777/QĐ" not in value for value in queries.values())
    assert_oracle_not_in_model_inputs(case, queries, tuple(queries.values()))
    assert CONTROLLED_INPUT_REFERENCE == "CONTROLLED_INPUT_REFERENCE"


class _CapturingEmbedder:
    def __init__(self) -> None:
        self.inputs: list[str] = []

    async def embed_query(self, value: str) -> SemanticEmbeddingBatch:
        self.inputs.append(value)
        return SemanticEmbeddingBatch(vectors=((1.0,) + (0.0,) * 383,))


class _CapturingReranker:
    def __init__(self) -> None:
        self.request: RerankRequest | None = None

    async def rerank(self, request: RerankRequest) -> RerankResult:
        self.request = request
        return RerankResult.from_request(request, (1.0,) * len(request.candidates))


@pytest.mark.asyncio
async def test_expected_number_reaches_only_d_control_not_fake_embedder_or_reranker() -> None:
    expected = "ORACLE-777/QĐ"
    case = ControlledCase("Q01", "What is the applicable rule?", "student policy", (expected,))
    queries = build_queries(case, ("UEB",))
    embedder = _CapturingEmbedder()
    reranker = _CapturingReranker()

    for query in queries.values():
        await embedder.embed_query(query)
    await reranker.rerank(
        RerankRequest(
            query=queries["A"],
            candidates=(RerankCandidate(chunk_id="opaque", text="hydrated controlled text"),),
        )
    )

    assert expected not in "\n".join(embedder.inputs)
    assert reranker.request is not None
    assert expected not in reranker.request.query
    assert expected not in reranker.request.candidates[0].text
    assert case.expected_documents == (expected,)  # D exact-number control's sole input.


def test_classifier_fixtures_cover_measured_stages() -> None:
    assert (
        classify_root_cause("MISSING", found_top50=False, found_final=False)[0] == "CORPUS_MISSING"
    )
    assert (
        classify_root_cause("QUARANTINED", found_top50=False, found_final=False)[0]
        == "CORPUS_QUARANTINED"
    )
    assert (
        classify_root_cause("INDEXED", found_top50=True, found_final=False)[0]
        == "CANDIDATE_WINDOW_MISS"
    )
    assert (
        classify_root_cause("INDEXED", found_top50=True, found_final=False, rerank_demoted=True)[0]
        == "RERANK_DEMOTION"
    )
    assert (
        classify_root_cause("INDEXED", found_top50=True, found_final=True)[0]
        == "EXPECTED_DOCUMENT_FOUND_FINAL"
    )
    assert (
        classify_root_cause(
            "INDEXED", found_top50=False, found_final=False, candidate_present=True
        )[0]
        == "INSUFFICIENCY_RECHECK_REQUIRED_CANDIDATE_PRESENT"
    )


def test_capability_codes_are_requested_only_by_failure_class() -> None:
    assert FulltextRootCauseEvaluator._capability_codes("NO_DECOMPOSITION") == [
        "REVIEW_HYPOTHESIS_MULTI_INTENT_CAPABILITY_ABSENT"
    ]
    assert FulltextRootCauseEvaluator._capability_codes("HIERARCHY") == [
        "REVIEW_HYPOTHESIS_LEGAL_HIERARCHY_CAPABILITY_ABSENT"
    ]
    assert FulltextRootCauseEvaluator._capability_codes("VERSION_AMENDMENT") == [
        "REVIEW_HYPOTHESIS_VERSION_STATUS_CAPABILITY_ABSENT"
    ]


def test_reports_are_content_free_and_have_required_trace_and_table(tmp_path: Path) -> None:
    question = "RAW_QUESTION_SENTINEL"
    answer = "RAW_ANSWER_SENTINEL"
    chunk = "RAW_CHUNK_SENTINEL"
    uuid = str(uuid4())
    url = "https://private.example/sentinel"
    candidate = Candidate(uuid4(), uuid4(), "UEB", "1/QĐ-UEB", "Safe title", "ACTIVE", 1, 0)
    trace = {
        "raw_question": question,
        "normalized_question": question,
        "intent": NOT_IMPLEMENTED,
        "entities": NOT_IMPLEMENTED,
        "org": NOT_IMPLEMENTED,
        "legal_topics": NOT_IMPLEMENTED,
        "sub_intents": NOT_IMPLEMENTED,
        "query_plan": NOT_IMPLEMENTED,
        "expanded_queries": NOT_EMITTED,
        "corpus_insight_policy": NOT_IMPLEMENTED,
        "corpus_insight_decision": NOT_IMPLEMENTED,
        "candidate_lanes": {},
        "exact_number_control": [],
        "selected_evidence_diagnostic_final_top3": [candidate.safe(1)],
    }
    result = {
        "cases": [
            {
                "case_id": "Q01",
                "expected_documents": ["1/QĐ-UEB"],
                "documents": [
                    {
                        "expected_display_number": "1/QĐ-UEB",
                        "inventory": [],
                        "inventory_classification": "INDEXED",
                        "duplicates": False,
                        "raw_normalization_mismatch": False,
                        "is_indexed": True,
                        "lane_ranks": {
                            "A_PRODUCTION_EQUIVALENT_NATURAL_QUESTION_SEMANTIC": 4,
                        },
                        "title_metadata_rank": None,
                        "d_exact_control_rank": None,
                        "merged_top50_rank": 1,
                        "pre_rerank_semantic_top8_rank": 1,
                        "post_rerank_final_top3_rank": 1,
                        "found_merged_diagnostic_top50": True,
                        "found_final_top3": True,
                        "failure_stage": "FINAL_SELECTION",
                        "root_causes": ["EXPECTED_DOCUMENT_FOUND_FINAL"],
                        "confidence": "HIGH",
                        "rejected_evidence": [],
                        "review_comment": question,
                        "review_direction": answer,
                    }
                ],
                "found_top50_count": 1,
                "expected_indexed_count": 1,
                "found_final_count": 1,
                "found_top50_ratio": "1/1",
                "found_final_ratio": "1/1",
                "found_top50": True,
                "found_final": True,
                "root_causes": ["EXPECTED_DOCUMENT_FOUND_FINAL"],
                "failure_stages": ["FINAL_SELECTION"],
                "trace": trace,
                "review_comment": question,
                "review_direction": answer,
            }
        ],
        "blockers": [],
        "private": [question, answer, chunk, uuid, url],
    }

    paths = write_reports(result, tmp_path)

    content = "\n".join(path.read_text(encoding="utf-8") for path in paths)
    # Reports only serialize the documented diagnostic contract, not arbitrary input keys.
    for sentinel in (question, answer, chunk, uuid, url):
        assert sentinel not in content
    assert "| Q | Expected direct docs | Found top50?" in paths[0].read_text(encoding="utf-8")
    assert NOT_IMPLEMENTED in content
    assert 'raw_question": "CONTROLLED_INPUT_REFERENCE' in content
    assert 'final_answer_state": "NOT_RUN_DIAGNOSTIC_ONLY' in content
    assert 'found_top50_ratio": "1/1' in content
    assert 'failure_stages": [' in content
    assert "B_REVIEW_TOPIC_CONTROL" in content
    assert "C_ORACLE_SOURCE_SCOPE_CONTROL" in content
    assert "APPROVED_CORPUS_METADATA_AND_CONTROLLED_EXPECTED_IDENTITIES" in content
    assert '"aggregate_availability"' in content
    assert '"production_equivalent_natural_question_semantic_top50"' in content
    assert '"ratio": "1/1"' in content
    assert "Production-equivalent A semantic top50 availability: 1/1." in content
