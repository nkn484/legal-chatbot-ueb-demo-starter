"""Focused unit coverage for the evaluation-only M2 retrieval evaluator."""
# ruff: noqa: E501

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[2]


def _module():
    path = ROOT / "scripts" / "evaluate_m2_retrieval.py"
    spec = importlib.util.spec_from_file_location("m2_retrieval_evaluator", path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _parent(evaluator, case_id: str = "Q01"):
    return evaluator.ParentCase(case_id, "private parent question", ("101/test", "202/test"))


def _observation(evaluator, *, set_name="A", case_id="Q01", mode="S0", numbers=(), **kwargs):
    return evaluator.RetrievalObservation(
        set_name=set_name,
        case_id=case_id,
        mode=mode,
        category=kwargs.get("category", "PARENT"),
        parent_case_id=kwargs.get("parent_case_id", "Q01"),
        expected_behavior="SAFE",
        decision="EVIDENCE_AVAILABLE",
        reason="SEMANTIC_EVIDENCE_AVAILABLE",
        strategy="v-test",
        latency_ms=1.0,
        evidence_available=bool(numbers),
        citation_count=len(numbers),
        source_count=len(kwargs.get("sources", ())),
        citation_sources=tuple(kwargs.get("sources", ())),
        citation_document_numbers=tuple(numbers),
        citations_resolvable=kwargs.get("resolvable", True),
        failure_code=kwargs.get("failure_code", ""),
        cleanup_status="COMPLETED",
        diagnostics=kwargs.get("diagnostics", evaluator.SafeDiagnostics()),
    )


def test_parsers_keep_questions_separate_from_expected_identity_metadata(tmp_path: Path) -> None:
    evaluator = _module()
    workbook = openpyxl.Workbook()
    results = workbook.active
    assert results is not None
    results.title = "Kết quả 10 câu"
    results.append(("Câu hỏi", "Văn bản lấy"))
    for number in range(1, 11):
        results.append((f"private question {number}", "not used"))
    score = workbook.create_sheet("Chấm điểm")
    score.append(("",) * 7)
    score.append(("",) * 7)
    score.append(("ID", "", "", "", "", "", "Văn bản đúng/kỳ vọng"))
    for number in range(1, 11):
        score.append((f"Q{number:02d}", "", "", "", "", "", f"{number}/TEST"))
    path = tmp_path / "expert.xlsx"
    workbook.save(path)
    workbook.close()
    parsed = evaluator.parse_expert_workbook(path)
    assert parsed[0].expected_numbers == ("1/test",)
    assert "1/test" not in parsed[0].question


@pytest.mark.asyncio
async def test_execute_case_resolves_then_cleans_up_on_success_and_failure() -> None:
    evaluator = _module()
    cleaned = []

    class Repository:
        async def retrieve_and_persist(self, request):
            assert request.query == "private parent question"
            return SimpleNamespace(
                retrieval_run_id="run-1",
                candidates=(SimpleNamespace(citation_id="citation-1"),),
                decision=SimpleNamespace(value="EVIDENCE_AVAILABLE"),
                reason=SimpleNamespace(value="SEMANTIC_EVIDENCE_AVAILABLE"),
            )

    class Resolver:
        async def resolve(self, citation_id, run_id):
            assert (citation_id, run_id) == ("citation-1", "run-1")
            return SimpleNamespace(source_id="UEB", document_number="101/TEST")

    async def cleanup(run_id):
        cleaned.append(run_id)
        return "COMPLETED"

    observation = await evaluator.execute_case(
        _parent(evaluator), mode="S0", repository=Repository(), resolver=Resolver(), cleanup=cleanup,
        diagnostics=[],
    )
    assert cleaned == ["run-1"]
    assert observation.citation_document_numbers == ("101/TEST",)
    assert "private parent question" not in repr(observation)

    class FailingResolver:
        async def resolve(self, citation_id, run_id):
            del citation_id, run_id
            raise RuntimeError("resolver failure")

    failed = await evaluator.execute_case(
        _parent(evaluator), mode="S0", repository=Repository(), resolver=FailingResolver(), cleanup=cleanup,
        diagnostics=[],
    )
    assert cleaned == ["run-1", "run-1"]
    assert failed.failure_code == "SAFE_EXCEPTION"


def test_set_a_deduplicates_identity_and_excludes_unindexed_blockers() -> None:
    evaluator = _module()
    parent = _parent(evaluator)
    observation = _observation(
        evaluator, numbers=("101/TEST", "101/TEST", "999/OTHER"), sources=("UEB",)
    )
    score = evaluator.score_set_a((parent,), (observation,), {"101/test": ("UEB",)})[0]
    assert score.denominator == 1
    assert score.hit_numbers == ("101/test",)
    assert score.non_expected_cited_count == 1
    assert score.source_denominator == 1


def test_set_b_overlap_and_set_c_safe_invariant() -> None:
    evaluator = _module()
    parent = _observation(evaluator, numbers=("101/TEST", "202/TEST"))
    paraphrase = _observation(
        evaluator, set_name="B", case_id="B-Q01-01", numbers=("202/TEST", "303/TEST")
    )
    score = evaluator.score_set_b((parent, paraphrase))[0]
    assert score.overlap_count == 1
    assert score.jaccard == pytest.approx(1 / 3)
    control = _observation(
        evaluator,
        set_name="C",
        case_id="C-SM-01",
        mode="S1",
        category="SYNTHETIC_METADATA_NUMBER",
        diagnostics=evaluator.SafeDiagnostics(identity_count=1),
    )
    invariant = evaluator.check_set_c_invariants((control,))[0]
    assert invariant.invariant_failures == ("ARBITRARY_IDENTITY_CANDIDATE",)


def test_reports_use_required_sheets_without_private_question_and_gate_is_mechanical(tmp_path: Path) -> None:
    evaluator = _module()
    parent = _parent(evaluator)
    s0 = _observation(evaluator, mode="S0", numbers=("101/TEST",), sources=("UEB",))
    s1 = _observation(evaluator, mode="S1", numbers=("101/TEST", "202/TEST"), sources=("UEB",))
    scores = evaluator.score_set_a((parent,), (s0, s1), {"101/test": ("UEB",), "202/test": ("UEB",)})
    xlsx, report_json, markdown = tmp_path / "report.xlsx", tmp_path / "report.json", tmp_path / "report.md"
    evaluator.write_reports(
        xlsx_path=xlsx, json_path=report_json, markdown_path=markdown, observations=(s0, s1),
        set_a=scores, set_b=(), set_c=(), arm_diagnostics=(),
    )
    workbook = openpyxl.load_workbook(xlsx, read_only=True)
    try:
        assert workbook.sheetnames == ["Summary", "Set A", "Set B", "Set C", "Arm Diagnostics"]
        assert "private parent question" not in str(list(workbook["Set A"].values))
    finally:
        workbook.close()
    assert "private parent question" not in report_json.read_text(encoding="utf-8")
    gate = evaluator.mechanical_gate(
        {"S0": {"set_a_identity_hits": 1, "set_a_non_expected_cited_rate": 0.5},
         "S1": {"set_a_identity_hits": 2, "set_a_non_expected_cited_rate": 0.5,
                "set_c_invariant_failures": 0}}
    )
    assert gate["mechanical_gates_pass"] is True
    assert gate["recommendation"] == "HOLD_PENDING_ORACLE"
