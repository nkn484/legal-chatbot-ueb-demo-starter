"""Focused unit coverage for the standalone legal-chatbot stress runner."""

from __future__ import annotations

import asyncio
import importlib.util
import sys
from pathlib import Path
from types import ModuleType, SimpleNamespace

import openpyxl
import pytest


def _runner() -> ModuleType:
    path = Path(__file__).parents[2] / "scripts" / "run_legal_chatbot_stress.py"
    spec = importlib.util.spec_from_file_location("stress_runner", path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _input_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    questions = workbook.active
    questions.title = "Câu hỏi"
    questions.append(["Title"])
    questions.append(["Method"])
    questions.append([])
    questions.append(["ID", "Câu hỏi"])
    grading = workbook.create_sheet("Khung chấm")
    grading.append(["Title"])
    grading.append(["Method"])
    grading.append([])
    grading.append(
        [
            "ID", "Chủ đề", "Nhóm nguồn kỳ vọng", "Văn bản kỳ vọng chính", "Tối thiểu",
            "Điều kiện PASS",
        ]
    )
    for number in range(1, 11):
        case_id = f"Q{number:02d}"
        questions.append([case_id, f"question-{case_id}"])
        grading.append(
            [
                case_id,
                "topic",
                "VBQPPL, UEB",
                "08/2021/TT-BGDĐT; 123/QĐ-UEB",
                "1",
                "grading-only-text",
            ]
        )
    workbook.save(path)
    workbook.close()


def test_workbook_parser_keeps_grading_out_of_question_input(tmp_path: Path) -> None:
    runner = _runner()
    input_path = tmp_path / "input.xlsx"
    _input_workbook(input_path)

    cases = runner.parse_stress_workbook(input_path)

    assert len(cases) == 10
    assert cases[0].question == "question-Q01"
    assert "grading-only-text" not in cases[0].question
    assert cases[0].expected_sources == ("VBQPPL", "UEB")
    assert cases[0].expected_documents == ("08/2021/TT-BGDĐT", "123/QĐ-UEB")


def test_percentiles_metrics_and_post_scoring() -> None:
    runner = _runner()
    assert runner.percentile([1, 2, 3, 4], 50) == 2
    assert runner.percentile([1, 2, 3, 4], 95) == 4
    case = runner.StressCase(
        "Q01", "question", expected_sources=("VBQPPL", "UEB"),
        expected_documents=("1/QĐ-UEB", "2/QĐ-UEB"),
    )
    score = runner.score_case(
        case,
        (runner.CitationView("UEB", "1/QĐ-UEB", None, None, "OFFICIAL_LEGAL"),),
        {"1/QĐ-UEB": "INDEXED", "2/QĐ-UEB": "QUARANTINED"},
    )
    assert score.source_coverage_percent == 50.0
    assert score.expected_document_hits == ("1/QĐ-UEB",)
    assert score.corpus_blockers == ("2/QĐ-UEB:QUARANTINED",)
    assert score.legal_correctness == "NOT_MEASURED_REQUIRES_HUMAN_REVIEW"


def test_report_has_required_sheet_structure(tmp_path: Path) -> None:
    runner = _runner()
    output = tmp_path / "report.xlsx"
    case = runner.StressCase("Q01", "question")
    row = runner.CallRow("MECHANICAL", "Q01", 1, 1, 5, "ANSWER", "ANSWER_GROUNDED", 1, 1, "")
    runner.write_report(
        output, (case,), [row], [], [], [runner.score_case(case, (), {})],
        cleanup_status="COMPLETED", infrastructure_status="COMPLETED",
    )
    workbook = openpyxl.load_workbook(output, read_only=True)
    try:
        assert workbook.sheetnames == [
            "Tổng hợp", "Câu hỏi & Kết quả", "Mechanical Load", "API Probe", "Corpus Coverage",
            "Planner Diagnostics", "Semantic Diagnostics", "Reranker Diagnostics",
            "Metadata Repair Diagnostics", "Provider Output Diagnostics", "Regression",
        ]
    finally:
        workbook.close()


def test_parser_flags_are_independent_and_default_disabled() -> None:
    runner = _runner()
    defaults = runner.build_parser().parse_args([])
    assert defaults.lexical_repair_enabled is False
    assert defaults.planner_enabled is False
    lexical = runner.build_parser().parse_args(["--lexical-repair-enabled"])
    planner = runner.build_parser().parse_args(["--planner-enabled"])
    both = runner.build_parser().parse_args(
        ["--lexical-repair-enabled", "--planner-enabled"]
    )
    assert (lexical.lexical_repair_enabled, lexical.planner_enabled) == (True, False)
    assert (planner.lexical_repair_enabled, planner.planner_enabled) == (False, True)
    assert (both.lexical_repair_enabled, both.planner_enabled) == (True, True)


def test_planner_diagnostic_sheet_is_content_free_and_exact(tmp_path: Path) -> None:
    runner = _runner()
    output = tmp_path / "report.xlsx"
    question_sentinel = "QUESTION_SENTINEL_MUST_NOT_BE_HERE"
    plan_sentinel = "PLAN_SENTINEL_MUST_NOT_BE_HERE"
    case = runner.StressCase("Q01", question_sentinel, expected_documents=("1/QĐ-UEB",))
    row = runner.CallRow(
        "REAL_SHINE", "Q01", 1, 1, 4, "REFUSAL", "INVALID_PROVIDER_OUTPUT", 0, 0,
        "INVALID_PROVIDER_OUTPUT", retrieval_run_id="run-1",
    )
    diagnostic = runner.RetrievalDiagnostic(
        True,
        "EVIDENCE_AVAILABLE",
        "LEXICAL_EVIDENCE_AVAILABLE",
        "v1",
        (
            runner.CitationView(
                "UEB", "1/QĐ-UEB", "title", "https://example.test", "OFFICIAL_LEGAL"
            ),
        ),
    )
    runner.apply_retrieval_diagnostics([row], {"run-1": diagnostic})
    coverage = runner.score_real_cases((case,), [row], {"1/QĐ-UEB": "INDEXED"})
    records = {
        "Q01": runner.PlannerRecord(
            "Q01", runner.QueryPlannerOutcome.PLANNED, 2.5
        )
    }
    counters = runner.ProviderCounterSnapshot(1, 1, 0)
    runner.write_report(
        output,
        (case,),
        [],
        [row],
        [],
        coverage,
        cleanup_status="COMPLETED",
        infrastructure_status="COMPLETED",
        planner_enabled=True,
        planner_records=records,
        planner_by_case={"Q01": counters},
        answer_by_case={"Q01": counters},
    )
    workbook = openpyxl.load_workbook(output, read_only=True)
    try:
        sheet = workbook["Planner Diagnostics"]
        headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        assert headers == (
            "Case ID", "Planner enabled", "Planner called", "Planner outcome",
            "Planner duration ms", "Planner provider calls", "Planner provider success",
            "Planner provider failure", "Answer provider calls", "Answer provider success",
            "Answer provider failure", "Chat outcome", "Chat reason", "Retrieval run present",
            "Retrieval decision", "Retrieval reason", "Retrieval strategy", "Citation count",
            "Source count", "Expected indexed-doc hit count",
        )
        values = " ".join(
            str(value or "")
            for row_values in sheet.iter_rows(values_only=True)
            for value in row_values
        )
        assert question_sentinel not in values
        assert plan_sentinel not in values
    finally:
        workbook.close()


class _NoRetentionProvider:
    def __init__(self) -> None:
        self.fail = False

    async def generate(self, request: object) -> object:
        del request
        if self.fail:
            raise RuntimeError("safe")
        return object()

    async def health_check(self) -> object:
        return object()


class _BlockingProvider:
    async def generate(self, request: object) -> object:
        del request
        await asyncio.sleep(60)
        return object()

    async def health_check(self) -> object:
        return object()


@pytest.mark.asyncio
async def test_label_only_provider_proxies_count_separately_without_text_retention() -> None:
    runner = _runner()
    delegate = _NoRetentionProvider()
    planner_counters = runner.ProviderCallCounters()
    answer_counters = runner.ProviderCallCounters()
    planner_proxy = runner.LabelOnlyProviderProxy("planner", delegate, planner_counters)
    answer_proxy = runner.LabelOnlyProviderProxy("answer", delegate, answer_counters)
    request_sentinel = "REQUEST_PROMPT_SENTINEL"

    await planner_proxy.generate(request_sentinel)
    await answer_proxy.generate(request_sentinel)
    delegate.fail = True
    with pytest.raises(RuntimeError):
        await planner_proxy.generate(request_sentinel)

    assert (
        planner_counters.calls,
        planner_counters.successes,
        planner_counters.failures,
    ) == (2, 1, 1)
    assert (answer_counters.calls, answer_counters.successes, answer_counters.failures) == (1, 1, 0)
    assert request_sentinel not in repr(planner_proxy.__dict__)
    assert request_sentinel not in repr(answer_proxy.__dict__)
    assert (planner_proxy.label, answer_proxy.label) == ("planner", "answer")
    assert not hasattr(planner_proxy, "request")
    await planner_proxy.aclose()
    assert delegate.fail is True


@pytest.mark.asyncio
async def test_label_only_provider_counts_timeout_cancellation_as_failure() -> None:
    runner = _runner()
    counters = runner.ProviderCallCounters()
    proxy = runner.LabelOnlyProviderProxy("planner", _BlockingProvider(), counters)

    with pytest.raises(TimeoutError):
        await asyncio.wait_for(proxy.generate(object()), timeout=0.01)

    assert (counters.calls, counters.successes, counters.failures) == (1, 0, 1)


class _PlannerDelegate:
    def __init__(self, outcome: object) -> None:
        self._outcome = outcome

    async def plan(self, question: str) -> object:
        del question
        return SimpleNamespace(outcome=self._outcome)


@pytest.mark.asyncio
async def test_recording_planner_retains_only_outcome_and_duration() -> None:
    runner = _runner()
    records: dict[str, object] = {}
    planner = runner.RecordingPlanner(
        _PlannerDelegate(runner.QueryPlannerOutcome.PLANNED), "Q01", records
    )
    await planner.plan("PLANNER_INPUT_SENTINEL")

    record = records["Q01"]
    assert record.outcome is runner.QueryPlannerOutcome.PLANNED
    assert record.duration_ms >= 0
    assert "PLANNER_INPUT_SENTINEL" not in repr(records)
    assert set(record.__dict__) == {"case_id", "outcome", "duration_ms"}


def test_invalid_answer_output_keeps_persisted_retrieval_diagnostics_and_coverage() -> None:
    runner = _runner()
    case = runner.StressCase(
        "Q01", "question", expected_sources=("UEB",), expected_documents=("1/QĐ-UEB",)
    )
    row = runner.CallRow(
        "REAL_SHINE", "Q01", 1, 1, 10, "REFUSAL", "INVALID_PROVIDER_OUTPUT", 0, 0,
        "INVALID_PROVIDER_OUTPUT", retrieval_run_id="captured-run",
    )
    evidence = runner.CitationView("UEB", "1/QĐ-UEB", "title", "url", "OFFICIAL_LEGAL")
    runner.apply_retrieval_diagnostics(
        [row],
        {
            "captured-run": runner.RetrievalDiagnostic(
                True, "EVIDENCE_AVAILABLE", "LEXICAL_EVIDENCE_AVAILABLE", "v1", (evidence,)
            )
        },
    )

    coverage = runner.score_real_cases((case,), [row], {"1/QĐ-UEB": "INDEXED"})[0]
    assert row.outcome == "REFUSAL"
    assert row.reason == "INVALID_PROVIDER_OUTPUT"
    assert row.retrieval_decision == "EVIDENCE_AVAILABLE"
    assert row.citations == (evidence,)
    assert coverage.retrieved_sources == ("UEB",)
    assert coverage.expected_document_hits == ("1/QĐ-UEB",)


def test_provider_output_diagnostics_are_content_free_and_counted(tmp_path: Path) -> None:
    runner = _runner()
    output = tmp_path / "report.xlsx"
    case = runner.StressCase("Q01", "QUESTION_SENTINEL")
    row = runner.CallRow(
        "REAL_SHINE",
        "Q01",
        1,
        1,
        5,
        "REFUSAL",
        "INVALID_PROVIDER_OUTPUT",
        1,
        1,
        "INVALID_PROVIDER_OUTPUT",
        answer="PROVIDER_OUTPUT_SENTINEL",
        retrieval_run_id="run-1",
        retrieval_run_present=True,
        provider_output_class="JSON_SYNTAX",
    )
    runner.write_report(
        output,
        (case,),
        [],
        [row],
        [],
        [runner.score_case(case, (), {})],
        cleanup_status="COMPLETED",
        infrastructure_status="COMPLETED",
        answer_by_case={"Q01": runner.ProviderCounterSnapshot(1, 0, 0)},
    )
    workbook = openpyxl.load_workbook(output, read_only=True)
    try:
        sheet = workbook["Provider Output Diagnostics"]
        assert next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) == (
            "Case ID",
            "Outcome",
            "Reason",
            "Retrieval run present",
            "Citation count",
            "Provider call count",
            "Output class",
        )
        assert next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)) == (
            "Q01",
            "REFUSAL",
            "INVALID_PROVIDER_OUTPUT",
            True,
            1,
            1,
            "JSON_SYNTAX",
        )
        values = " ".join(
            str(value or "")
            for row_values in sheet.iter_rows(values_only=True)
            for value in row_values
        )
        assert "QUESTION_SENTINEL" not in values
        assert "PROVIDER_OUTPUT_SENTINEL" not in values
        summary = workbook["Tổng hợp"]
        assert ("Provider output class JSON_SYNTAX", 1) in tuple(summary.values)
    finally:
        workbook.close()


def test_label_only_provider_output_parser_retains_only_safe_class() -> None:
    runner = _runner()
    records: dict[str, str] = {}
    parser = runner.LabelOnlyProviderOutputParser(
        runner.StrictProviderJsonParser(), "Q01", records
    )

    with pytest.raises(runner.ChatError):
        parser.parse('{"answer":"OUTPUT_SENTINEL https://example.test"}')

    assert records == {"Q01": "ANSWER_URL"}
    assert "OUTPUT_SENTINEL" not in repr(parser.__dict__)


def test_label_only_provider_output_parser_records_none_for_safe_multiline_answer() -> None:
    runner = _runner()
    records: dict[str, str] = {}
    parser = runner.LabelOnlyProviderOutputParser(
        runner.StrictProviderJsonParser(), "Q01", records
    )

    parser.parse('{"answer":"First paragraph.\\r\\nSecond paragraph.\\tDetail."}')

    assert records == {"Q01": "NONE"}


@pytest.mark.asyncio
async def test_label_only_provider_output_provider_records_only_preparser_class() -> None:
    runner = _runner()
    records: dict[str, str] = {}
    provider = runner.LabelOnlyProviderOutputProvider(
        _NoRetentionProvider(), "Q01", records, max_response_bytes=1024
    )

    await provider.generate(object())

    assert records == {"Q01": "PORT_RESULT_TYPE"}
    assert not hasattr(provider, "result")


def test_provider_output_diagnostics_mark_mechanical_and_nonrun_cases(tmp_path: Path) -> None:
    runner = _runner()
    output = tmp_path / "report.xlsx"
    cases = (runner.StressCase("Q01", "question"), runner.StressCase("Q02", "question"))
    mechanical = [
        runner.CallRow("MECHANICAL", "Q01", 1, 1, 5, "ANSWER", "ANSWER_GROUNDED", 1, 1, "")
    ]
    runner.write_report(
        output,
        cases,
        mechanical,
        [],
        [],
        [runner.score_case(case, (), {}) for case in cases],
        cleanup_status="COMPLETED",
        infrastructure_status="COMPLETED",
    )
    workbook = openpyxl.load_workbook(output, read_only=True)
    try:
        sheet = workbook["Provider Output Diagnostics"]
        assert tuple(sheet.iter_rows(min_row=2, max_row=3, values_only=True)) == (
            ("Q01", "ANSWER", "ANSWER_GROUNDED", False, 1, 0, "NOT_APPLICABLE"),
            ("Q02", "NOT_RUN", "NOT_RUN", False, 0, 0, "NOT_RUN"),
        )
    finally:
        workbook.close()
