"""Unit coverage for the read-only Gate-3 exact-versus-HNSW evaluator."""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path
from uuid import uuid4

import openpyxl
import pytest


def _module():
    path = Path("scripts/evaluate_semantic_ann.py")
    spec = importlib.util.spec_from_file_location("semantic_ann_evaluator", path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    questions = workbook.active
    questions.title = "Câu hỏi"
    questions.append(("", ""))
    questions.append(("", ""))
    questions.append(("", ""))
    questions.append(("", ""))
    grading = workbook.create_sheet("Khung chấm")
    for _ in range(4):
        grading.append(("", "", "", ""))
    for number in range(1, 11):
        case_id = f"Q{number:02d}"
        questions.append((case_id, f"question {number}"))
        grading.append((case_id, "topic", "VBQPPL", f"{number}/TEST"))
    workbook.save(path)
    workbook.close()


def test_semantic_ann_parser_keeps_grading_metadata_separate(tmp_path: Path) -> None:
    evaluator = _module()
    workbook = tmp_path / "input.xlsx"
    _workbook(workbook)
    cases = evaluator.parse_cases(workbook)
    assert len(cases) == 10
    assert cases[0].question == "question 1"
    assert cases[0].expected_documents == ("1/TEST",)
    assert "1/TEST" not in cases[0].question


def test_semantic_ann_metrics_and_recursive_explain_index_detection() -> None:
    evaluator = _module()
    first, second, third = uuid4(), uuid4(), uuid4()
    assert evaluator.overlap_percent((first, second), (second, third)) == 100 / 3
    assert evaluator.recall_at_k((first, second), (second, third)) == 0.5
    plan = [{"Plan": {"Plans": [{"Index Name": "ix_chunk_embeddings_embedding_hnsw_cosine"}]}}]
    assert evaluator._plan_contains_index(plan, "ix_chunk_embeddings_embedding_hnsw_cosine")
    assert not evaluator._plan_contains_index(plan, "another_index")


@pytest.mark.asyncio
async def test_semantic_ann_incomplete_coverage_writes_safe_blocker_without_embedding(
    tmp_path: Path,
) -> None:
    evaluator = _module()
    case = evaluator.EvaluationCase("Q01", "private question", ("1/TEST",))

    class Embedder:
        async def embed_query(self, text: str) -> object:
            raise AssertionError(text)

    async def incomplete(session_factory, sources) -> bool:
        del session_factory, sources
        return False

    result = await evaluator.evaluate(
        (case,),
        session_factory=object(),
        embedder=Embedder(),
        top_k=8,
        ef_search=40,
        coverage_checker=incomplete,
    )
    assert result.blocker == "SEMANTIC_COVERAGE_INCOMPLETE"
    assert result.cases == ()
    output = tmp_path / "report.xlsx"
    evaluator.write_report(output, result, top_k=8, ef_search=40)
    workbook = openpyxl.load_workbook(output, read_only=True)
    try:
        assert workbook.sheetnames == ["Summary", "Cases"]
        summary = workbook["Summary"]
        values = [row[1] for row in summary.iter_rows(min_row=2, values_only=True)]
        assert "SEMANTIC_COVERAGE_INCOMPLETE" in values
        assert "private question" not in str(values)
    finally:
        workbook.close()
