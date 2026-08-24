"""Read-only Phase B.1 retrieval-engine root-cause diagnostic helpers.

This module is deliberately evaluation-only.  Query strings, tsquery text, SQL,
plans, vectors, and evidence text remain short-lived and are never serialized.
"""

# ruff: noqa: E501
from __future__ import annotations

import json
import re
from collections import Counter
from collections.abc import Iterable, Mapping
from dataclasses import dataclass, field
from enum import StrEnum
from pathlib import Path
from typing import Any

import openpyxl

from legal_chatbot.documents.metadata_normalization import normalize_document_number

SOURCE_IDS = ("VBQPPL", "VNU", "UEB")
POOL_SIZES = (8, 12, 16, 20)
_LEXEME = re.compile(r"'(?:[^']|'')*'")
_TOKEN = re.compile(r"[^\W_]+", re.UNICODE)


class RootCauseCode(StrEnum):
    CONFIG_MISMATCH = "FTS_INDEX_CONFIGURATION_MISMATCH"
    FILTER_EXCLUSION = "FTS_FILTER_EXCLUSION"
    IDENTITY_COLLAPSE_LOSS = "FTS_IDENTITY_COLLAPSE_LOSS"
    FUSION_LOSS = "FTS_FUSION_LOSS"
    QUERY_CONSTRUCTION_FAILURE = "FTS_QUERY_CONSTRUCTION_FAILURE"
    FTS_NO_MATCH = "FTS_NO_MATCH"
    NO_ROOT_CAUSE_OBSERVED = "NO_FTS_ROOT_CAUSE_OBSERVED"


class UnsafeReportFieldError(ValueError):
    """Raised with only the unsafe field name, never its private value."""

    def __init__(self, field_name: str) -> None:
        super().__init__("unsafe report field")
        self.field_name = field_name


@dataclass(frozen=True)
class DiagnosticCase:
    case_id: str
    question: str = field(repr=False)
    expected_numbers: tuple[str, ...]


@dataclass(frozen=True)
class TsqueryStructure:
    """Content-free shape summary; lexeme values are never retained."""

    lexeme_count: int
    and_operator_count: int
    phrase_operator_count: int
    numnode: int
    token_alias_distribution: dict[str, int]
    accented_token_count: int
    abbreviation_like_token_count: int
    truncated_to_32: bool = False

    def safe(self) -> dict[str, object]:
        return {
            "lexeme_count": self.lexeme_count,
            "and_operator_count": self.and_operator_count,
            "phrase_operator_count": self.phrase_operator_count,
            "numnode": self.numnode,
            "token_alias_distribution": self.token_alias_distribution,
            "accented_token_count": self.accented_token_count,
            "abbreviation_like_token_count": self.abbreviation_like_token_count,
            "truncated_to_32": self.truncated_to_32,
        }


@dataclass(frozen=True)
class PlanSummary:
    root_node_type: str
    node_types: tuple[str, ...]
    scan_types: tuple[str, ...]
    index_names: tuple[str, ...]
    actual_rows: int
    planning_ms: float
    execution_ms: float
    shared_hit: int
    shared_read: int
    temp_read: int
    temp_written: int
    limit_above_scan: bool

    def safe(self) -> dict[str, object]:
        return {
            "root_node_type": self.root_node_type,
            "node_types": list(self.node_types),
            "scan_types": list(self.scan_types),
            "index_names": list(self.index_names),
            "actual_rows": self.actual_rows,
            "planning_ms": self.planning_ms,
            "execution_ms": self.execution_ms,
            "buffers": {
                "shared_hit": self.shared_hit,
                "shared_read": self.shared_read,
                "temp_read": self.temp_read,
                "temp_written": self.temp_written,
            },
            "limit_above_scan": self.limit_above_scan,
        }


@dataclass(frozen=True)
class LaneEvidence:
    config_matches_simple: bool
    gin_index_valid: bool
    natural_unfiltered_rows: int
    natural_filtered_rows: int
    natural_unfiltered_expected: bool
    natural_filtered_expected: bool
    collapsed_expected: bool
    fused_expected: bool
    or_expected: bool


def parse_tsquery_structure(
    tsquery_text: str, *, source_question: str = "", max_lexemes: int = 32
) -> TsqueryStructure:
    """Return counts only from private tsquery/question text; never return tokens or lexemes."""
    lexeme_count = len(_LEXEME.findall(tsquery_text))
    tokens = _TOKEN.findall(source_question)
    aliases = Counter(
        "NUMERIC"
        if token.isdecimal()
        else "ALPHA_NUMERIC"
        if any(char.isdigit() for char in token)
        else "ALPHA"
        for token in tokens
    )
    return TsqueryStructure(
        lexeme_count=lexeme_count,
        and_operator_count=tsquery_text.count("&"),
        phrase_operator_count=tsquery_text.count("<->") + tsquery_text.count("<"),
        numnode=lexeme_count
        + tsquery_text.count("&")
        + tsquery_text.count("|")
        + tsquery_text.count("<"),
        token_alias_distribution=dict(sorted(aliases.items())),
        accented_token_count=sum(any(ord(char) > 127 for char in token) for token in tokens),
        abbreviation_like_token_count=sum(
            len(token) <= 5 and token.isupper() for token in source_question.split()
        ),
        truncated_to_32=lexeme_count > max_lexemes,
    )


def safe_plan_summary(payload: object) -> PlanSummary:
    """Extract allowlisted numeric/operator metadata without retaining a raw SQL plan."""
    root = payload
    if isinstance(root, str):
        try:
            root = json.loads(root)
        except json.JSONDecodeError:
            root = {}
    if isinstance(root, list):
        root = root[0] if root else {}
    if not isinstance(root, Mapping):
        root = {}
    plan = root.get("Plan", root)
    if not isinstance(plan, Mapping):
        plan = {}
    node_types: list[str] = []
    scan_types: list[str] = []
    indexes: list[str] = []
    rows = 0
    buffers = Counter()

    def visit(node: object) -> None:
        nonlocal rows
        if not isinstance(node, Mapping):
            return
        node_type = str(node.get("Node Type", "UNKNOWN"))
        node_types.append(node_type)
        if "Scan" in node_type:
            scan_types.append(node_type)
        index_name = node.get("Index Name")
        if isinstance(index_name, str):
            indexes.append(index_name)
        rows += int(node.get("Actual Rows", 0) or 0)
        for key, target in (
            ("Shared Hit Blocks", "hit"),
            ("Shared Read Blocks", "read"),
            ("Temp Read Blocks", "temp_read"),
            ("Temp Written Blocks", "temp_written"),
        ):
            buffers[target] += int(node.get(key, 0) or 0)
        for child in node.get("Plans", ()) or ():
            visit(child)

    visit(plan)
    return PlanSummary(
        root_node_type=str(plan.get("Node Type", "UNKNOWN")),
        node_types=tuple(node_types),
        scan_types=tuple(scan_types),
        index_names=tuple(sorted(set(indexes))),
        actual_rows=rows,
        planning_ms=float(root.get("Planning Time", 0) or 0),
        execution_ms=float(root.get("Execution Time", 0) or 0),
        shared_hit=buffers["hit"],
        shared_read=buffers["read"],
        temp_read=buffers["temp_read"],
        temp_written=buffers["temp_written"],
        limit_above_scan=limit_above_scan(plan),
    )


def limit_above_scan(plan: Mapping[str, object]) -> bool:
    """Prove a LIMIT occurs above at least one scan node in an already-private plan object."""

    def has_scan(node: object) -> bool:
        if not isinstance(node, Mapping):
            return False
        if "Scan" in str(node.get("Node Type", "")):
            return True
        return any(has_scan(child) for child in node.get("Plans", ()) or ())

    if str(plan.get("Node Type", "")) == "Limit" and has_scan(plan):
        return True
    return any(
        limit_above_scan(child)
        for child in plan.get("Plans", ()) or ()
        if isinstance(child, Mapping)
    )


def classify_lane(evidence: LaneEvidence) -> RootCauseCode:
    """Apply the approved, ordered classification precedence exactly."""
    if not evidence.config_matches_simple or not evidence.gin_index_valid:
        return RootCauseCode.CONFIG_MISMATCH
    if evidence.natural_unfiltered_expected and not evidence.natural_filtered_expected:
        return RootCauseCode.FILTER_EXCLUSION
    if evidence.natural_filtered_expected and not evidence.collapsed_expected:
        return RootCauseCode.IDENTITY_COLLAPSE_LOSS
    if evidence.collapsed_expected and not evidence.fused_expected:
        return RootCauseCode.FUSION_LOSS
    if not evidence.natural_filtered_expected and evidence.or_expected:
        return RootCauseCode.QUERY_CONSTRUCTION_FAILURE
    if not evidence.natural_filtered_expected and not evidence.or_expected:
        return RootCauseCode.FTS_NO_MATCH
    return RootCauseCode.NO_ROOT_CAUSE_OBSERVED


def aggregate_root_cause(rows: Iterable[Mapping[str, object]], lane: str) -> dict[str, object]:
    """Return a primary cause only for a decisive (strict-majority) observed distribution."""
    distribution = Counter(str(row["classification"]) for row in rows if row.get("lane") == lane)
    total = sum(distribution.values())
    primary, count = distribution.most_common(1)[0] if distribution else (None, 0)
    return {
        "distribution": dict(sorted(distribution.items())),
        "primary": primary if primary is not None and count > total / 2 else "MIXED",
        "decisive": bool(primary is not None and count > total / 2),
    }


def q6_trace(
    *,
    expected_numbers: Iterable[str],
    diagnostic: Iterable[Any],
    pool20: Iterable[Any],
    final3: Iterable[Any],
) -> tuple[dict[str, object], ...]:
    """Post-score-only Q6 trace with approved document numbers, ranks, and safe scores."""
    expected = tuple(expected_numbers)
    diagnostic, pool20, final3 = tuple(diagnostic), tuple(pool20), tuple(final3)
    result = []
    for number in expected:
        diagnostic_item = next(
            (item for item in diagnostic if item.identity.document_number_normalized == number),
            None,
        )
        pool_item = next(
            (item for item in pool20 if item.identity.document_number_normalized == number), None
        )
        final_item = next(
            (item for item in final3 if item.identity.document_number_normalized == number), None
        )
        if diagnostic_item is None:
            reason = "NOT_IN_DIAGNOSTIC_TOP50"
        elif pool_item is None:
            reason = "POOL20_CUTOFF"
        elif final_item is None:
            reason = "FINAL_TOP3_CUTOFF"
        else:
            reason = "SELECTED_FINAL_TOP3"
        item = diagnostic_item or pool_item or final_item
        aggregates = (
            {aggregate.lane.value: aggregate for aggregate in item.lane_aggregates} if item else {}
        )
        result.append(
            {
                "document_number": number,
                "diagnostic_rank": diagnostic.index(diagnostic_item) + 1
                if diagnostic_item
                else None,
                "pool20_rank": pool20.index(pool_item) + 1 if pool_item else None,
                "final_rank": final3.index(final_item) + 1 if final_item else None,
                "fusion_score": item.fusion_score if item else None,
                "semantic_rank": aggregates.get("SEMANTIC").best_rank
                if "SEMANTIC" in aggregates
                else None,
                "semantic_score": aggregates.get("SEMANTIC").best_score
                if "SEMANTIC" in aggregates
                else None,
                "content_rank": aggregates.get("CONTENT_FTS").best_rank
                if "CONTENT_FTS" in aggregates
                else None,
                "content_score": aggregates.get("CONTENT_FTS").best_score
                if "CONTENT_FTS" in aggregates
                else None,
                "title_rank": aggregates.get("TITLE_FTS").best_rank
                if "TITLE_FTS" in aggregates
                else None,
                "title_score": aggregates.get("TITLE_FTS").best_score
                if "TITLE_FTS" in aggregates
                else None,
                "rejection_reason": reason,
            }
        )
    return tuple(result)


def percentile(values: Iterable[float], percent: int) -> float:
    values = sorted(values)
    if not values:
        return 0.0
    return round(values[max(0, min(len(values) - 1, (len(values) * percent + 99) // 100 - 1))], 3)


def gate_b1(
    *,
    lane_rows: Iterable[dict[str, object]],
    lane_aggregates: Mapping[str, Mapping[str, object]],
    complete_fields: Mapping[str, bool],
) -> str:
    """Pass only when every derived Phase-B1 evidence field is complete."""

    rows = tuple(lane_rows)
    material = [row for row in rows if row.get("lane") in {"CONTENT_FTS", "TITLE_FTS"}]
    valid_codes = {code.value for code in RootCauseCode}
    classified = bool(material) and all(
        row.get("classification") in valid_codes for row in material
    )
    decisive_lanes = all(
        bool(lane_aggregates.get(lane, {}).get("decisive"))
        and lane_aggregates.get(lane, {}).get("primary") not in {None, "MIXED"}
        for lane in ("CONTENT_FTS", "TITLE_FTS")
    )
    return (
        "PASS_ROOT_CAUSE_PROVEN"
        if classified
        and decisive_lanes
        and bool(complete_fields)
        and all(complete_fields.values())
        else "NO_GO_B1_INCONCLUSIVE"
    )


def parse_expert_workbook(path: Path) -> tuple[DiagnosticCase, ...]:
    """Read Q01..Q10 inputs and oracle identities; keep identity scoring out of retrieval."""
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        questions = [
            str(row[0]).strip()
            for row in book["Kết quả 10 câu"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        ]
        expected: dict[str, tuple[str, ...]] = {}
        for row in book["Chấm điểm"].iter_rows(min_row=4, values_only=True):
            if not row or str(row[0] or "").strip() not in {f"Q{i:02d}" for i in range(1, 11)}:
                continue
            import re

            numbers = []
            for raw in re.findall(
                r"\d+(?:\s*/\s*[\wÀ-ỹĐđ.-]+){1,3}", str(row[6] if len(row) > 6 else "")
            ):
                normalized = normalize_document_number(raw)
                if normalized and normalized not in numbers:
                    numbers.append(normalized)
            expected[str(row[0]).strip()] = tuple(numbers)
        if len(questions) != 10 or len(expected) != 10:
            raise ValueError("expert workbook must contain Q01-Q10")
        return tuple(
            DiagnosticCase(f"Q{i:02d}", question, expected[f"Q{i:02d}"])
            for i, question in enumerate(questions, 1)
        )
    finally:
        book.close()


def combine_fts_lane_rows(fts_result: Any, latency_result: Any) -> tuple[dict[str, object], ...]:
    """Combine standalone FTS search facts with post-score collapse/fusion facts."""
    latency_cases = {case.case_id: case for case in latency_result.cases}
    rows: list[dict[str, object]] = []
    for case in fts_result.cases:
        latency_case = latency_cases.get(case.case_id)
        if latency_case is None:
            continue
        diagnostic = latency_case.diagnostic
        for lane_name, probe, gin_valid in (
            ("CONTENT_FTS", case.content, fts_result.inventory.content_gin_valid),
            ("TITLE_FTS", case.title, fts_result.inventory.title_gin_valid),
        ):
            evidence = LaneEvidence(
                config_matches_simple=fts_result.inventory.config_matches_simple,
                gin_index_valid=gin_valid,
                natural_unfiltered_rows=probe.natural_unfiltered_rows,
                natural_filtered_rows=probe.natural_filtered_rows,
                natural_unfiltered_expected=probe.natural_unfiltered_expected,
                natural_filtered_expected=probe.natural_filtered_expected,
                collapsed_expected=diagnostic.lane_collapsed_expected.get(lane_name, False),
                fused_expected=diagnostic.fused_expected,
                or_expected=probe.or_filtered_expected,
            )
            rows.append(
                {
                    "case_id": case.case_id,
                    "lane": lane_name,
                    "classification": classify_lane(evidence).value,
                    "config_matches_simple": evidence.config_matches_simple,
                    "gin_index_valid": evidence.gin_index_valid,
                    "natural_unfiltered_rows": evidence.natural_unfiltered_rows,
                    "natural_filtered_rows": evidence.natural_filtered_rows,
                    "natural_unfiltered_expected": evidence.natural_unfiltered_expected,
                    "natural_filtered_expected": evidence.natural_filtered_expected,
                    "collapsed_expected": evidence.collapsed_expected,
                    "fused_expected": evidence.fused_expected,
                    "or_expected": evidence.or_expected,
                    "or_filtered_rows": probe.or_filtered_rows,
                    "natural_plan_summary": probe.natural_filtered_plan.safe(),
                    "or_plan_summary": probe.or_filtered_plan.safe(),
                    "actual_index_used": probe.actual_index_used,
                    "or_index_used": probe.or_filtered_index_used,
                    "capability_index_used": probe.index_capability_used,
                }
            )
    return tuple(rows)


def prior_phase_b_8425_breakdown(path: Path) -> dict[str, object]:
    """Extract, rather than hardcode, the prior Phase-B measured p95 fields."""
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"available": False}
    pools = payload.get("pool_measurements", [])
    p95 = next(
        (
            float(row["p95_latency_ms"])
            for row in pools
            if isinstance(row, Mapping) and isinstance(row.get("p95_latency_ms"), (int, float))
        ),
        None,
    )
    set_a = payload.get("set_a", [])
    row = next(
        (
            candidate
            for candidate in set_a
            if isinstance(candidate, Mapping)
            and candidate.get("retrieval_eval_ms") == p95
        ),
        {},
    )
    embedding_ms = _numeric_value(row, "embedding_ms")
    ranking_ms = _numeric_value(row, "ranking_ms")
    data_lane_elapsed_ms = (
        round(max(p95 - embedding_ms - ranking_ms, 0.0), 3)
        if p95 is not None and embedding_ms is not None and ranking_ms is not None
        else None
    )
    return {
        "available": p95 is not None,
        "original_latency_statistic": "COLD_NEAREST_RANK_P95",
        "original_cold_nearest_rank_p95_ms": p95,
        "original_p95_ms": p95,
        "retrieval_eval_formula": "embedding_ms + data_lane_elapsed_ms + ranking_ms",
        "embedding_ms": embedding_ms,
        "data_lane_elapsed_ms": data_lane_elapsed_ms,
        "ranking_ms": ranking_ms,
        "component_sum_ms": (
            round(embedding_ms + data_lane_elapsed_ms + ranking_ms, 3)
            if embedding_ms is not None
            and data_lane_elapsed_ms is not None
            and ranking_ms is not None
            else None
        ),
        "retrieval_eval_explain_included": False,
        "reader_wall_with_explain_ms": _numeric_value(row, "reader_ms"),
        "end_to_end_with_explain_ms": _numeric_value(row, "end_to_end_ms"),
        "reader_wall_explain_included": bool(_numeric_value(row, "explain_query_count")),
        "timings_are_not_all_additive": True,
    }


def semantic_latency_root(latency_result: Any) -> dict[str, object]:
    """Return a bounded semantic conclusion strictly from timing and plan summaries."""
    plans = [plan for case in latency_result.cases for plan in case.plans]
    exact = [plan for plan in plans if plan.label.startswith("EXACT_SEMANTIC")]
    ann = [plan for plan in plans if plan.label.startswith("ANN_CONTROL")]
    exact_seq_scan_count = sum("Seq Scan" in plan.plan.scan_types for plan in exact)
    exact_index_scan_count = sum(
        any("Index" in scan_type for scan_type in plan.plan.scan_types) for plan in exact
    )
    exact_seq_scan = bool(exact) and exact_seq_scan_count == len(exact)
    exact_hnsw_used = any(plan.hnsw_index_used for plan in exact)
    ann_hnsw_used = any(plan.hnsw_index_used for plan in ann)
    exact_limit_evidence = bool(exact) and all(plan.limit_evidence for plan in exact)
    ann_limit_evidence = bool(ann) and all(plan.limit_evidence for plan in ann)
    p95 = latency_result.aggregates
    controlled_stage_names = (
        "embedding_ms",
        "phase4_sql_ms",
        "phase4_collapse_ms",
        "diagnostic_semantic_ms",
        "diagnostic_content_ms",
        "diagnostic_title_ms",
        "diagnostic_collapse_ms",
        "diagnostic_fusion_ms",
        "diagnostic_transaction_other_ms",
        "analyzer_ms",
        "hydration_ms",
    )
    controlled_stages = {
        key: float(p95.get(key, {}).get("p95_ms", 0.0))
        for key in controlled_stage_names
        if isinstance(p95.get(key), Mapping)
    }
    major_stage, major_stage_ms = max(
        controlled_stages.items(), key=lambda item: item[1], default=("NOT_MEASURED", 0.0)
    )
    return {
        "code": (
            "SEMANTIC_EXACT_SCAN_FORCED_SEQSCAN_ANN_CAPABILITY_CONFIRMED"
            if exact_seq_scan and not exact_hnsw_used and ann_hnsw_used
            else "SEMANTIC_PLAN_CAPABILITY_INCONCLUSIVE"
        ),
        "exact_scan_mode": "INDEX_AND_BITMAP_SCANS_DISABLED_BY_DIAGNOSTIC",
        "exact_hnsw_absence_is_planner_failure": False,
        "exact_scans_disabled_seq_scan": exact_seq_scan,
        "exact_scans_disabled_plan_count": len(exact),
        "exact_seq_scan_plan_count": exact_seq_scan_count,
        "exact_index_scan_plan_count": exact_index_scan_count,
        "exact_index_names": sorted(
            {name for plan in exact for name in plan.plan.index_names}
        ),
        "exact_hnsw_actual": exact_hnsw_used,
        "ann_control_plan_count": len(ann),
        "ann_control_hnsw_actual": ann_hnsw_used,
        "ann_control_scope": "BARE_CHUNK_EMBEDDING_CAPABILITY_NOT_PRODUCTION_EQUIVALENT",
        "exact_limit_above_scan": exact_limit_evidence,
        "ann_control_limit_above_scan": ann_limit_evidence,
        "model_and_database_warmed_single_pass_case_count": len(latency_result.cases),
        "descriptive_nearest_rank_p95": True,
        "dominant_controlled_data_stage_p95": major_stage,
        "dominant_controlled_data_stage_p95_ms": major_stage_ms,
        "explain_wall_p95_ms": float(p95.get("explain_wall_ms", {}).get("p95_ms", 0.0)),
        "explain_overhead_p95_ms": float(
            p95.get("explain_overhead_ms", {}).get("p95_ms", 0.0)
        ),
        "description": (
            "Exact seq scans are diagnostic-by-design; the separate ANN control proves only "
            "bare index capability. Timings are descriptive over the fixed ten-case pass."
        ),
    }


def _numeric_value(row: object, key: str) -> float | None:
    value = row.get(key) if isinstance(row, Mapping) else None
    return float(value) if isinstance(value, (int, float)) else None


def write_reports(result: Mapping[str, object], markdown_path: Path, json_path: Path) -> None:
    """Write safe evidence reports and reject raw input, evidence, query, and plan fields."""
    _assert_safe_report(result)
    for path in (markdown_path, json_path):
        path.parent.mkdir(parents=True, exist_ok=True)
    encoded = json.dumps(result, ensure_ascii=False, indent=2)
    json_path.write_text(encoded + "\n", encoding="utf-8")
    aggregates = result.get("fts_aggregates", {})
    inventory = result.get("inventory", {})
    lane_rows = result.get("lane_rows", ())
    latency = result.get("latency", {})
    timings = latency.get("aggregates", {}) if isinstance(latency, Mapping) else {}
    latency_counts = latency.get("counts", {}) if isinstance(latency, Mapping) else {}
    q6 = result.get("q6", {})
    trace = q6.get("trace", []) if isinstance(q6, Mapping) else []
    wrong_final = q6.get("wrong_final_numbers", []) if isinstance(q6, Mapping) else []
    invariants = result.get("invariants", {})
    stage_names = (
        "embedding_ms",
        "phase4_transaction_setup_ms",
        "phase4_sql_ms",
        "phase4_collapse_ms",
        "phase4_total_ms",
        "diagnostic_semantic_ms",
        "diagnostic_content_ms",
        "diagnostic_title_ms",
        "diagnostic_collapse_ms",
        "diagnostic_fusion_ms",
        "diagnostic_transaction_other_ms",
        "diagnostic_total_ms",
        "explain_wall_ms",
        "explain_overhead_ms",
        "analyzer_ms",
        "hydration_ms",
    )
    lines = [
        "# Phase B.1 retrieval-engine root cause",
        "",
        "## Evidence protocol",
        "The embedding model and database candidate paths were warmed once with an internal "
        "constant before a fixed, sequential ten-case pass. Reported p50/p95 values are "
        "descriptive nearest-rank statistics, not repeated-trial confidence estimates.",
        "",
        "## Actual FTS configuration/index inventory",
        json.dumps(inventory, ensure_ascii=False),
        "",
        "## TITLE_FTS root cause",
        json.dumps(aggregates.get("TITLE_FTS", {}), ensure_ascii=False),
        "",
        "## CONTENT_FTS root cause",
        json.dumps(aggregates.get("CONTENT_FTS", {}), ensure_ascii=False),
        "",
        "Natural conjunction misses followed by bounded OR-control recovery are classified as "
        "query-construction recall limitations; this does not assert that OR is the intended "
        "production query semantics.",
        "",
        "## Per-case FTS evidence",
        "| Case | Lane | Classification | Natural rows | Natural expected | Collapsed expected | Fused expected | OR expected | Index used |",
        "|---|---|---|---:|---|---|---|---|---|",
    ]
    if isinstance(lane_rows, (list, tuple)):
        lines.extend(
            "| {case_id} | {lane} | {classification} | {natural_filtered_rows} | "
            "{natural_filtered_expected} | {collapsed_expected} | {fused_expected} | "
            "{or_expected} | {actual_index_used} |".format(**row)
            for row in lane_rows
            if isinstance(row, Mapping)
        )
    lines.extend(
        [
        "",
        "## Original cold nearest-rank p95 breakdown",
        json.dumps(result.get("prior_phase_b_8425_breakdown", {}), ensure_ascii=False),
        "",
        "The frozen retrieval-evaluation p95 excludes EXPLAIN and follows the evaluator formula. "
        "Reader wall/end-to-end timings that include EXPLAIN are separate, non-additive observations.",
        "",
        "## Model-and-database-warmed single-pass p50/p95",
        "| Stage | p50 ms | p95 ms |",
        "|---|---:|---:|",
        ]
    )
    if isinstance(timings, Mapping):
        lines.extend(
            f"| {stage} | {timings.get(stage, {}).get('p50_ms')} | {timings.get(stage, {}).get('p95_ms')} |"
            for stage in stage_names
            if isinstance(timings.get(stage), Mapping)
        )
    lines.extend(
        [
        "",
        "Query-count accounting: " + json.dumps(latency_counts, ensure_ascii=False),
        "",
        "## Semantic plan/capability evidence",
        json.dumps(result.get("semantic_latency_root", {}), ensure_ascii=False),
        "",
        "Exact-path seq scans are forced by disabled index/bitmap scans. The bare ANN control "
        "proves HNSW capability only; it is not production-equivalent planner evidence. Nested "
        "plan-summary row/buffer values are not PostgreSQL plan-total resource usage.",
        "",
        "## Additional hypothetical Q6 trace (fresh EXPLAIN_FALSE run)",
        "This is not a replay of a frozen Phase-B selected configuration; Phase B had NO_SELECTION.",
        "",
        "| Document number | Diagnostic | Pool20 | Final3 | Fusion | Semantic rank | Semantic score | Rejection |",
        "|---|---:|---:|---:|---:|---:|---:|---|",
        ]
    )
    lines.extend(
        "| {document_number} | {diagnostic_rank} | {pool20_rank} | {final_rank} | {fusion_score} | {semantic_rank} | {semantic_score} | {rejection_reason} |".format(
            **row
        )
        for row in trace
    )
    lines.extend(
        [
            "",
            "Wrong final document numbers: " + json.dumps(wrong_final, ensure_ascii=False),
            "",
            "## Read-only/default-off invariants",
            json.dumps(invariants, ensure_ascii=False),
            "",
            "## recommended remediation1-3",
            "1. Separately approve an FTS query-construction experiment only where evidence supports it.",
            "2. Investigate exact semantic SQL, index capability, and warm-cache behavior.",
            "3. Establish an evaluation warmup and latency measurement protocol.",
            "",
            "## Gate",
            f"**{result.get('gate', 'NO_GO_B1_INCONCLUSIVE')}**",
            "",
            "## explicit no tuning",
            "No retrieval, index, SQL, model, or runtime tuning was implemented by this diagnostic.",
            "",
        ]
    )
    markdown_path.write_text("\n".join(lines), encoding="utf-8")


def _assert_safe_report(value: object) -> None:
    raw_value_keys = {
        "question",
        "query",
        "tsquery",
        "lexeme",
        "chunk",
        "title",
        "answer",
        "url",
        "uuid",
        "vector",
        "sql",
        "plan",
        "raw_plan",
    }
    if isinstance(value, Mapping):
        for key, item in value.items():
            if _unsafe_report_key(str(key), raw_value_keys):
                raise UnsafeReportFieldError(str(key))
            _assert_safe_report(item)
    elif isinstance(value, (list, tuple)):
        for item in value:
            _assert_safe_report(item)


def _unsafe_report_key(key: str, raw_value_keys: set[str]) -> bool:
    """Reject raw value fields without rejecting safe derived names such as ``sql_ms``.

    This deliberately uses exact keys (plus an explicit ``raw_*`` form), rather
    than substring matching: summaries, counts, and timing fields are safe
    allowlisted derivatives, not raw query/plan payloads.
    """

    normalized = key.casefold()
    if normalized in {code.value.casefold() for code in RootCauseCode}:
        return False
    if normalized in raw_value_keys:
        return True
    if normalized.startswith("raw_"):
        return True
    safe_exact = {
        "plans",
        "plan_summary",
        "token_alias_distribution",
        "expected_numbers",
        "wrong_final_numbers",
        "document_number",
        "content_fts",
        "title_fts",
        "semantic_plan_capability_complete",
    }
    safe_suffixes = (
        "_count",
        "_counts",
        "_ms",
        "_rank",
        "_score",
        "_rows",
        "_names",
        "_types",
        "_summary",
        "_available",
        "_enabled",
        "_used",
        "_valid",
        "_ready",
        "_evidence",
        "_distribution",
    )
    if normalized in safe_exact or normalized.endswith(safe_suffixes):
        return False
    return any(part in raw_value_keys for part in normalized.split("_"))
