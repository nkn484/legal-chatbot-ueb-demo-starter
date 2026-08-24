"""Deterministic reader for the three approved demo-data worksheets."""

from __future__ import annotations

import json
import re
from datetime import UTC, date, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

import openpyxl

from legal_chatbot.demo_corpus.models import CatalogEntry, CorpusFileKind

_SHEETS = {
    "RAWDATA_QPPL": "VBQPPL",
    "RAWDATA_VNU": "VNU",
    "RAWDATA_UEB": "UEB",
}


def _text(value: Any) -> str | None:
    if value is None:
        return None
    normalized = re.sub(r"\s+", " ", str(value)).strip()
    return normalized or None


def _date(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=UTC) if value.tzinfo is None else value.astimezone(UTC)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=UTC)
    text = str(value).strip()
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    return parsed.replace(tzinfo=UTC) if parsed.tzinfo is None else parsed.astimezone(UTC)


def classify_file_reference(label: str | None, url: str | None) -> CorpusFileKind:
    if url:
        path = urlsplit(url).path
        if "/folders/" in path:
            return CorpusFileKind.FOLDER
        if "/file/d/" in path or "/document/d/" in path:
            return CorpusFileKind.DIRECT_FILE
        return CorpusFileKind.UNRESOLVED
    return CorpusFileKind.UNRESOLVED if label else CorpusFileKind.MISSING


def direct_download_url(url: str) -> str | None:
    match = re.search(r"/document/d/([^/?]+)", url)
    if match:
        return f"https://docs.google.com/document/d/{match.group(1)}/export?format=pdf"
    match = re.search(r"/file/d/([^/?]+)", url)
    if match:
        return f"https://drive.usercontent.google.com/download?id={match.group(1)}&export=download"
    return None


def _entry(path: Path, sheet: Any, row: int, dataset_id: str) -> CatalogEntry:
    source_id = _SHEETS[sheet.title]
    file_cell = sheet.cell(row, 13)
    file_label = _text(file_cell.value)
    file_url = file_cell.hyperlink.target if file_cell.hyperlink else None
    authority = (
        _text(sheet.cell(row, 9).value)
        if source_id == "VBQPPL"
        else (
            "Đại học Quốc gia Hà Nội" if source_id == "VNU" else "Trường Đại học Kinh tế - ĐHQGHN"
        )
    )
    payload = {
        "dataset_id": dataset_id,
        "source_id": source_id,
        "workbook_name": path.name,
        "sheet_name": sheet.title,
        "source_row": row,
        "document_number": _text(sheet.cell(row, 4).value),
        "title": _text(sheet.cell(row, 8).value),
        "document_type": _text(sheet.cell(row, 6).value),
        "issuing_authority": authority,
        "issue_date": _date(sheet.cell(row, 5).value),
        "effective_date": _date(sheet.cell(row, 10).value) if source_id == "VBQPPL" else None,
        "legal_status": _text(sheet.cell(row, 11).value),
        "file_label": file_label,
        "file_url": file_url,
        "file_kind": classify_file_reference(file_label, file_url),
    }
    canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    return CatalogEntry(
        **payload,
        external_id=f"{dataset_id}:{sheet.title}:{row}",
        record_sha256=sha256(canonical).hexdigest(),
    )


def load_demo_catalog(
    demo_data_path: Path = Path("demo_data"), *, dataset_id: str = "demo-data-v1"
) -> tuple[CatalogEntry, ...]:
    paths = (
        demo_data_path / "VBQPPL - So sánh.xlsx",
        demo_data_path / "Số hóa VB VNU và UEB.xlsx",
    )
    entries: list[CatalogEntry] = []
    for path in paths:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
        try:
            for sheet_name in _SHEETS:
                if sheet_name not in workbook.sheetnames:
                    continue
                sheet = workbook[sheet_name]
                for row in range(2, sheet.max_row + 1):
                    if sheet.cell(row, 1).value is not None:
                        entries.append(_entry(path, sheet, row, dataset_id))
        finally:
            workbook.close()
    return tuple(entries)
