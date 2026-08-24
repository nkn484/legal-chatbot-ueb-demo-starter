"""Export safe metadata for every indexed manual snapshot corpus document."""

from __future__ import annotations

import argparse
import asyncio
from collections import Counter
from collections.abc import Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from legal_chatbot.core.config import Settings
from legal_chatbot.db.session import create_engine, create_session_factory
from legal_chatbot.demo_corpus.config import DemoCorpusSettings
from legal_chatbot.documents.orm import CorpusCatalogEntry

_HEADERS = (
    "STT",
    "Nguồn",
    "Workbook",
    "Sheet",
    "Dòng nguồn",
    "Số văn bản",
    "Tên/Trích yếu",
    "Loại văn bản",
    "Cơ quan/Đơn vị",
    "Ngày ban hành",
    "Ngày hiệu lực",
    "Tình trạng",
    "File văn bản",
    "URL file",
    "SHA-256 file",
    "Legal document ID",
    "Document version ID",
    "Cập nhật trạng thái lúc",
)


def _value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return value
    return value


def _style_sheet(sheet: Any) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = (8, 12, 32, 20, 12, 28, 70, 22, 32, 16, 16, 20, 28, 65, 68, 40, 40, 22)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


async def run(output_path: Path) -> int:
    settings = Settings()  # type: ignore[call-arg]
    corpus_settings = DemoCorpusSettings()
    engine = create_engine(settings)
    session_factory = create_session_factory(engine)
    try:
        async with session_factory() as session:
            entries = tuple(
                (
                    await session.scalars(
                        select(CorpusCatalogEntry)
                        .where(
                            CorpusCatalogEntry.dataset_id == corpus_settings.dataset_id,
                            CorpusCatalogEntry.processing_status == "INDEXED",
                        )
                        .order_by(
                            CorpusCatalogEntry.source_id,
                            CorpusCatalogEntry.sheet_name,
                            CorpusCatalogEntry.source_row,
                        )
                    )
                ).all()
            )
    finally:
        await engine.dispose()

    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is not None:
        workbook.remove(active_sheet)
    counts = Counter(entry.source_id for entry in entries)

    summary = workbook.create_sheet("Tổng hợp")
    summary.append(("Nguồn", "Số văn bản đã index"))
    for source_id in ("VBQPPL", "VNU", "UEB"):
        summary.append((source_id, counts[source_id]))
    summary.append(("TỔNG", len(entries)))
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 24
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    for source_id in ("VBQPPL", "VNU", "UEB"):
        sheet = workbook.create_sheet(source_id)
        sheet.append(_HEADERS)
        source_entries = tuple(entry for entry in entries if entry.source_id == source_id)
        for index, entry in enumerate(source_entries, start=1):
            sheet.append(
                tuple(
                    _value(value)
                    for value in (
                        index,
                        entry.source_id,
                        entry.workbook_name,
                        entry.sheet_name,
                        entry.source_row,
                        entry.document_number,
                        entry.title,
                        entry.document_type,
                        entry.issuing_authority,
                        entry.issue_date,
                        entry.effective_date,
                        entry.legal_status,
                        entry.file_label,
                        entry.file_url,
                        entry.file_sha256,
                        str(entry.legal_document_id) if entry.legal_document_id else None,
                        str(entry.document_version_id) if entry.document_version_id else None,
                        entry.updated_at,
                    )
                )
            )
            url_cell = sheet.cell(sheet.max_row, 14)
            if entry.file_url:
                url_cell.hyperlink = entry.file_url
                url_cell.style = "Hyperlink"
        _style_sheet(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(
        f"indexed_export={output_path.as_posix()} total={len(entries)} "
        f"VBQPPL={counts['VBQPPL']} VNU={counts['VNU']} UEB={counts['UEB']}"
    )
    return 0


def main(argv: Sequence[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Export indexed demo corpus metadata to Excel")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("docs/evidence/demo-corpus-indexed-documents.xlsx"),
    )
    args = parser.parse_args(argv)
    raise SystemExit(asyncio.run(run(args.output)))


if __name__ == "__main__":
    main()
