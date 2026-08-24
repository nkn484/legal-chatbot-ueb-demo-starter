"""Validate the non-authoritative M2 evaluation controls without runtime access.

This tool is evaluation-only.  It reads the controlled workbook solely to extract
benchmark document-number tokens for leakage linting; it never emits workbook
questions, answers, or review comments.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import unicodedata
from collections import Counter
from collections.abc import Iterable
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_JSON = ROOT / "docs" / "evals" / "m2_evaluation_set.json"
DEFAULT_XLSX = ROOT / "docs" / "evals" / "m2_evaluation_set.xlsx"
DEFAULT_BENCHMARK_WORKBOOK = ROOT / "docs" / "Stress_test_Legal_Chatbot_UEB_10_cau.xlsx"
SCHEMA_VERSION = "M2-EVALUATION-SET-1"
CASE_FIELDS = frozenset(
    {"case_id", "parent_case_id", "category", "question", "expected_behavior", "notes"}
)
CATEGORIES = frozenset(
    {
        "SINGLE_SOURCE_SUFFICIENCY",
        "NO_EVIDENCE",
        "UNRELATED_TO_UEB",
        "UEB_MENTION_NO_ALL_SOURCES",
        "AMBIGUOUS_DOCUMENT_IDENTITY",
        "GENERAL_ADMINISTRATIVE_NONLEGAL",
        "SYNTHETIC_METADATA_NUMBER",
    }
)
CONTROL_BEHAVIORS = frozenset(
    {"MAY_CLARIFY_OR_REFUSE", "MAY_RESPOND_WITH_AVAILABLE_EVIDENCE"}
)
URL_PATTERN = re.compile(r"(?:https?://|www\.)", re.IGNORECASE)
UUID_PATTERN = re.compile(
    r"\b[0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}\b",
    re.IGNORECASE,
)
DOCUMENT_NUMBER_PATTERN = re.compile(
    r"\b\d{1,8}(?:\s*/\s*[A-Za-zÀ-ỹĐđ][\wÀ-ỹĐđ.-]*){1,3}\b", re.UNICODE
)


class ValidationError(ValueError):
    """Raised when an M2 evaluation artifact is not a valid control set."""


def canonical_json_bytes(payload: Any) -> bytes:
    """Return the documented UTF-8 canonical JSON representation."""

    return json.dumps(
        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")


def canonical_json_sha256(path: Path) -> str:
    """Hash parsed JSON, rather than incidental whitespace in its source file."""

    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    return hashlib.sha256(canonical_json_bytes(payload)).hexdigest()


def load_evaluation_set(path: Path) -> dict[str, Any]:
    """Load a UTF-8 evaluation manifest."""

    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise ValidationError("evaluation manifest must be a JSON object")
    return payload


def _normalise(value: str) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", value).strip()).casefold()


def _has_control_character(value: str) -> bool:
    return any(unicodedata.category(character) == "Cc" for character in value)


def _require_safe_string(field: str, value: object) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValidationError(f"{field} must be a non-empty string")
    if _has_control_character(value) or URL_PATTERN.search(value) or UUID_PATTERN.search(value):
        raise ValidationError(f"{field} contains disallowed control content")
    return value


def cases_by_set(payload: dict[str, Any]) -> dict[str, tuple[dict[str, str], ...]]:
    """Return typed cases after structural validation, without changing their order."""

    if payload.get("schema_version") != SCHEMA_VERSION:
        raise ValidationError("unexpected schema_version")
    sets = payload.get("sets")
    if not isinstance(sets, dict) or set(sets) != {"B", "C"}:
        raise ValidationError("manifest must contain only sets B and C")

    result: dict[str, tuple[dict[str, str], ...]] = {}
    for set_name in ("B", "C"):
        definition = sets[set_name]
        if not isinstance(definition, dict) or set(definition) != {"description", "cases"}:
            raise ValidationError(f"set {set_name} has an invalid shape")
        _require_safe_string(f"set {set_name} description", definition["description"])
        raw_cases = definition["cases"]
        if not isinstance(raw_cases, list):
            raise ValidationError(f"set {set_name} cases must be a list")
        typed_cases: list[dict[str, str]] = []
        for index, case in enumerate(raw_cases, start=1):
            if not isinstance(case, dict) or set(case) != CASE_FIELDS:
                raise ValidationError(f"set {set_name} case {index} has an invalid shape")
            typed: dict[str, str] = {}
            for field in CASE_FIELDS:
                value = case[field]
                if field == "parent_case_id" and value is None:
                    typed[field] = ""
                else:
                    typed[field] = _require_safe_string(
                        f"set {set_name} case {index} {field}", value
                    )
            question_length = len(typed["question"].strip())
            if not 12 <= question_length <= 500:
                raise ValidationError(
                    f"set {set_name} case {index} question is outside length bounds"
                )
            typed_cases.append(typed)
        result[set_name] = tuple(typed_cases)
    return result


def load_benchmark_document_numbers(path: Path) -> frozenset[str]:
    """Extract only benchmark document-number tokens from the grading column."""

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Khung chấm" not in workbook.sheetnames:
            raise ValidationError("benchmark workbook lacks the grading sheet")
        numbers: set[str] = set()
        for row in workbook["Khung chấm"].iter_rows(min_row=5, values_only=True):
            if len(row) < 4 or row[3] is None:
                continue
            for match in DOCUMENT_NUMBER_PATTERN.findall(str(row[3])):
                numbers.add(re.sub(r"\s+", "", match).casefold())
        if not numbers:
            raise ValidationError("benchmark workbook has no parseable document numbers")
        return frozenset(numbers)
    finally:
        workbook.close()


def find_document_number_leaks(
    cases: Iterable[dict[str, str]], benchmark_numbers: Iterable[str]
) -> tuple[str, ...]:
    """Return case IDs whose text contains an exact normalized benchmark number."""

    normalized_numbers = tuple(
        sorted({_normalise(number).replace(" ", "") for number in benchmark_numbers})
    )
    leaked: list[str] = []
    for case in cases:
        search_text = _normalise(" ".join(case.values())).replace(" ", "")
        if any(number in search_text for number in normalized_numbers):
            leaked.append(case["case_id"])
    return tuple(leaked)


def validate_cases(
    sets: dict[str, tuple[dict[str, str], ...]], benchmark_numbers: Iterable[str]
) -> None:
    """Enforce M2 counts, labels, uniqueness, and benchmark-leakage boundaries."""

    set_b = sets["B"]
    set_c = sets["C"]
    if len(set_b) != 30:
        raise ValidationError("set B must contain exactly 30 paraphrases")
    if len(set_c) < 20:
        raise ValidationError("set C must contain at least 20 controls")

    ids = [case["case_id"] for cases in sets.values() for case in cases]
    if len(ids) != len(set(ids)):
        raise ValidationError("case IDs must be unique")
    questions = [_normalise(case["question"]) for cases in sets.values() for case in cases]
    if len(questions) != len(set(questions)):
        raise ValidationError("questions must be unique after normalization")

    parents = Counter(case["parent_case_id"] for case in set_b)
    expected_parents = {f"Q{number:02d}" for number in range(1, 11)}
    if set(parents) != expected_parents or any(count != 3 for count in parents.values()):
        raise ValidationError("set B must contain three paraphrases for each Q01 through Q10")
    for case in set_b:
        parent = case["parent_case_id"]
        if not re.fullmatch(rf"B-{parent}-0[1-3]", case["case_id"]):
            raise ValidationError("set B case ID does not match its parent")
        if case["category"] != "PARAPHRASE":
            raise ValidationError("set B category must be PARAPHRASE")
        if case["expected_behavior"] != "RETRIEVAL_COMPARABLE_TO_PARENT":
            raise ValidationError("set B expected_behavior is invalid")

    categories = Counter(case["category"] for case in set_c)
    if set(categories) != CATEGORIES:
        raise ValidationError("set C must cover exactly the M2 control categories")
    for case in set_c:
        if case["parent_case_id"]:
            raise ValidationError("set C parent_case_id must be null")
        if not re.fullmatch(r"C-[A-Z0-9-]+-\d{2}", case["case_id"]):
            raise ValidationError("set C case ID is invalid")
        if case["expected_behavior"] not in CONTROL_BEHAVIORS:
            raise ValidationError("set C expected_behavior is invalid")

    all_cases = (case for cases in sets.values() for case in cases)
    leaked = find_document_number_leaks(all_cases, benchmark_numbers)
    if leaked:
        raise ValidationError("benchmark document-number leakage detected")


def _xlsx_rows(path: Path) -> tuple[tuple[str, ...], ...]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Evaluation cases" not in workbook.sheetnames:
            raise ValidationError("XLSX lacks the Evaluation cases sheet")
        sheet = workbook["Evaluation cases"]
        header_row = next(sheet.iter_rows(values_only=True))
        header = tuple("" if value is None else str(value) for value in header_row)
        expected_header = (
            "Set",
            "Case ID",
            "Parent Case ID",
            "Category",
            "Question",
            "Expected Behavior",
            "Notes",
        )
        if header != expected_header:
            raise ValidationError("XLSX evaluation header is invalid")
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(value is None for value in row):
                continue
            rows.append(tuple("" if value is None else str(value) for value in row[:7]))
        return tuple(rows)
    finally:
        workbook.close()


def validate_xlsx_parity(path: Path, sets: dict[str, tuple[dict[str, str], ...]]) -> None:
    """Check that review XLSX cells equal JSON cases, in the same stable order."""

    expected = tuple(
        (
            set_name,
            case["case_id"],
            case["parent_case_id"],
            case["category"],
            case["question"],
            case["expected_behavior"],
            case["notes"],
        )
        for set_name in ("B", "C")
        for case in sets[set_name]
    )
    if _xlsx_rows(path) != expected:
        raise ValidationError("XLSX and JSON evaluation cases are not identical")


def validate_evaluation_set(json_path: Path, xlsx_path: Path, benchmark_workbook: Path) -> None:
    """Run all artifact checks with no network, model, database, or runtime calls."""

    sets = cases_by_set(load_evaluation_set(json_path))
    validate_cases(sets, load_benchmark_document_numbers(benchmark_workbook))
    validate_xlsx_parity(xlsx_path, sets)


def write_xlsx(path: Path, payload: dict[str, Any]) -> None:
    """Create the reviewer workbook directly from the machine-readable manifest."""

    sets = cases_by_set(payload)
    workbook = openpyxl.Workbook()
    manifest = workbook.active
    if manifest is None:
        raise RuntimeError("openpyxl did not create a manifest worksheet")
    manifest.title = "Manifest"
    manifest.append(("Field", "Value"))
    manifest.append(("Schema version", payload["schema_version"]))
    manifest.append(
        ("Canonical JSON SHA-256", hashlib.sha256(canonical_json_bytes(payload)).hexdigest())
    )
    manifest.append(("Set B cases", len(sets["B"])))
    manifest.append(("Set C cases", len(sets["C"])))
    manifest.append(("Use", "Non-authoritative evaluation control; human review required."))

    sheet = workbook.create_sheet("Evaluation cases")
    header = (
        "Set",
        "Case ID",
        "Parent Case ID",
        "Category",
        "Question",
        "Expected Behavior",
        "Notes",
    )
    sheet.append(header)
    for set_name in ("B", "C"):
        for case in sets[set_name]:
            sheet.append(
                (
                    set_name,
                    case["case_id"],
                    case["parent_case_id"],
                    case["category"],
                    case["question"],
                    case["expected_behavior"],
                    case["notes"],
                )
            )
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for target in (manifest, sheet):
        for cell in target[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        target.freeze_panes = "A2"
        target.auto_filter.ref = target.dimensions
    for column, width in {"A": 11, "B": 18, "C": 18, "D": 34, "E": 68, "F": 38, "G": 34}.items():
        sheet.column_dimensions[column].width = width
    sheet.sheet_view.showGridLines = False
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Validate M2 evaluation control artifacts.")
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--benchmark-workbook", type=Path, default=DEFAULT_BENCHMARK_WORKBOOK)
    parser.add_argument("--write-xlsx", action="store_true")
    args = parser.parse_args(argv)
    payload = load_evaluation_set(args.json)
    sets = cases_by_set(payload)
    benchmark_numbers = load_benchmark_document_numbers(args.benchmark_workbook)
    validate_cases(sets, benchmark_numbers)
    if args.write_xlsx:
        write_xlsx(args.xlsx, payload)
    validate_xlsx_parity(args.xlsx, sets)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ValidationError as error:
        print(f"validation failed: {error}", file=sys.stderr)
        raise SystemExit(1) from error
