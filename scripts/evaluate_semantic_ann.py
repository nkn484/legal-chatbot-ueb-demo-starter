"""Read-only Gate-3 comparison of exact cosine ranking and PostgreSQL HNSW ANN.

This is an operator tool, not a runtime retrieval path.  It never creates
retrieval runs/citations and never writes query text or embeddings to the report.
"""

from __future__ import annotations

import argparse
import asyncio
import re
import sys
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from pathlib import Path
from time import perf_counter
from typing import Any
from uuid import UUID

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import and_, func, or_, select, text

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "src") not in sys.path:
    sys.path.insert(0, str(ROOT / "src"))

from legal_chatbot.core.config import Settings  # noqa: E402
from legal_chatbot.db.session import create_engine, create_session_factory  # noqa: E402
from legal_chatbot.documents.hybrid_retrieval_repository import (  # noqa: E402
    PostgresHybridRetrievalRepository,
)
from legal_chatbot.documents.orm import (  # noqa: E402
    ChunkEmbedding,
    DocumentChunk,
    DocumentVersion,
    LegalDocument,
    SourceProvenanceRecord,
)
from legal_chatbot.semantic.config import SemanticSettings  # noqa: E402
from legal_chatbot.semantic.constants import SEMANTIC_DIMENSION, SEMANTIC_PROFILE_ID  # noqa: E402
from legal_chatbot.semantic.fastembed_adapter import FastEmbedSemanticAdapter  # noqa: E402

DEFAULT_INPUT = Path("docs/Stress_test_Legal_Chatbot_UEB_10_cau.xlsx")
DEFAULT_OUTPUT = Path("docs/Stress_test_Legal_Chatbot_UEB_10_cau_G3_ANN.xlsx")
SOURCES = ("VBQPPL", "VNU", "UEB")
CASE_IDS = tuple(f"Q{number:02d}" for number in range(1, 11))


@dataclass(frozen=True)
class EvaluationCase:
    """Question with grading metadata kept separate until post-query scoring."""

    case_id: str
    question: str
    expected_documents: tuple[str, ...]


@dataclass(frozen=True)
class Candidate:
    """Internal identifier/metadata-only candidate; text and vectors are excluded."""

    chunk_id: UUID
    source_id: str
    document_number: str | None


@dataclass(frozen=True)
class CaseResult:
    case_id: str
    exact: tuple[Candidate, ...]
    ann: tuple[Candidate, ...]
    exact_ms: float
    ann_ms: float
    ann_index_used: bool
    expected_documents: tuple[str, ...]


@dataclass(frozen=True)
class EvaluationResult:
    cases: tuple[CaseResult, ...]
    blocker: str | None = None


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _expected_documents(value: object) -> tuple[str, ...]:
    """Extract grading-only document labels without feeding them to the embedder."""

    matches = re.findall(r"\d+(?:\s*/\s*[\wÀ-ỹĐđ.-]+){1,3}", _text(value), re.UNICODE)
    return tuple(dict.fromkeys(re.sub(r"\s+", "", item) for item in matches))


def parse_cases(path: Path) -> tuple[EvaluationCase, ...]:
    """Read Q01--Q10 questions only; scoring fields remain independent metadata."""

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        questions_sheet = workbook["Câu hỏi"]
        grading_sheet = workbook["Khung chấm"]
        questions = {
            _text(row[0]): _text(row[1])
            for row in questions_sheet.iter_rows(min_row=5, values_only=True)
            if len(row) >= 2 and _text(row[0]) in CASE_IDS and _text(row[1])
        }
        expected = {
            _text(row[0]): _expected_documents((tuple(row) + (None,) * 4)[3])
            for row in grading_sheet.iter_rows(min_row=5, values_only=True)
            if row and _text(row[0]) in CASE_IDS
        }
        if tuple(sorted(questions)) != CASE_IDS or tuple(sorted(expected)) != CASE_IDS:
            raise ValueError("workbook must contain Q01 through Q10 in both required sheets")
        return tuple(
            EvaluationCase(case_id, questions[case_id], expected[case_id])
            for case_id in CASE_IDS
        )
    finally:
        workbook.close()


def overlap_percent(exact_ids: Iterable[UUID], ann_ids: Iterable[UUID]) -> float:
    exact = set(exact_ids)
    ann = set(ann_ids)
    return 100.0 if not exact else 100.0 * len(exact & ann) / len(exact | ann)


def recall_at_k(exact_ids: Iterable[UUID], ann_ids: Iterable[UUID]) -> float:
    exact = set(exact_ids)
    return 1.0 if not exact else len(exact & set(ann_ids)) / len(exact)


def _plan_contains_index(value: object, index_name: str) -> bool:
    """Recursively inspect PostgreSQL JSON EXPLAIN output without retaining its SQL."""

    if isinstance(value, dict):
        if value.get("Index Name") == index_name:
            return True
        return any(_plan_contains_index(item, index_name) for item in value.values())
    if isinstance(value, list):
        return any(_plan_contains_index(item, index_name) for item in value)
    return False


def _latest_version():
    return (
        select(func.max(DocumentVersion.version_number))
        .where(DocumentVersion.document_id == LegalDocument.id)
        .correlate(LegalDocument)
        .scalar_subquery()
    )


def _strict_provenance():
    return (
        select(SourceProvenanceRecord.id)
        .where(
            SourceProvenanceRecord.document_version_id == DocumentVersion.id,
            SourceProvenanceRecord.transport_trust_mode == "STRICT_TLS",
            or_(
                SourceProvenanceRecord.provenance_type == "source_fetch",
                and_(
                    SourceProvenanceRecord.provenance_type == "manual_snapshot",
                    SourceProvenanceRecord.transport_trust_mode == "STRICT_TLS",
                ),
            ),
        )
        .correlate(DocumentVersion)
        .exists()
    )


def semantic_statement(vector: tuple[float, ...], top_k: int):
    """Build the shared exact/ANN candidate statement with only trusted current vectors."""

    distance = ChunkEmbedding.embedding.cosine_distance(list(vector))
    return (
        select(DocumentChunk.id, LegalDocument.source_id, DocumentVersion.document_number)
        .select_from(DocumentChunk)
        .join(DocumentVersion, DocumentChunk.document_version_id == DocumentVersion.id)
        .join(LegalDocument, DocumentVersion.document_id == LegalDocument.id)
        .join(
            ChunkEmbedding,
            and_(
                ChunkEmbedding.document_chunk_id == DocumentChunk.id,
                ChunkEmbedding.embedding_model_id == SEMANTIC_PROFILE_ID,
                ChunkEmbedding.embedding_kind == "semantic",
                ChunkEmbedding.dimension == SEMANTIC_DIMENSION,
            ),
        )
        .where(
            LegalDocument.source_id.in_(SOURCES),
            DocumentVersion.version_number == _latest_version(),
            _strict_provenance(),
        )
        .order_by(distance.asc(), DocumentChunk.id.asc())
        .limit(top_k)
    )


async def _candidates(session: Any, statement: object) -> tuple[Candidate, ...]:
    rows = (await session.execute(statement)).all()
    return tuple(Candidate(row[0], row[1], row[2]) for row in rows)


async def _exact_candidates(
    session_factory: Any, vector: tuple[float, ...], top_k: int
) -> tuple[Candidate, ...]:
    async with session_factory() as session:
        async with session.begin():
            await session.execute(text("SET LOCAL enable_indexscan = off"))
            await session.execute(text("SET LOCAL enable_bitmapscan = off"))
            return await _candidates(session, semantic_statement(vector, top_k))


async def _ann_candidates(
    session_factory: Any, vector: tuple[float, ...], top_k: int, ef_search: int
) -> tuple[Candidate, ...]:
    async with session_factory() as session:
        async with session.begin():
            await session.execute(text(f"SET LOCAL hnsw.ef_search = {ef_search}"))
            return await _candidates(session, semantic_statement(vector, top_k))


async def _ann_plan_uses_hnsw(session_factory: Any, vector: tuple[float, ...], top_k: int) -> bool:
    """Ask PostgreSQL for an ANN plan using a transient vector literal, never report it."""

    vector_literal = "[" + ",".join(str(item) for item in vector) + "]"
    statement = text(
        "EXPLAIN (FORMAT JSON) SELECT id FROM chunk_embeddings "
        "WHERE embedding_model_id = :profile AND embedding_kind = 'semantic' "
        "AND dimension = 384 ORDER BY embedding <=> CAST(:vector AS vector) LIMIT :limit"
    )
    async with session_factory() as session:
        async with session.begin():
            rows = await session.execute(
                statement,
                {"profile": SEMANTIC_PROFILE_ID, "vector": vector_literal, "limit": top_k},
            )
            plan = rows.scalar_one()
    return _plan_contains_index(plan, "ix_chunk_embeddings_embedding_hnsw_cosine")


async def evaluate(
    cases: tuple[EvaluationCase, ...],
    *,
    session_factory: Any,
    embedder: Any,
    top_k: int,
    ef_search: int,
    coverage_checker: Any = PostgresHybridRetrievalRepository.coverage_complete_for,
) -> EvaluationResult:
    """Sequentially embed each question once and issue only read-only candidate queries."""

    if not await coverage_checker(session_factory, SOURCES):
        return EvaluationResult((), blocker="SEMANTIC_COVERAGE_INCOMPLETE")
    results: list[CaseResult] = []
    for case in cases:
        batch = await embedder.embed_query(case.question)
        vector = batch.vectors[0]
        exact_started = perf_counter()
        exact = await _exact_candidates(session_factory, vector, top_k)
        exact_ms = (perf_counter() - exact_started) * 1000
        ann_started = perf_counter()
        ann = await _ann_candidates(session_factory, vector, top_k, ef_search)
        ann_ms = (perf_counter() - ann_started) * 1000
        index_used = await _ann_plan_uses_hnsw(session_factory, vector, top_k)
        results.append(
            CaseResult(
                case.case_id,
                exact,
                ann,
                exact_ms,
                ann_ms,
                index_used,
                case.expected_documents,
            )
        )
    return EvaluationResult(tuple(results))


def _percentile(values: Iterable[float], percent: int) -> float:
    ordered = sorted(values)
    if not ordered:
        return 0.0
    index = max(0, min(len(ordered) - 1, (len(ordered) * percent + 99) // 100 - 1))
    return ordered[index]


def _hits(candidates: tuple[Candidate, ...], expected: tuple[str, ...]) -> int:
    document_numbers = {item.document_number for item in candidates if item.document_number}
    return len(set(expected) & document_numbers)


def _autosize(sheet: Any) -> None:
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(
            42, max(12, max(len(str(cell.value or "")) for cell in column) + 2)
        )


def _table(sheet: Any, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    sheet.append(list(headers))
    for row in rows:
        sheet.append(list(row))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    _autosize(sheet)


def write_report(output: Path, result: EvaluationResult, *, top_k: int, ef_search: int) -> None:
    """Write only score diagnostics, source/document metadata, and a coverage blocker."""

    workbook = openpyxl.Workbook()
    summary = workbook.active
    assert summary is not None
    summary.title = "Summary"
    recalls = [
        recall_at_k((x.chunk_id for x in row.exact), (x.chunk_id for x in row.ann))
        for row in result.cases
    ]
    exact_times = [row.exact_ms for row in result.cases]
    ann_times = [row.ann_ms for row in result.cases]
    parity = sum(
        _hits(row.exact, row.expected_documents) == _hits(row.ann, row.expected_documents)
        for row in result.cases
    )
    exact_percentiles = f"{_percentile(exact_times, 50):.2f} / {_percentile(exact_times, 95):.2f}"
    ann_percentiles = f"{_percentile(ann_times, 50):.2f} / {_percentile(ann_times, 95):.2f}"
    _table(
        summary,
        ("Metric", "Value"),
        (
            ("Reference", "Exact cosine is the Gate-3 reference; HNSW remains runtime-disabled."),
            ("Profile", SEMANTIC_PROFILE_ID),
            ("Top K", top_k),
            ("HNSW ef_search", ef_search),
            ("Blocker", result.blocker or "NONE"),
            ("Exact p50/p95 ms", exact_percentiles),
            ("ANN p50/p95 ms", ann_percentiles),
            ("Mean ANN recall@K", sum(recalls) / len(recalls) if recalls else 0.0),
            ("Minimum ANN recall@K", min(recalls) if recalls else 0.0),
            ("Expected-hit parity cases", f"{parity}/{len(result.cases)}"),
            ("ANN HNSW index-used cases", sum(row.ann_index_used for row in result.cases)),
        ),
    )
    cases_sheet = workbook.create_sheet("Cases")
    _table(
        cases_sheet,
        (
            "Case ID", "Exact latency ms", "ANN latency ms", "Top-K overlap %", "ANN recall@K",
            "Exact source IDs", "ANN source IDs", "Exact document numbers", "ANN document numbers",
            "Exact expected indexed hits", "ANN expected indexed hits", "ANN HNSW index used",
        ),
        (
            (
                row.case_id,
                round(row.exact_ms, 3),
                round(row.ann_ms, 3),
                round(
                    overlap_percent(
                        (x.chunk_id for x in row.exact), (x.chunk_id for x in row.ann)
                    ),
                    2,
                ),
                round(
                    recall_at_k((x.chunk_id for x in row.exact), (x.chunk_id for x in row.ann)), 4
                ),
                "; ".join(sorted({x.source_id for x in row.exact})),
                "; ".join(sorted({x.source_id for x in row.ann})),
                "; ".join(x.document_number or "" for x in row.exact),
                "; ".join(x.document_number or "" for x in row.ann),
                _hits(row.exact, row.expected_documents),
                _hits(row.ann, row.expected_documents),
                row.ann_index_used,
            )
            for row in result.cases
        ),
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()


async def run(args: argparse.Namespace) -> int:
    cases = parse_cases(args.input)
    engine: Any = None
    result = EvaluationResult((), blocker="INFRASTRUCTURE_FAILURE")
    try:
        engine = create_engine(Settings())
        result = await evaluate(
            cases,
            session_factory=create_session_factory(engine),
            embedder=FastEmbedSemanticAdapter(SemanticSettings()),
            top_k=args.top_k,
            ef_search=args.ef_search,
        )
    except Exception:
        result = EvaluationResult((), blocker="INFRASTRUCTURE_FAILURE")
    finally:
        if engine is not None:
            await engine.dispose()
        write_report(args.output, result, top_k=args.top_k, ef_search=args.ef_search)
    return 0 if result.blocker is None else 2


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Read-only exact versus HNSW semantic evaluator")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--top-k", type=int, default=8)
    parser.add_argument("--ef-search", type=int, default=40)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    if not 1 <= args.top_k <= 8 or not 1 <= args.ef_search <= 1_000:
        return 2
    return asyncio.run(run(args))


if __name__ == "__main__":
    raise SystemExit(main())
