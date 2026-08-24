"""Evaluation-only M2 retrieval comparison for S0 and S1.

Questions are used only as bounded in-memory retrieval inputs.  This script never
calls chat, providers, or network services, and its reports deliberately exclude
questions, answers, titles, URLs, chunk text, UUIDs, prompts, and model payloads.
"""
# ruff: noqa: E501

from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
from collections.abc import Awaitable, Callable, Iterable
from dataclasses import asdict, dataclass, field
from pathlib import Path
from time import perf_counter
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import and_, delete, func, select

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "src") not in sys.path:
    sys.path.insert(0, str(ROOT / "src"))

from legal_chatbot.core.config import Settings  # noqa: E402
from legal_chatbot.db.session import create_engine, create_session_factory  # noqa: E402
from legal_chatbot.documents.citation_resolver import PostgresCitationResolver  # noqa: E402
from legal_chatbot.documents.metadata_normalization import normalize_document_number  # noqa: E402
from legal_chatbot.documents.metadata_repair_repository import (  # noqa: E402
    PostgresMetadataRepairRetrievalRepository,
)
from legal_chatbot.documents.orm import (  # noqa: E402
    ChunkEmbedding,
    CitationRecord,
    CorpusCatalogEntry,
    DocumentChunk,
    DocumentVersion,
    LegalDocument,
    RetrievalRun,
    SourceProvenanceRecord,
)
from legal_chatbot.documents.reranked_semantic_repository import (  # noqa: E402
    PostgresRerankedSemanticRepository,
)
from legal_chatbot.reranking.config import RerankerSettings  # noqa: E402
from legal_chatbot.reranking.fastembed_adapter import FastEmbedRerankerAdapter  # noqa: E402
from legal_chatbot.retrieval.models import RetrievalRequest  # noqa: E402
from legal_chatbot.semantic.config import SemanticSettings  # noqa: E402
from legal_chatbot.semantic.constants import SEMANTIC_DIMENSION, SEMANTIC_PROFILE_ID  # noqa: E402
from legal_chatbot.semantic.fastembed_adapter import FastEmbedSemanticAdapter  # noqa: E402
from legal_chatbot.semantic.models import SemanticEmbeddingBatch  # noqa: E402

DEFAULT_EXPERT = ROOT / "docs" / "Ket_qua_cham_toan_van_Stress_test_Legal_Chatbot_UEB_2026-08-22.xlsx"
DEFAULT_SET = ROOT / "docs" / "evals" / "m2_evaluation_set.json"
DEFAULT_XLSX = ROOT / "docs" / "evals" / "M2_retrieval_evaluation.xlsx"
DEFAULT_JSON = ROOT / "docs" / "evals" / "M2_retrieval_evaluation.json"
DEFAULT_MARKDOWN = ROOT / "docs" / "evals" / "M2_retrieval_before_after.md"
SOURCE_IDS = ("VBQPPL", "VNU", "UEB")
PARENT_IDS = tuple(f"Q{number:02d}" for number in range(1, 11))
DOCUMENT_NUMBER_PATTERN = re.compile(r"\d+(?:\s*/\s*[\wÀ-ỹĐđ.-]+){1,3}", re.UNICODE)
CONTROL_IDENTITY_CATEGORIES = frozenset(
    {"SYNTHETIC_METADATA_NUMBER", "AMBIGUOUS_DOCUMENT_IDENTITY"}
)


class EvaluationError(ValueError):
    """Raised for invalid evaluation-only input artifacts."""


@dataclass(frozen=True)
class ParentCase:
    case_id: str
    question: str = field(repr=False)
    expected_numbers: tuple[str, ...]


@dataclass(frozen=True)
class ControlCase:
    set_name: str
    case_id: str
    question: str = field(repr=False)
    category: str
    parent_case_id: str | None
    expected_behavior: str


@dataclass(frozen=True)
class SafeDiagnostics:
    semantic_count: int = 0
    identity_count: int = 0
    title_count: int = 0
    ambiguous_identity_count: int = 0
    metadata_no_support_count: int = 0
    pre_rerank_count: int = 0
    post_collapse_count: int = 0
    reranker_input_count: int = 0
    final_count: int = 0
    reranker_fallback: bool = False
    arm_contributions: tuple[tuple[str, int], ...] = ()
    rejection_reasons: tuple[tuple[str, int], ...] = ()


@dataclass(frozen=True)
class RetrievalObservation:
    set_name: str
    case_id: str
    mode: str
    category: str
    parent_case_id: str | None
    expected_behavior: str
    decision: str
    reason: str
    strategy: str
    latency_ms: float
    evidence_available: bool
    citation_count: int
    source_count: int
    citation_sources: tuple[str, ...]
    citation_document_numbers: tuple[str, ...]
    citations_resolvable: bool
    failure_code: str
    cleanup_status: str
    diagnostics: SafeDiagnostics


@dataclass(frozen=True)
class SetAScore:
    case_id: str
    mode: str
    expected_indexed_numbers: tuple[str, ...]
    inventory_blockers: tuple[str, ...]
    expected_sources: tuple[str, ...]
    cited_numbers: tuple[str, ...]
    hit_numbers: tuple[str, ...]
    denominator: int
    case_hit: bool
    non_expected_cited_count: int
    cited_identity_count: int
    source_hit_count: int
    source_denominator: int


@dataclass(frozen=True)
class SetBScore:
    case_id: str
    parent_case_id: str
    mode: str
    parent_evidence_available: bool
    paraphrase_evidence_available: bool
    overlap_count: int
    jaccard: float | None
    stability: str


@dataclass(frozen=True)
class SetCInvariant:
    case_id: str
    mode: str
    category: str
    invariant_failures: tuple[str, ...]


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _numbers(value: object) -> tuple[str, ...]:
    values: list[str] = []
    for match in DOCUMENT_NUMBER_PATTERN.findall(_text(value)):
        normalized = normalize_document_number(match)
        if normalized is not None and normalized not in values:
            values.append(normalized)
    return tuple(values)


def parse_expert_workbook(path: Path) -> tuple[ParentCase, ...]:
    """Separate parent questions from expected-number scoring metadata in the workbook."""

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Kết quả 10 câu" not in workbook.sheetnames or "Chấm điểm" not in workbook.sheetnames:
            raise EvaluationError("expert workbook lacks required evaluation sheets")
        questions = [
            _text(row[0])
            for row in workbook["Kết quả 10 câu"].iter_rows(min_row=2, values_only=True)
            if row and _text(row[0])
        ]
        expected: dict[str, tuple[str, ...]] = {}
        for row in workbook["Chấm điểm"].iter_rows(min_row=4, values_only=True):
            if not row or _text(row[0]) not in PARENT_IDS:
                continue
            padded = tuple(row) + (None,) * 7
            expected[_text(padded[0])] = _numbers(padded[6])
        if len(questions) != 10 or tuple(sorted(expected)) != PARENT_IDS:
            raise EvaluationError("expert workbook must provide exactly ten parent cases")
        if any(not expected[case_id] for case_id in PARENT_IDS):
            raise EvaluationError("each parent case requires at least one expected identity")
        return tuple(
            ParentCase(case_id, question, expected[case_id])
            for case_id, question in zip(PARENT_IDS, questions, strict=True)
        )
    finally:
        workbook.close()


def parse_m2_set(path: Path) -> tuple[ControlCase, ...]:
    """Load only question inputs and policy labels from the validated M2 JSON set."""

    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    try:
        sets = payload["sets"]
        result = []
        for set_name in ("B", "C"):
            for item in sets[set_name]["cases"]:
                result.append(
                    ControlCase(
                        set_name=set_name,
                        case_id=str(item["case_id"]),
                        question=str(item["question"]),
                        category=str(item["category"]),
                        parent_case_id=item["parent_case_id"],
                        expected_behavior=str(item["expected_behavior"]),
                    )
                )
    except (KeyError, TypeError) as error:
        raise EvaluationError("M2 set has an invalid shape") from error
    controls = tuple(result)
    if len([item for item in controls if item.set_name == "B"]) != 30:
        raise EvaluationError("M2 Set B must contain 30 controls")
    if len([item for item in controls if item.set_name == "C"]) < 20:
        raise EvaluationError("M2 Set C must contain at least 20 controls")
    if any(not item.question.strip() for item in controls):
        raise EvaluationError("M2 control questions must not be blank")
    return controls


def _safe_diagnostics(value: object | None) -> SafeDiagnostics:
    if value is None:
        return SafeDiagnostics()
    contributions = getattr(value, "arm_contribution_counts", {}) or {}
    rejections = getattr(value, "rejection_reason_counts", {}) or {}
    return SafeDiagnostics(
        semantic_count=int(
            getattr(value, "semantic_candidate_count", getattr(value, "pre_rerank_chunk_candidate_count", 0))
        ),
        identity_count=int(getattr(value, "exact_identity_candidate_count", 0)),
        title_count=int(getattr(value, "title_candidate_count", 0)),
        ambiguous_identity_count=int(getattr(value, "ambiguous_identity_count", 0)),
        metadata_no_support_count=int(getattr(value, "metadata_no_support_count", 0)),
        pre_rerank_count=int(
            getattr(value, "pre_dedup_count", getattr(value, "pre_rerank_chunk_candidate_count", 0))
        ),
        post_collapse_count=int(
            getattr(value, "post_document_collapse_count", getattr(value, "post_collapse_document_version_count", 0))
        ),
        reranker_input_count=int(getattr(value, "reranker_input_count", 0)),
        final_count=int(
            getattr(value, "final_count", getattr(value, "final_citation_document_version_count", 0))
        ),
        reranker_fallback=bool(getattr(value, "reranker_fallback", False)),
        arm_contributions=tuple(sorted((str(key), int(count)) for key, count in contributions.items())),
        rejection_reasons=tuple(sorted((str(key), int(count)) for key, count in rejections.items())),
    )


async def cleanup_run(session_factory: Any, retrieval_run_id: Any) -> str:
    """Delete only this evaluator-created run and its citations."""

    try:
        async with session_factory.begin() as session:
            await session.execute(
                delete(CitationRecord).where(CitationRecord.retrieval_run_id == retrieval_run_id)
            )
            await session.execute(delete(RetrievalRun).where(RetrievalRun.id == retrieval_run_id))
        return "COMPLETED"
    except Exception:
        return "FAILED_SAFE"


async def execute_case(
    case: ParentCase | ControlCase,
    *,
    mode: str,
    repository: Any,
    resolver: Any,
    cleanup: Callable[[Any], Awaitable[str]],
    diagnostics: list[object],
) -> RetrievalObservation:
    """Execute one bounded retrieval and remove its exact persisted evidence afterwards."""

    started = perf_counter()
    result: Any | None = None
    failure_code = ""
    citations_resolvable = True
    sources: set[str] = set()
    numbers: set[str] = set()
    diagnostics_start = len(diagnostics)
    try:
        retrieval_result = await repository.retrieve_and_persist(
            RetrievalRequest(query=case.question, top_k=3)
        )
        result = retrieval_result
        for candidate in retrieval_result.candidates:
            resolved = await resolver.resolve(candidate.citation_id, retrieval_result.retrieval_run_id)
            sources.add(str(resolved.source_id))
            if resolved.document_number:
                numbers.add(str(resolved.document_number))
    except Exception:
        citations_resolvable = False
        failure_code = "SAFE_EXCEPTION"
    finally:
        cleanup_status = "NOT_NEEDED"
        if result is not None:
            cleanup_status = await cleanup(result.retrieval_run_id)
            if cleanup_status != "COMPLETED":
                failure_code = failure_code or "CLEANUP_FAILED"
    diagnostic = diagnostics[-1] if len(diagnostics) > diagnostics_start else None
    return RetrievalObservation(
        set_name="A" if isinstance(case, ParentCase) else case.set_name,
        case_id=case.case_id,
        mode=mode,
        category="PARENT" if isinstance(case, ParentCase) else case.category,
        parent_case_id=case.case_id if isinstance(case, ParentCase) else case.parent_case_id,
        expected_behavior="RETRIEVAL_SCORING_ONLY"
        if isinstance(case, ParentCase)
        else case.expected_behavior,
        decision=str(getattr(getattr(result, "decision", None), "value", "SAFE_EXCEPTION")),
        reason=str(getattr(getattr(result, "reason", None), "value", "SAFE_EXCEPTION")),
        strategy=str(getattr(diagnostic, "strategy_version", "NOT_OBSERVED")),
        latency_ms=round((perf_counter() - started) * 1000, 3),
        evidence_available=bool(result is not None and getattr(result, "candidates", ())),
        citation_count=len(getattr(result, "candidates", ())) if citations_resolvable else 0,
        source_count=len(sources) if citations_resolvable else 0,
        citation_sources=tuple(sorted(sources)) if citations_resolvable else (),
        citation_document_numbers=tuple(sorted(numbers)) if citations_resolvable else (),
        citations_resolvable=citations_resolvable,
        failure_code=failure_code,
        cleanup_status=cleanup_status,
        diagnostics=_safe_diagnostics(diagnostic),
    )


async def evaluate_sequentially(
    parents: Iterable[ParentCase],
    controls: Iterable[ControlCase],
    repositories: dict[str, Any],
    resolver: Any,
    cleanup: Callable[[Any], Awaitable[str]],
    diagnostics_by_mode: dict[str, list[object]],
) -> tuple[RetrievalObservation, ...]:
    """Run A, B, and C sequentially for each mode; no query is retained in output."""

    observations: list[RetrievalObservation] = []
    for mode, repository in repositories.items():
        diagnostics = diagnostics_by_mode[mode]
        for case in (*tuple(parents), *tuple(controls)):
            observations.append(
                await execute_case(
                    case,
                    mode=mode,
                    repository=repository,
                    resolver=resolver,
                    cleanup=cleanup,
                    diagnostics=diagnostics,
                )
            )
    return tuple(observations)


async def indexed_oracle(
    session_factory: Any, expected_numbers: Iterable[str]
) -> dict[str, tuple[str, ...]]:
    """Find expected identities proven linked, indexed, and chunked in the current DB."""

    expected = tuple(sorted(set(expected_numbers)))
    if not expected:
        return {}
    latest = (
        select(func.max(DocumentVersion.version_number))
        .where(DocumentVersion.document_id == LegalDocument.id)
        .correlate(LegalDocument)
        .scalar_subquery()
    )
    statement = (
        select(DocumentVersion.document_number_normalized, LegalDocument.source_id)
        .select_from(DocumentVersion)
        .join(LegalDocument)
        .join(DocumentChunk)
        .join(CorpusCatalogEntry, CorpusCatalogEntry.document_version_id == DocumentVersion.id)
        .join(
            ChunkEmbedding,
            and_(
                ChunkEmbedding.document_chunk_id == DocumentChunk.id,
                ChunkEmbedding.embedding_model_id == SEMANTIC_PROFILE_ID,
                ChunkEmbedding.embedding_kind == "semantic",
                ChunkEmbedding.dimension == SEMANTIC_DIMENSION,
            ),
        )
        .where(
            DocumentVersion.document_number_normalized.in_(expected),
            DocumentVersion.version_number == latest,
            LegalDocument.source_id.in_(SOURCE_IDS),
            CorpusCatalogEntry.processing_status == "INDEXED",
        )
        .distinct()
    )
    async with session_factory() as session:
        rows = (await session.execute(statement)).all()
    values: dict[str, set[str]] = {}
    for number, source_id in rows:
        if number:
            values.setdefault(str(number), set()).add(str(source_id))
    return {number: tuple(sorted(sources)) for number, sources in values.items()}


async def read_recall_diagnostics(
    session_factory: Any,
    embedder: Any,
    metadata_repository: Any,
    parents: Iterable[ParentCase],
) -> tuple[dict[str, Any], ...]:
    """Read natural-query candidates without persistence or oracle filtering.

    Candidate selection is deliberately independent of expected identities. Expected
    identities are compared only after the exact-semantic top-50 and S1 merged
    pre-rerank candidates have been selected.
    """

    latest = (
        select(func.max(DocumentVersion.version_number))
        .where(DocumentVersion.document_id == LegalDocument.id)
        .correlate(LegalDocument)
        .scalar_subquery()
    )
    strict = (
        select(SourceProvenanceRecord.id)
        .where(
            SourceProvenanceRecord.document_version_id == DocumentVersion.id,
            SourceProvenanceRecord.transport_trust_mode == "STRICT_TLS",
            SourceProvenanceRecord.provenance_type.in_(("source_fetch", "manual_snapshot")),
        )
        .correlate(DocumentVersion)
        .exists()
    )
    output: list[dict[str, Any]] = []
    for parent in parents:
        availability = {8: False, 16: False, 20: False, 50: False}
        s1_merged_at_16 = False
        try:
            batch = await embedder.embed_query(parent.question)
            if not isinstance(batch, SemanticEmbeddingBatch) or len(batch.vectors) != 1:
                raise EvaluationError("semantic embedding unavailable")
            vector = batch.vectors[0]
            if len(vector) != SEMANTIC_DIMENSION:
                raise EvaluationError("semantic embedding dimension mismatch")
            distance = ChunkEmbedding.embedding.cosine_distance(list(vector))
            statement = (
                select(DocumentVersion.document_number_normalized)
                .select_from(DocumentChunk)
                .join(DocumentVersion)
                .join(LegalDocument)
                .join(
                    ChunkEmbedding,
                    and_(
                        ChunkEmbedding.document_chunk_id == DocumentChunk.id,
                        ChunkEmbedding.embedding_model_id == SEMANTIC_PROFILE_ID,
                        ChunkEmbedding.embedding_kind == "semantic",
                        ChunkEmbedding.dimension == SEMANTIC_DIMENSION,
                    ),
                )
                .where(
                    LegalDocument.source_id.in_(SOURCE_IDS),
                    DocumentVersion.version_number == latest,
                    strict,
                )
                .order_by(distance.asc(), DocumentChunk.ordinal.asc(), DocumentChunk.id.asc())
                .limit(50)
            )
            async with session_factory() as session:
                numbers = tuple(str(row[0]) for row in (await session.execute(statement)).all() if row[0])
            expected = set(parent.expected_numbers)
            for limit in availability:
                availability[limit] = bool(expected & set(numbers[:limit]))

            merged, _diagnostics = await metadata_repository._read(  # noqa: SLF001
                RetrievalRequest(query=parent.question, top_k=3), tuple(vector)
            )
            version_ids = tuple(item.version_id for item in merged[:16])
            if version_ids:
                async with session_factory() as session:
                    merged_numbers = {
                        str(row[1])
                        for row in (
                            await session.execute(
                                select(
                                    DocumentVersion.id, DocumentVersion.document_number_normalized
                                ).where(DocumentVersion.id.in_(version_ids))
                            )
                        ).all()
                        if row[1]
                    }
                s1_merged_at_16 = bool(expected & merged_numbers)
        except Exception:
            # Diagnostic availability remains false; this is not an evaluation crash.
            pass
        output.append(
            {
                "case_id": parent.case_id,
                "mode": "S0",
                "exact_semantic_available_at_8": availability[8],
                "exact_semantic_available_at_16": availability[16],
                "exact_semantic_available_at_20": availability[20],
                "exact_semantic_available_at_50": availability[50],
            }
        )
        output.append(
            {
                "case_id": parent.case_id,
                "mode": "S1",
                "pre_rerank_merged_available_at_16": s1_merged_at_16,
            }
        )
    return tuple(output)


def score_set_a(
    parents: Iterable[ParentCase],
    observations: Iterable[RetrievalObservation],
    oracle: dict[str, tuple[str, ...]],
) -> tuple[SetAScore, ...]:
    """Score retrieved top-three identities only after retrieval has completed."""

    parents_by_id = {case.case_id: case for case in parents}
    scores = []
    for observation in observations:
        if observation.set_name != "A":
            continue
        parent = parents_by_id[observation.case_id]
        expected_indexed = tuple(number for number in parent.expected_numbers if number in oracle)
        blockers = tuple(number for number in parent.expected_numbers if number not in oracle)
        expected_sources = tuple(sorted({source for number in expected_indexed for source in oracle[number]}))
        cited_normalized = {
            normalized
            for number in observation.citation_document_numbers
            if (normalized := normalize_document_number(number)) is not None
        }
        hits = tuple(sorted(set(expected_indexed) & cited_normalized))
        cited_sources = set(observation.citation_sources)
        non_expected = len(cited_normalized - set(parent.expected_numbers))
        scores.append(
            SetAScore(
                case_id=observation.case_id,
                mode=observation.mode,
                expected_indexed_numbers=expected_indexed,
                inventory_blockers=blockers,
                expected_sources=expected_sources,
                cited_numbers=tuple(sorted(cited_normalized)),
                hit_numbers=hits,
                denominator=len(expected_indexed),
                case_hit=bool(hits),
                non_expected_cited_count=non_expected,
                cited_identity_count=len(cited_normalized),
                source_hit_count=len(cited_sources & set(expected_sources)),
                source_denominator=len(expected_sources),
            )
        )
    return tuple(scores)


def score_set_b(observations: Iterable[RetrievalObservation]) -> tuple[SetBScore, ...]:
    """Compare B citation-number sets against their same-mode parent retrieval sets."""

    all_observations = tuple(observations)
    parents = {(item.mode, item.case_id): item for item in all_observations if item.set_name == "A"}
    scores = []
    for item in all_observations:
        if item.set_name != "B" or item.parent_case_id is None:
            continue
        parent = parents[(item.mode, item.parent_case_id)]
        left, right = set(parent.citation_document_numbers), set(item.citation_document_numbers)
        if not left or not right:
            jaccard, stability = None, "PARENT_OR_PARAPHRASE_NO_EVIDENCE"
        else:
            overlap = len(left & right)
            jaccard = overlap / len(left | right)
            stability = "EXACT" if left == right else "PARTIAL" if overlap else "DISJOINT"
        scores.append(
            SetBScore(
                case_id=item.case_id,
                parent_case_id=item.parent_case_id,
                mode=item.mode,
                parent_evidence_available=parent.evidence_available,
                paraphrase_evidence_available=item.evidence_available,
                overlap_count=len(left & right),
                jaccard=jaccard,
                stability=stability,
            )
        )
    return tuple(scores)


def check_set_c_invariants(observations: Iterable[RetrievalObservation]) -> tuple[SetCInvariant, ...]:
    """Check only contract-safe control invariants, never answer/legal correctness."""

    results = []
    for item in observations:
        if item.set_name != "C":
            continue
        failures = []
        if item.failure_code:
            failures.append(item.failure_code)
        if not item.citations_resolvable:
            failures.append("UNRESOLVABLE_CITATION")
        if item.category in CONTROL_IDENTITY_CATEGORIES and item.diagnostics.identity_count:
            failures.append("ARBITRARY_IDENTITY_CANDIDATE")
        results.append(SetCInvariant(item.case_id, item.mode, item.category, tuple(failures)))
    return tuple(results)


def _percentile(values: Iterable[float], percent: int) -> float:
    ordered = sorted(values)
    if not ordered:
        return 0.0
    index = max(0, min(len(ordered) - 1, (len(ordered) * percent + 99) // 100 - 1))
    return round(ordered[index], 3)


def summary_metrics(
    set_a: Iterable[SetAScore], set_b: Iterable[SetBScore], set_c: Iterable[SetCInvariant], observations: Iterable[RetrievalObservation]
) -> dict[str, dict[str, float | int]]:
    """Aggregate safe, retrieval-only metrics for the summary and mechanical gate."""

    result: dict[str, dict[str, float | int]] = {}
    observations_by_mode: dict[str, list[RetrievalObservation]] = {}
    for item in observations:
        observations_by_mode.setdefault(item.mode, []).append(item)
    a_by_mode: dict[str, list[SetAScore]] = {}
    b_by_mode: dict[str, list[SetBScore]] = {}
    c_by_mode: dict[str, list[SetCInvariant]] = {}
    for item in set_a:
        a_by_mode.setdefault(item.mode, []).append(item)
    for item in set_b:
        b_by_mode.setdefault(item.mode, []).append(item)
    for item in set_c:
        c_by_mode.setdefault(item.mode, []).append(item)
    for mode in sorted(observations_by_mode):
        a_rows, b_rows, c_rows = a_by_mode.get(mode, []), b_by_mode.get(mode, []), c_by_mode.get(mode, [])
        hits = sum(len(row.hit_numbers) for row in a_rows)
        denominator = sum(row.denominator for row in a_rows)
        non_expected = sum(row.non_expected_cited_count for row in a_rows)
        cited = sum(row.cited_identity_count for row in a_rows)
        source_hits = sum(row.source_hit_count for row in a_rows)
        source_denominator = sum(row.source_denominator for row in a_rows)
        observed = observations_by_mode[mode]
        evidence_both = [row for row in b_rows if row.jaccard is not None]
        result[mode] = {
            "set_a_identity_hits": hits,
            "set_a_inventory_denominator": denominator,
            "set_a_hit_rate": round(hits / denominator, 6) if denominator else 0.0,
            "set_a_case_hits": sum(row.case_hit for row in a_rows),
            "set_a_non_expected_cited_rate": round(non_expected / cited, 6) if cited else 0.0,
            "set_a_source_coverage": round(source_hits / source_denominator, 6)
            if source_denominator
            else 0.0,
            "set_a_p50_latency_ms": _percentile(
                (item.latency_ms for item in observed if item.set_name == "A"), 50
            ),
            "set_a_p95_latency_ms": _percentile(
                (item.latency_ms for item in observed if item.set_name == "A"), 95
            ),
            "set_b_mean_jaccard": round(
                sum(row.jaccard or 0.0 for row in evidence_both) / len(evidence_both), 6
            )
            if evidence_both
            else 0.0,
            "set_b_evidence_rate": round(
                sum(row.paraphrase_evidence_available for row in b_rows) / len(b_rows), 6
            )
            if b_rows
            else 0.0,
            "set_c_invariant_failures": sum(len(row.invariant_failures) for row in c_rows),
            "set_c_ambiguous_identity_count": sum(
                item.diagnostics.ambiguous_identity_count
                for item in observed
                if item.set_name == "C"
            ),
        }
    return result


def mechanical_gate(metrics: dict[str, dict[str, float | int]]) -> dict[str, bool | str]:
    """Compute a non-release comparison gate; it intentionally never approves release."""

    s0, s1 = metrics.get("S0", {}), metrics.get("S1", {})
    hit_improves = s1.get("set_a_identity_hits", 0) > s0.get("set_a_identity_hits", 0)
    non_expected_not_worse = s1.get("set_a_non_expected_cited_rate", 0.0) <= s0.get(
        "set_a_non_expected_cited_rate", 0.0
    )
    no_invariant_failures = s1.get("set_c_invariant_failures", 0) == 0
    return {
        "s1_hit_improves": bool(hit_improves),
        "s1_non_expected_not_worse": bool(non_expected_not_worse),
        "s1_no_invariant_failures": bool(no_invariant_failures),
        "mechanical_gates_pass": bool(hit_improves and non_expected_not_worse and no_invariant_failures),
        "recommendation": "HOLD_PENDING_ORACLE",
    }


def _table(sheet: Any, header: tuple[str, ...], rows: Iterable[tuple[Any, ...]]) -> None:
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(
            42, max(12, max(len(str(cell.value or "")) for cell in column) + 2)
        )


def write_reports(
    *,
    xlsx_path: Path,
    json_path: Path,
    markdown_path: Path,
    observations: tuple[RetrievalObservation, ...],
    set_a: tuple[SetAScore, ...],
    set_b: tuple[SetBScore, ...],
    set_c: tuple[SetCInvariant, ...],
    arm_diagnostics: tuple[dict[str, Any], ...],
) -> None:
    """Write report-only safe metadata; question and evidence text are intentionally absent."""

    metrics = summary_metrics(set_a, set_b, set_c, observations)
    gate = mechanical_gate(metrics)
    for mode in ("S0", "S1"):
        mode_rows = [item for item in arm_diagnostics if item.get("mode") == mode]
        for key in (
            "exact_semantic_available_at_8",
            "exact_semantic_available_at_16",
            "exact_semantic_available_at_20",
            "exact_semantic_available_at_50",
            "pre_rerank_merged_available_at_16",
        ):
            values = [bool(item[key]) for item in mode_rows if key in item]
            if values:
                metrics.setdefault(mode, {})[f"{key}_rate"] = round(sum(values) / len(values), 6)
    payload = {
        "report_schema_version": "M2-RETRIEVAL-EVALUATION-1",
        "methodology_limitations": [
            "Retrieval behavior only; no legal correctness, authority, completeness, or currentness claim.",
            "Set B measures citation-set stability, not legal-answer correctness.",
            "Set C checks only contract-safe invariants, not no-evidence accuracy.",
            "Expected identity scoring is post-retrieval and inventory-limited to linked/indexed/chunked identities.",
        ],
        "summary": metrics,
        "mechanical_gate": gate,
        "set_a": [asdict(item) for item in set_a],
        "set_b": [asdict(item) for item in set_b],
        "set_c": [asdict(item) for item in set_c],
        "observations": [asdict(item) for item in observations],
        "arm_diagnostics": list(arm_diagnostics),
    }
    for path in (xlsx_path, json_path, markdown_path):
        path.parent.mkdir(parents=True, exist_ok=True)
    with json_path.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")

    workbook = openpyxl.Workbook()
    summary = workbook.active
    if summary is None:
        raise RuntimeError("openpyxl did not create Summary")
    summary.title = "Summary"
    summary_rows = [
        ("Scope", "Retrieval-only; no provider/chat/network calls."),
        ("Recommendation", gate["recommendation"]),
        ("Mechanical gates pass", gate["mechanical_gates_pass"]),
        ("Limitation", "No legal correctness, authority, completeness, or currentness claim."),
    ]
    for mode in ("S0", "S1"):
        for metric, value in sorted(metrics.get(mode, {}).items()):
            summary_rows.append((f"{mode} {metric}", value))
    _table(summary, ("Metric", "Value"), summary_rows)
    observations_by_key = {(item.mode, item.case_id): item for item in observations}
    _table(
        workbook.create_sheet("Set A"),
        (
            "Case ID", "Mode", "Expected indexed identities", "Inventory blockers", "Expected sources", "Cited identities",
            "Hit identities", "Inventory denominator", "Case hit", "Non-expected cited count",
            "Cited identity count", "Source hits", "Source denominator", "Decision", "Reason",
            "Strategy", "Latency ms",
        ),
        tuple(
            (
                item.case_id, item.mode, "; ".join(item.expected_indexed_numbers),
                "; ".join(item.inventory_blockers),
                "; ".join(item.expected_sources), "; ".join(item.cited_numbers),
                "; ".join(item.hit_numbers), item.denominator, item.case_hit,
                item.non_expected_cited_count, item.cited_identity_count, item.source_hit_count,
                item.source_denominator,
                observations_by_key[(item.mode, item.case_id)].decision,
                observations_by_key[(item.mode, item.case_id)].reason,
                observations_by_key[(item.mode, item.case_id)].strategy,
                observations_by_key[(item.mode, item.case_id)].latency_ms,
            )
            for item in set_a
        ),
    )
    _table(
        workbook.create_sheet("Set B"),
        (
            "Case ID", "Parent case ID", "Mode", "Parent evidence available",
            "Paraphrase evidence available", "Overlap count", "Jaccard", "Stability",
        ),
        (
            (
                item.case_id, item.parent_case_id, item.mode, item.parent_evidence_available,
                item.paraphrase_evidence_available, item.overlap_count, item.jaccard, item.stability,
            )
            for item in set_b
        ),
    )
    _table(
        workbook.create_sheet("Set C"),
        (
            "Case ID", "Mode", "Category", "Decision", "Evidence available", "Source count",
            "Citation count", "Identity ambiguity count", "Semantic arm count", "Title arm count",
            "Strategy", "Latency ms", "Invariant failures",
        ),
        (
            (
                item.case_id, item.mode, item.category,
                observations_by_key[(item.mode, item.case_id)].decision,
                observations_by_key[(item.mode, item.case_id)].evidence_available,
                observations_by_key[(item.mode, item.case_id)].source_count,
                observations_by_key[(item.mode, item.case_id)].citation_count,
                observations_by_key[(item.mode, item.case_id)].diagnostics.ambiguous_identity_count,
                observations_by_key[(item.mode, item.case_id)].diagnostics.semantic_count,
                observations_by_key[(item.mode, item.case_id)].diagnostics.title_count,
                observations_by_key[(item.mode, item.case_id)].strategy,
                observations_by_key[(item.mode, item.case_id)].latency_ms,
                "; ".join(item.invariant_failures),
            )
            for item in set_c
        ),
    )
    _table(
        workbook.create_sheet("Arm Diagnostics"),
        (
            "Set", "Case ID", "Mode", "Strategy", "Semantic count", "Identity count", "Title count",
            "Ambiguous identity count", "Metadata no-support", "Pre-rerank count", "Post-collapse count",
            "Reranker input count", "Final count", "Fallback", "Arm contributions", "Rejections",
        ),
        tuple(
            (
                item.set_name, item.case_id, item.mode, item.strategy,
                item.diagnostics.semantic_count, item.diagnostics.identity_count,
                item.diagnostics.title_count, item.diagnostics.ambiguous_identity_count,
                item.diagnostics.metadata_no_support_count, item.diagnostics.pre_rerank_count,
                item.diagnostics.post_collapse_count, item.diagnostics.reranker_input_count,
                item.diagnostics.final_count, item.diagnostics.reranker_fallback,
                "; ".join(f"{key}:{value}" for key, value in item.diagnostics.arm_contributions),
                "; ".join(f"{key}:{value}" for key, value in item.diagnostics.rejection_reasons),
            )
            for item in observations
        )
        + tuple(
            (
                "RECALL", item["case_id"], item["mode"], "READ_ONLY_CANDIDATE_DIAGNOSTIC",
                "", "", "", "", "", "", "", "", "", "",
                "; ".join(f"{key}:{int(value)}" for key, value in item.items() if key not in {"case_id", "mode"}),
                "",
            )
            for item in arm_diagnostics
        ),
    )
    workbook.save(xlsx_path)
    workbook.close()
    markdown_path.write_text(
        "# M2 retrieval before/after\n\n"
        "Recommendation: **HOLD_PENDING_ORACLE** (not a release decision).\n\n"
        f"Mechanical gates pass: `{gate['mechanical_gates_pass']}`. S1 expected-identity hits improve: "
        f"`{gate['s1_hit_improves']}`; S1 non-expected cited rate is not worse: "
        f"`{gate['s1_non_expected_not_worse']}`; S1 has no Set C invariant failures: "
        f"`{gate['s1_no_invariant_failures']}`.\n\n"
        "This is retrieval-only measurement. It does not establish legal correctness, authority, "
        "completeness, currentness, no-evidence accuracy, or a release decision. Set B measures "
        "citation-set stability only. Questions, answers, titles, URLs, chunks, UUIDs, and prompts "
        "are excluded from the report.\n",
        encoding="utf-8",
    )


async def run(args: argparse.Namespace) -> int:
    parents = parse_expert_workbook(args.expert_workbook)
    controls = parse_m2_set(args.m2_set)
    engine: Any | None = None
    try:
        engine = create_engine(Settings())  # type: ignore[call-arg]
        session_factory = create_session_factory(engine)
        semantic = FastEmbedSemanticAdapter(SemanticSettings())
        reranker = FastEmbedRerankerAdapter(RerankerSettings())
        diagnostics = {"S0": [], "S1": []}
        repositories = {
            "S0": PostgresRerankedSemanticRepository(
                session_factory, SOURCE_IDS, semantic, reranker, observer=diagnostics["S0"].append
            ),
            "S1": PostgresMetadataRepairRetrievalRepository(
                session_factory, SOURCE_IDS, semantic, reranker, observer=diagnostics["S1"].append
            ),
        }
        observations = await evaluate_sequentially(
            parents,
            controls,
            repositories,
            PostgresCitationResolver(session_factory),
            lambda run_id: cleanup_run(session_factory, run_id),
            diagnostics,
        )
        expected = (number for parent in parents for number in parent.expected_numbers)
        oracle = await indexed_oracle(session_factory, expected)
        set_a = score_set_a(parents, observations, oracle)
        set_b = score_set_b(observations)
        set_c = check_set_c_invariants(observations)
        recall = await read_recall_diagnostics(session_factory, semantic, repositories["S1"], parents)
        write_reports(
            xlsx_path=args.output,
            json_path=args.json_output,
            markdown_path=args.markdown_output,
            observations=observations,
            set_a=set_a,
            set_b=set_b,
            set_c=set_c,
            arm_diagnostics=recall,
        )
        return 0
    except Exception:
        return 2
    finally:
        if engine is not None:
            await engine.dispose()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run M2 retrieval-only evaluation.")
    parser.add_argument("--expert-workbook", type=Path, default=DEFAULT_EXPERT)
    parser.add_argument("--m2-set", type=Path, default=DEFAULT_SET)
    parser.add_argument("--output", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--json-output", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--markdown-output", type=Path, default=DEFAULT_MARKDOWN)
    return parser


def main() -> int:
    try:
        return asyncio.run(run(build_parser().parse_args()))
    except Exception:
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
