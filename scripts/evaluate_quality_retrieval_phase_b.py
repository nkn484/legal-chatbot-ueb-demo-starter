"""Offline-only Phase B4/B5 quality-retrieval candidate evaluator.

The evaluator deliberately keeps queries and candidate text in memory.  Reports are
content-free: they contain case identifiers, approved document numbers, counts and
timings only.  It is not a production retrieval path.
"""

# ruff: noqa: E501, E402
from __future__ import annotations

import argparse
import ast
import asyncio
import hashlib
import json
import sys
from collections import defaultdict
from collections.abc import Iterable
from dataclasses import dataclass, field
from pathlib import Path
from time import perf_counter
from typing import Any
from uuid import UUID, uuid4

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import and_, delete, func, select, text

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "src") not in sys.path:
    sys.path.insert(0, str(ROOT / "src"))

from legal_chatbot.core.config import Settings  # noqa: E402
from legal_chatbot.db.session import create_engine, create_session_factory  # noqa: E402
from legal_chatbot.documents.citation_resolver import PostgresCitationResolver  # noqa: E402
from legal_chatbot.documents.metadata_normalization import normalize_document_number  # noqa: E402
from legal_chatbot.documents.orm import (  # noqa: E402
    ChunkEmbedding,
    CitationRecord,
    CorpusCatalogEntry,
    DocumentChunk,
    DocumentVersion,
    LegalDocument,
    RetrievalRun,
    ReviewedLegalEffectAssertion,
    ReviewedLegalEffectEvent,
    ReviewedLegalEffectFamily,
    ReviewedLegalEffectImport,
    SourceProvenanceRecord,
)
from legal_chatbot.documents.quality_candidate_reader import (
    FTSQueryMode,  # noqa: E402
    PostgresQualityCandidateReader,  # noqa: E402
)
from legal_chatbot.retrieval.config import RetrievalSettings  # noqa: E402
from legal_chatbot.retrieval.quality_repair.models import (  # noqa: E402
    CandidateEvidence,
    CollapsedDocumentCandidate,
    RetrievalLane,
)
from legal_chatbot.retrieval.quality_repair.ranking import (  # noqa: E402
    PoolMeasurementSummary,
    PoolReferenceSummary,
    PoolSelectionStatus,
    build_lane_document_pool,
    fused_diagnostic_top50,
    merge_chunk_candidates,
    select_final_top3,
    select_pareto_pool,
    with_lane_unique_contributions,
)
from legal_chatbot.semantic.config import SemanticSettings  # noqa: E402
from legal_chatbot.semantic.constants import SEMANTIC_DIMENSION, SEMANTIC_PROFILE_ID  # noqa: E402
from legal_chatbot.semantic.fastembed_adapter import FastEmbedSemanticAdapter  # noqa: E402
from legal_chatbot.semantic.models import SemanticEmbeddingBatch  # noqa: E402

DEFAULT_EXPERT = (
    ROOT / "docs" / "Ket_qua_cham_toan_van_Stress_test_Legal_Chatbot_UEB_2026-08-22.xlsx"
)
DEFAULT_SET = ROOT / "docs" / "evals" / "m2_evaluation_set.json"
DEFAULT_OUTPUT = ROOT / "docs" / "evals" / "quality-retrieval" / "phase-b-candidate-evaluation.xlsx"
SOURCE_IDS = ("VBQPPL", "VNU", "UEB")
POOL_SIZES = (8, 12, 16, 20)
PARENT_IDS = tuple(f"Q{value:02d}" for value in range(1, 11))
EXPECTED_DENOMINATOR = 29
_REVIEWED_COUNT_TABLES = {
    "reviewed_effect_imports": ReviewedLegalEffectImport,
    "reviewed_effect_families": ReviewedLegalEffectFamily,
    "reviewed_effect_assertions": ReviewedLegalEffectAssertion,
    "reviewed_effect_events": ReviewedLegalEffectEvent,
    "retrieval_runs": RetrievalRun,
    "citations": CitationRecord,
}
_REVIEWED_REGISTRY_COUNT_NAMES = tuple(
    name for name in _REVIEWED_COUNT_TABLES if name.startswith("reviewed_effect_")
)
_QUALITY_FLAG_NAMES = (
    "lexical_repair_enabled",
    "semantic_hybrid_enabled",
    "rerank_enabled",
    "metadata_repair_enabled",
    "quality_repair_enabled",
    "quality_title_search_enabled",
    "quality_hybrid_fusion_enabled",
    "quality_query_planner_enabled",
    "quality_dynamic_evidence_enabled",
    "quality_repair_retrieval_enabled",
)


class EvaluationError(RuntimeError):
    """A controlled evaluator/infrastructure invariant failure."""


async def _state_counts(session_factory: Any) -> dict[str, int]:
    """Snapshot evaluator write targets in one explicit repeatable-read transaction."""

    async with session_factory() as session:
        async with session.begin():
            await session.execute(
                text("SET TRANSACTION ISOLATION LEVEL REPEATABLE READ, READ ONLY")
            )
            return {
                name: int((await session.scalar(select(func.count()).select_from(table))) or 0)
                for name, table in _REVIEWED_COUNT_TABLES.items()
            }


def _runtime_state_flags() -> dict[str, object]:
    """Inspect default/active retrieval settings and static runtime imports without activation."""

    active_settings = RetrievalSettings()
    defaults = {
        **{
            name: RetrievalSettings.model_fields[name].default for name in _QUALITY_FLAG_NAMES
        },
        "quality_strategy": RetrievalSettings.model_fields["quality_strategy"].default,
    }
    active = {
        **{name: getattr(active_settings, name) for name in _QUALITY_FLAG_NAMES},
        "quality_strategy": active_settings.quality_strategy,
    }
    service = ast.parse(
        (ROOT / "src" / "legal_chatbot" / "retrieval" / "service.py").read_text(encoding="utf-8")
    )
    imported_modules = {
        module
        for node in ast.walk(service)
        for module in (
            ([node.module] if isinstance(node, ast.ImportFrom) and node.module else [])
            or ([alias.name for alias in node.names] if isinstance(node, ast.Import) else [])
        )
    }
    static_runtime = {
        "runtime_service_imports_reviewed_effects": any(
            module.startswith("legal_chatbot.legal_effects") for module in imported_modules
        ),
        "runtime_service_imports_quality_execution": any(
            module.startswith("legal_chatbot.retrieval.quality_repair") for module in imported_modules
        ),
    }
    quality_defaults_off = not any(defaults[name] for name in _QUALITY_FLAG_NAMES) and (
        defaults["quality_strategy"] == "disabled"
    )
    quality_active_off = not any(active[name] for name in _QUALITY_FLAG_NAMES) and (
        active["quality_strategy"] == "disabled"
    )
    reviewed_effects_off = not any(static_runtime.values())
    return {
        "defaults": defaults,
        "active": active,
        "static_runtime": static_runtime,
        "quality_defaults_off": quality_defaults_off,
        "quality_active_off": quality_active_off,
        "reviewed_effects_off": reviewed_effects_off,
        "flags_off": quality_defaults_off and quality_active_off and reviewed_effects_off,
    }


def _state_invariants(
    before_counts: dict[str, int], after_counts: dict[str, int], flags: dict[str, object]
) -> dict[str, object]:
    """Build a content-free evaluator state assertion from read-only snapshots."""

    return {
        "before_counts": before_counts,
        "after_counts": after_counts,
        "counts_unchanged": before_counts == after_counts,
        "reviewed_effect_registry_zero": all(
            before_counts[name] == 0 and after_counts[name] == 0
            for name in _REVIEWED_REGISTRY_COUNT_NAMES
        ),
        "flags": flags,
    }


@dataclass(frozen=True)
class ParentCase:
    case_id: str
    question: str = field(repr=False)
    expected_numbers: tuple[str, ...]


@dataclass(frozen=True)
class ControlCase:
    set_name: str
    case_id: str
    question: str = field(repr=False)
    category: str
    parent_case_id: str | None


@dataclass(frozen=True)
class QueryResult:
    """Private per-query material. Candidate identities never reach report serialization."""

    case_id: str
    set_name: str
    category: str
    embedding_ms: float
    reader_ms: float
    ranking_ms: float
    retrieval_eval_ms: float
    end_to_end_ms: float
    transaction_elapsed_ms: float
    data_query_count: int
    explain_query_count: int
    query_count: int
    lane_candidates: dict[RetrievalLane, tuple[CandidateEvidence, ...]] = field(repr=False)
    lane_metrics: tuple[Any, ...] = field(repr=False)
    diagnostic: Any = field(repr=False)
    pools: dict[int, Any] = field(repr=False)
    finals: dict[int, tuple[CollapsedDocumentCandidate, ...]] = field(repr=False)
    requested_fts_query_mode: FTSQueryMode = FTSQueryMode.NATURAL
    applied_fts_query_mode: FTSQueryMode = FTSQueryMode.NATURAL
    fts_preparation_query_count: int = 0
    fts_preparation_elapsed_ms: float = 0.0
    bounded_or_selected_lexeme_count: int = 0
    bounded_or_source_lexeme_count: int = 0
    bounded_or_truncated: bool = False
    bounded_or_empty_query: bool = False
    bounded_or_natural_fallback_used: bool = False


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _numbers(value: object) -> tuple[str, ...]:
    """Normalize workbook metadata, never use it as a retrieval input."""
    import re

    values: list[str] = []
    for raw in re.findall(r"\d+(?:\s*/\s*[\wÀ-ỹĐđ.-]+){1,3}", _text(value)):
        normalized = normalize_document_number(raw)
        if normalized and normalized not in values:
            values.append(normalized)
    return tuple(values)


def parse_expert_workbook(path: Path) -> tuple[ParentCase, ...]:
    """Parse questions separately from oracle identities; they are joined only post-score."""
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Kết quả 10 câu" not in book.sheetnames or "Chấm điểm" not in book.sheetnames:
            raise EvaluationError("expert workbook lacks required sheets")
        questions = [
            _text(row[0])
            for row in book["Kết quả 10 câu"].iter_rows(min_row=2, values_only=True)
            if row and _text(row[0])
        ]
        expected: dict[str, tuple[str, ...]] = {}
        for row in book["Chấm điểm"].iter_rows(min_row=4, values_only=True):
            row_values = tuple(row or ())
            if row_values and _text(row_values[0]) in PARENT_IDS:
                expected[_text(row_values[0])] = _numbers(
                    row_values[6] if len(row_values) > 6 else None
                )
        if (
            len(questions) != 10
            or tuple(sorted(expected)) != PARENT_IDS
            or any(not expected[key] for key in PARENT_IDS)
        ):
            raise EvaluationError("expert workbook must provide ten complete parent cases")
        return tuple(
            ParentCase(case_id, question, expected[case_id])
            for case_id, question in zip(PARENT_IDS, questions, strict=True)
        )
    finally:
        book.close()


def parse_m2_set(path: Path) -> tuple[ControlCase, ...]:
    """Load B/C query inputs and labels without interpreting expected legal outcomes."""
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))["sets"]
        cases = tuple(
            ControlCase(
                set_name,
                str(item["case_id"]),
                str(item["question"]),
                str(item["category"]),
                item.get("parent_case_id"),
            )
            for set_name in ("B", "C")
            for item in payload[set_name]["cases"]
        )
    except (KeyError, TypeError, json.JSONDecodeError) as error:
        raise EvaluationError("invalid M2 evaluation set") from error
    if (
        len([case for case in cases if case.set_name == "B"]) != 30
        or len([case for case in cases if case.set_name == "C"]) != 24
    ):
        raise EvaluationError("M2 set must contain Set B30 and Set C24")
    if any(not case.question.strip() for case in cases):
        raise EvaluationError("M2 cases must have nonblank questions")
    return cases


def _identity_number(candidate: CollapsedDocumentCandidate) -> str | None:
    return candidate.identity.document_number_normalized


def _ids(candidates: Iterable[CollapsedDocumentCandidate]) -> set[UUID]:
    return {candidate.identity.document_version_id for candidate in candidates}


def _merged(result: QueryResult) -> tuple[CandidateEvidence, ...]:
    """Reuse already-read candidates without an additional reader transaction."""
    return merge_chunk_candidates(
        candidate for lane in RetrievalLane for candidate in result.lane_candidates.get(lane, ())
    ).candidates


def _semantic_final(result: QueryResult) -> tuple[CollapsedDocumentCandidate, ...]:
    """Natural semantic-only reference final selection, not a fused release pool."""
    return build_lane_document_pool(_merged(result), RetrievalLane.SEMANTIC, 8).candidates[:3]


def _lane_pool(
    result: QueryResult, lane: RetrievalLane, size: int
) -> tuple[CollapsedDocumentCandidate, ...]:
    return build_lane_document_pool(_merged(result), lane, size).candidates


def _raw_lane(result: QueryResult, lane: RetrievalLane) -> tuple[CandidateEvidence, ...]:
    """Return the reader's rank-ordered raw chunk budget for one lane."""
    return tuple(
        sorted(
            result.lane_candidates.get(lane, ()),
            key=lambda candidate: next(
                observation.rank
                for observation in candidate.observations
                if observation.lane is lane
            ),
        )
    )


def _expected_numbers(
    candidates: Iterable[CollapsedDocumentCandidate | CandidateEvidence], expected: Iterable[str]
) -> set[str]:
    expected_set = set(expected)
    return {
        number
        for candidate in candidates
        if (number := candidate.identity.document_number_normalized) is not None
        and number in expected_set
    }


def _nonexpected_identity_count(
    candidates: Iterable[CollapsedDocumentCandidate], expected: Iterable[str]
) -> int:
    expected_set = set(expected)
    return sum(
        candidate.identity.document_number_normalized not in expected_set
        for candidate in {
            candidate.identity.document_version_id: candidate for candidate in candidates
        }.values()
    )


def fts_read_metadata(results: Iterable[QueryResult]) -> tuple[dict[str, Any], ...]:
    """Return content-free per-read FTS construction facts for B2A comparison."""

    return tuple(
        {
            "case_id": result.case_id,
            "set_name": result.set_name,
            "category": result.category,
            "requested_fts_query_mode": result.requested_fts_query_mode.value,
            "applied_fts_query_mode": result.applied_fts_query_mode.value,
            "fts_preparation_query_count": result.fts_preparation_query_count,
            "fts_preparation_elapsed_ms": result.fts_preparation_elapsed_ms,
            "bounded_or_selected_lexeme_count": result.bounded_or_selected_lexeme_count,
            "bounded_or_source_lexeme_count": result.bounded_or_source_lexeme_count,
            "bounded_or_truncated": result.bounded_or_truncated,
            "bounded_or_empty_query": result.bounded_or_empty_query,
            "bounded_or_natural_fallback_used": result.bounded_or_natural_fallback_used,
            "data_query_count": result.data_query_count,
            "explain_query_count": result.explain_query_count,
            "query_count": result.query_count,
            "reader_ms": result.reader_ms,
            "retrieval_eval_ms": result.retrieval_eval_ms,
            "transaction_elapsed_ms": result.transaction_elapsed_ms,
        }
        for result in results
    )


def phase_b2a_comparison_contract(
    parents: Iterable[ParentCase],
    controls: Iterable[ControlCase],
    oracle: dict[str, tuple[str, ...]],
    denominator: int,
    fts_query_mode: FTSQueryMode,
) -> dict[str, Any]:
    """Create a content-free comparability contract for paired B2A evaluator runs."""

    parent_values, control_values = tuple(parents), tuple(controls)
    inventory = "\n".join(
        f"{number}:{','.join(sources)}" for number, sources in sorted(oracle.items())
    )
    return {
        "fts_query_mode": fts_query_mode.value,
        "case_ids": {
            "A": [case.case_id for case in parent_values],
            "B": [case.case_id for case in control_values if case.set_name == "B"],
            "C": [case.case_id for case in control_values if case.set_name == "C"],
        },
        "frozen_expected_denominator": denominator,
        "pool_sizes": list(POOL_SIZES),
        "evaluation_source_scope": list(SOURCE_IDS),
        "semantic_model_id": SEMANTIC_PROFILE_ID,
        "eligible_expected_inventory_fingerprint": hashlib.sha256(
            inventory.encode("utf-8")
        ).hexdigest(),
    }


def _assert_candidates(candidates: Iterable[CollapsedDocumentCandidate]) -> None:
    """Ensure collapse/fusion did not sever identity, source, or strict provenance."""
    seen: dict[UUID, Any] = {}
    for candidate in candidates:
        identity = candidate.identity
        if identity.document_version_id in seen and seen[identity.document_version_id] != identity:
            raise EvaluationError("document version identity/provenance mismatch")
        if (
            identity.source_id.value not in SOURCE_IDS
            or not identity.latest_ingested
            or identity.transport_trust_mode.value != "STRICT_TLS"
        ):
            raise EvaluationError("candidate outside evaluator identity envelope")
        if candidate.representative.identity != identity:
            raise EvaluationError("candidate representative identity mismatch")
        seen[identity.document_version_id] = identity


async def evaluate_query(
    case: ParentCase | ControlCase,
    *,
    reader: Any,
    embedder: Any,
    explain: bool,
    fts_query_mode: FTSQueryMode = FTSQueryMode.NATURAL,
) -> QueryResult:
    """Embed exactly once, read three top-50 lanes once, then rank every pool in memory."""
    started = perf_counter()
    embed_started = perf_counter()
    batch = await embedder.embed_query(case.question)
    embedding_ms = (perf_counter() - embed_started) * 1000
    if (
        not isinstance(batch, SemanticEmbeddingBatch)
        or len(batch.vectors) != 1
        or len(batch.vectors[0]) != SEMANTIC_DIMENSION
    ):
        raise EvaluationError("embedder did not return exactly one E5 vector")
    reader_started = perf_counter()
    read = await reader.read_candidates(
        case.question,
        SOURCE_IDS,
        tuple(batch.vectors[0]),
        50,
        explain=explain,
        fts_query_mode=fts_query_mode,
    )
    reader_ms = (perf_counter() - reader_started) * 1000
    ranking_started = perf_counter()
    merged = merge_chunk_candidates(
        candidate for lane in RetrievalLane for candidate in read.lane_candidates.get(lane, ())
    )
    diagnostic_lanes = tuple(
        build_lane_document_pool(merged.candidates, lane, 50) for lane in RetrievalLane
    )
    diagnostic = fused_diagnostic_top50(diagnostic_lanes)
    _assert_candidates(diagnostic.candidates)
    pools: dict[int, Any] = {}
    finals: dict[int, tuple[CollapsedDocumentCandidate, ...]] = {}
    for size in POOL_SIZES:
        lanes = tuple(
            build_lane_document_pool(merged.candidates, lane, size) for lane in RetrievalLane
        )
        pool = with_lane_unique_contributions(lanes, size)
        _assert_candidates(pool.candidates)
        final = select_final_top3(pool)
        _assert_candidates(final)
        if len(_ids(final)) != len(final):
            raise EvaluationError("duplicate final document version")
        pools[size], finals[size] = pool, final
    ranking_ms = (perf_counter() - ranking_started) * 1000
    data_lane_elapsed_ms = sum(float(metric.elapsed_ms) for metric in read.lane_metrics)
    return QueryResult(
        case.case_id,
        "A" if isinstance(case, ParentCase) else case.set_name,
        "PARENT" if isinstance(case, ParentCase) else case.category,
        round(embedding_ms, 3),
        round(reader_ms, 3),
        round(ranking_ms, 3),
        round(
            embedding_ms + data_lane_elapsed_ms + read.fts_preparation_elapsed_ms + ranking_ms,
            3,
        ),
        round((perf_counter() - started) * 1000, 3),
        round(read.transaction_elapsed_ms, 3),
        read.data_query_count,
        read.explain_query_count,
        read.query_count,
        read.lane_candidates,
        read.lane_metrics,
        diagnostic,
        pools,
        finals,
        read.requested_fts_query_mode,
        read.applied_fts_query_mode,
        read.fts_preparation_query_count,
        round(read.fts_preparation_elapsed_ms, 3),
        read.bounded_or_selected_lexeme_count,
        read.bounded_or_source_lexeme_count,
        read.bounded_or_truncated,
        read.bounded_or_empty_query,
        read.bounded_or_natural_fallback_used,
    )


async def evaluate_sequentially(
    cases: Iterable[ParentCase | ControlCase],
    *,
    reader: Any,
    embedder: Any,
    explain_a: bool = True,
    fts_query_mode: FTSQueryMode = FTSQueryMode.NATURAL,
) -> tuple[QueryResult, ...]:
    """No concurrent reader/model use: the evaluator is intentionally sequential."""
    output = []
    for case in cases:
        output.append(
            await evaluate_query(
                case,
                reader=reader,
                embedder=embedder,
                explain=explain_a and isinstance(case, ParentCase),
                fts_query_mode=fts_query_mode,
            )
        )
    return tuple(output)


async def inventory_oracle(
    session_factory: Any, expected_numbers: Iterable[str]
) -> dict[str, tuple[str, ...]]:
    """Post-score-only expected inventory; duplicate occurrence remains in each case denominator."""
    expected = tuple(sorted(set(expected_numbers)))
    if not expected:
        return {}
    latest = (
        select(func.max(DocumentVersion.version_number))
        .where(DocumentVersion.document_id == LegalDocument.id)
        .correlate(LegalDocument)
        .scalar_subquery()
    )
    strict = (
        select(SourceProvenanceRecord.id)
        .where(
            SourceProvenanceRecord.document_version_id == DocumentVersion.id,
            SourceProvenanceRecord.source_id == LegalDocument.source_id,
            SourceProvenanceRecord.transport_trust_mode == "STRICT_TLS",
            SourceProvenanceRecord.provenance_type.in_(("source_fetch", "manual_snapshot")),
        )
        .correlate(DocumentVersion, LegalDocument)
        .exists()
    )
    statement = (
        select(DocumentVersion.document_number_normalized, LegalDocument.source_id)
        .select_from(DocumentVersion)
        .join(LegalDocument)
        .join(DocumentChunk)
        .join(
            ChunkEmbedding,
            and_(
                ChunkEmbedding.document_chunk_id == DocumentChunk.id,
                ChunkEmbedding.embedding_model_id == SEMANTIC_PROFILE_ID,
                ChunkEmbedding.embedding_kind == "semantic",
                ChunkEmbedding.dimension == SEMANTIC_DIMENSION,
            ),
        )
        .join(CorpusCatalogEntry, CorpusCatalogEntry.document_version_id == DocumentVersion.id)
        .where(
            DocumentVersion.document_number_normalized.in_(expected),
            LegalDocument.source_id.in_(SOURCE_IDS),
            DocumentVersion.version_number == latest,
            strict,
            CorpusCatalogEntry.processing_status == "INDEXED",
        )
        .distinct()
    )
    async with session_factory() as session:
        rows = (await session.execute(statement)).all()
    found: dict[str, set[str]] = defaultdict(set)
    for number, source in rows:
        if number:
            found[str(number)].add(str(source))
    return {number: tuple(sorted(sources)) for number, sources in found.items()}


def set_a_rows(
    parents: Iterable[ParentCase],
    results: Iterable[QueryResult],
    oracle: dict[str, tuple[str, ...]],
) -> tuple[dict[str, Any], ...]:
    """Score expected identities only after all candidate selection has completed."""
    by_id = {result.case_id: result for result in results}
    rows: list[dict[str, Any]] = []
    for parent in parents:
        result = by_id[parent.case_id]
        expected = tuple(
            number for number in parent.expected_numbers if number in oracle
        )  # occurrence-preserving
        semantic_raw = _raw_lane(result, RetrievalLane.SEMANTIC)
        semantic_collapsed = {
            size: _lane_pool(result, RetrievalLane.SEMANTIC, size) for size in (8, 20, 50)
        }
        semantic_top50_numbers = _expected_numbers(semantic_collapsed[50], expected)
        title_top50_numbers = _expected_numbers(
            _lane_pool(result, RetrievalLane.TITLE_FTS, 50), expected
        )
        content_top50_numbers = _expected_numbers(
            _lane_pool(result, RetrievalLane.CONTENT_FTS, 50), expected
        )
        for pool_size in POOL_SIZES:
            pool, final = result.pools[pool_size].candidates, result.finals[pool_size]
            candidate_numbers = {
                number for candidate in pool if (number := _identity_number(candidate))
            }
            final_numbers = {
                number for candidate in final if (number := _identity_number(candidate))
            }
            lane_hits = {
                lane.value: len(
                    set(expected)
                    & {
                        _identity_number(item)
                        for item in build_lane_document_pool(
                            merge_chunk_candidates(
                                candidate
                                for values in result.lane_candidates.values()
                                for candidate in values
                            ).candidates,
                            lane,
                            50,
                        ).candidates
                        if _identity_number(item)
                    }
                )
                for lane in RetrievalLane
            }
            unique_expected = {}
            for contribution in result.pools[pool_size].lane_unique_contributions:
                unique_expected[contribution.lane.value] = len(
                    set(contribution.document_version_ids)
                    & {
                        candidate.identity.document_version_id
                        for candidate in pool
                        if _identity_number(candidate) in expected
                    }
                )
            rows.append(
                {
                    "case_id": parent.case_id,
                    "pool_size": pool_size,
                    "expected_indexed_numbers": expected,
                    "inventory_blockers": tuple(
                        number for number in parent.expected_numbers if number not in oracle
                    ),
                    "candidate_hits": len(set(expected) & candidate_numbers),
                    "final_hits": len(set(expected) & final_numbers),
                    "final_case_hit": bool(set(expected) & final_numbers),
                    "lane_hits_at_50": lane_hits,
                    "fused_diagnostic_hits_at_50": len(
                        set(expected)
                        & {
                            _identity_number(item)
                            for item in result.diagnostic.candidates
                            if _identity_number(item)
                        }
                    ),
                    "direct_title_hit_identities": lane_hits[RetrievalLane.TITLE_FTS.value],
                    "nonexpected_candidate_count": _nonexpected_identity_count(pool, expected),
                    "candidate_identity_count": len(_ids(pool)),
                    "nonexpected_final_count": _nonexpected_identity_count(final, expected),
                    "final_identity_count": len(_ids(final)),
                    "source_coverage": len(
                        {candidate.identity.source_id.value for candidate in pool}
                        & {source for number in expected for source in oracle[number]}
                    ),
                    "source_denominator": len(
                        {source for number in expected for source in oracle[number]}
                    ),
                    "candidate_to_final_loss": len(set(expected) & candidate_numbers)
                    - len(set(expected) & final_numbers),
                    "pre_collapse_chunks": sum(
                        len(values) for values in result.lane_candidates.values()
                    ),
                    "post_collapse_versions": len(pool),
                    "duplicate_final_versions": len(final) - len(_ids(final)),
                    "unique_expected_contribution": unique_expected,
                    "generic_lane_unique": {
                        item.lane.value: item.unique_count
                        for item in result.pools[pool_size].lane_unique_contributions
                    },
                    "embedding_ms": result.embedding_ms,
                    "reader_ms": result.reader_ms,
                    "ranking_ms": result.ranking_ms,
                    "end_to_end_ms": result.end_to_end_ms,
                    "query_count": result.query_count,
                    "data_query_count": result.data_query_count,
                    "explain_query_count": result.explain_query_count,
                    "retrieval_eval_ms": result.retrieval_eval_ms,
                    "semantic_raw_chunk_hits": {
                        size: len(_expected_numbers(semantic_raw[:size], expected))
                        for size in (8, 20, 50)
                    },
                    "semantic_document_collapse_hits": {
                        size: len(_expected_numbers(semantic_collapsed[size], expected))
                        for size in (8, 20, 50)
                    },
                    "semantic_collapse_lift": {
                        size: len(_expected_numbers(semantic_collapsed[size], expected))
                        - len(_expected_numbers(semantic_raw[:size], expected))
                        for size in (8, 20, 50)
                    },
                    "title_unique_rescue_numbers": tuple(
                        sorted(title_top50_numbers - semantic_top50_numbers)
                    ),
                    "lexical_unique_rescue_numbers": tuple(
                        sorted(content_top50_numbers - semantic_top50_numbers)
                    ),
                    "direct_title_hit_total": len(title_top50_numbers),
                    "lexical_expected_hit_total": len(content_top50_numbers),
                    "transaction_elapsed_ms": result.transaction_elapsed_ms,
                    "lane_metrics": [
                        metric.model_dump(mode="json") for metric in result.lane_metrics
                    ],
                }
            )
    return tuple(rows)


def _percentile(values: Iterable[float], percent: int = 95) -> float:
    data = sorted(values)
    if not data:
        return 0.0
    return round(data[max(0, (len(data) * percent + 99) // 100 - 1)], 3)


def semantic_reference_summary(
    parents: Iterable[ParentCase],
    results: Iterable[QueryResult],
    oracle: dict[str, tuple[str, ...]],
) -> PoolReferenceSummary:
    """Measure the natural semantic-only diagnostic reference from current read results.

    Semantic collapse is in-memory and too small to isolate from the fused ranking timer;
    latency therefore uses embedding plus the semantic lane's measured data elapsed only.
    """
    by_id = {result.case_id: result for result in results}
    hits = nonexpected = identities = 0
    latency: list[float] = []
    for parent in parents:
        result = by_id[parent.case_id]
        expected = tuple(number for number in parent.expected_numbers if number in oracle)
        pool = _lane_pool(result, RetrievalLane.SEMANTIC, 50)
        hits += len(_expected_numbers(pool, expected))
        nonexpected += _nonexpected_identity_count(pool, expected)
        identities += len(_ids(pool))
        semantic_elapsed = next(
            (
                float(metric.elapsed_ms)
                for metric in result.lane_metrics
                if metric.lane is RetrievalLane.SEMANTIC
            ),
            0.0,
        )
        latency.append(result.embedding_ms + semantic_elapsed)
    failures = 0
    for result in by_id.values():
        if result.set_name != "C":
            continue
        final = _semantic_final(result)
        if len(_ids(final)) != len(final) or any(
            candidate.identity.source_id.value not in SOURCE_IDS
            or not candidate.identity.latest_ingested
            or candidate.identity.transport_trust_mode.value != "STRICT_TLS"
            for candidate in final
        ):
            failures += 1
    return PoolReferenceSummary(
        candidate_identity_count=hits,
        nonexpected_candidate_rate=nonexpected / identities if identities else 0.0,
        p95_latency_ms=_percentile(latency),
        query_count=1,
        set_c_failure_count=failures,
    )


def set_b_rows(results: Iterable[QueryResult]) -> tuple[dict[str, Any], ...]:
    values = tuple(results)
    parents = {item.case_id: item for item in values if item.set_name == "A"}
    rows = []
    for item in values:
        if item.set_name != "B":
            continue
        # B IDs are bound in the M2 parser, retained externally by execution order map.
        parent_id = item.case_id.split("-")[1] if "-" in item.case_id else ""
        parent = parents.get(parent_id)
        for size in POOL_SIZES:
            left = (
                {
                    _identity_number(candidate)
                    for candidate in parent.finals[size]
                    if _identity_number(candidate)
                }
                if parent
                else set()
            )
            right = {
                _identity_number(candidate)
                for candidate in item.finals[size]
                if _identity_number(candidate)
            }
            semantic_left = (
                {
                    _identity_number(candidate)
                    for candidate in _semantic_final(parent)
                    if _identity_number(candidate)
                }
                if parent
                else set()
            )
            semantic_right = {
                _identity_number(candidate)
                for candidate in _semantic_final(item)
                if _identity_number(candidate)
            }
            jaccard = len(left & right) / len(left | right) if left or right else None
            rows.append(
                {
                    "case_id": item.case_id,
                    "parent_case_id": parent_id,
                    "pool_size": size,
                    "jaccard": jaccard,
                    "semantic_reference_jaccard": (
                        len(semantic_left & semantic_right) / len(semantic_left | semantic_right)
                        if semantic_left or semantic_right
                        else None
                    ),
                    "parent_evidence": bool(left),
                    "paraphrase_evidence": bool(right),
                    "decision_consistent": bool(left) == bool(right),
                    "semantic_decision_consistent": bool(semantic_left) == bool(semantic_right),
                }
            )
    return tuple(rows)


def set_c_rows(results: Iterable[QueryResult]) -> tuple[dict[str, Any], ...]:
    rows = []
    for item in results:
        if item.set_name != "C":
            continue
        for size in POOL_SIZES:
            final = item.finals[size]
            failures = []
            if len(_ids(final)) != len(final):
                failures.append("DUPLICATE_VERSION")
            if any(
                candidate.identity.source_id.value not in SOURCE_IDS
                or not candidate.identity.latest_ingested
                or candidate.identity.transport_trust_mode.value != "STRICT_TLS"
                for candidate in final
            ):
                failures.append("IDENTITY_ENVELOPE")
            if item.data_query_count > 12:
                failures.append("QUERY_COUNT")
            rows.append(
                {
                    "case_id": item.case_id,
                    "category": item.category,
                    "pool_size": size,
                    "evidence_selected": bool(final),
                    "data_query_count": item.data_query_count,
                    "explain_query_count": item.explain_query_count,
                    "query_count": item.query_count,
                    "invariant_failures": tuple(failures),
                }
            )
    return tuple(rows)


def attach_citation_invariants(
    c_rows: Iterable[dict[str, Any]], citation_rows: Iterable[dict[str, Any]]
) -> tuple[dict[str, Any], ...]:
    """Make Set C cleanup/resolve failures visible in the corresponding invariant row."""
    index = {
        (row["case_id"], row["pool_size"]): row for row in citation_rows if row["set_name"] == "C"
    }
    output = []
    for row in c_rows:
        citation = index.get((row["case_id"], row["pool_size"]))
        failures = list(row["invariant_failures"])
        if citation and not (
            citation["resolvable"]
            and citation["cleanup"] == "COMPLETED"
            and citation["global_counts_match"]
        ):
            failures.append("CITATION_INVARIANT")
        output.append({**row, "invariant_failures": tuple(failures)})
    return tuple(output)


def citation_score_mapping(
    candidate: CollapsedDocumentCandidate,
) -> tuple[float | None, float | None]:
    """Map only content FTS and semantic lane scores to the persisted citation schema."""
    lexical = semantic = None
    for aggregate in candidate.lane_aggregates:
        if aggregate.lane is RetrievalLane.CONTENT_FTS:
            lexical = aggregate.best_score
        elif aggregate.lane is RetrievalLane.SEMANTIC:
            semantic = aggregate.best_score
    if semantic is None:
        semantic = candidate.representative.supporting_semantic_score
    if lexical is None and semantic is None:
        raise EvaluationError("citation candidate has neither lexical nor semantic score")
    return lexical, semantic


async def _global_counts(session_factory: Any) -> tuple[int, int]:
    async with session_factory() as session:
        return (
            int(
                (await session.execute(select(func.count()).select_from(RetrievalRun))).scalar_one()
            ),
            int(
                (
                    await session.execute(select(func.count()).select_from(CitationRecord))
                ).scalar_one()
            ),
        )


async def cleanup_run(session_factory: Any, run_id: UUID) -> str:
    try:
        async with session_factory.begin() as session:
            await session.execute(
                delete(CitationRecord).where(CitationRecord.retrieval_run_id == run_id)
            )
            await session.execute(delete(RetrievalRun).where(RetrievalRun.id == run_id))
        return "COMPLETED"
    except Exception:
        return "FAILED_SAFE"


async def citation_smoke(
    session_factory: Any,
    resolver: Any,
    candidates: tuple[CollapsedDocumentCandidate, ...],
    *,
    strategy_version: str,
    case_id: str = "",
    set_name: str = "",
    pool_size: int = 0,
) -> dict[str, Any]:
    """Persist evaluator-owned evidence, resolve it, and remove exactly that run in finally."""
    if not 1 <= len(strategy_version) <= 64:
        raise EvaluationError("evaluation strategy_version must be 1..64 characters")
    before = await _global_counts(session_factory)
    run_id = uuid4()
    citation_ids: list[UUID] = []
    cleanup = "NOT_STARTED"
    failure = ""
    run_created = False
    try:
        async with session_factory.begin() as session:
            run = RetrievalRun(
                id=run_id,
                strategy="evaluation_only_phase_b",
                strategy_version=strategy_version,
                scope="LATEST_INGESTED",
                trust_scope="STRICT_TLS_ONLY",
                query_max_chars=1,
                top_k=3,
                candidate_count=len(candidates),
                citation_count=len(candidates),
                evidence_decision="EVIDENCE_AVAILABLE" if candidates else "NO_RESULTS",
                evidence_reason="EVALUATION_ONLY",
            )
            session.add(run)
            for rank, candidate in enumerate(candidates, 1):
                lexical, semantic = citation_score_mapping(candidate)
                citation_id = uuid4()
                citation_ids.append(citation_id)
                session.add(
                    CitationRecord(
                        id=citation_id,
                        retrieval_run_id=run_id,
                        document_chunk_id=candidate.representative.chunk_id,
                        source_provenance_record_id=candidate.identity.provenance_record_id,
                        rank=rank,
                        lexical_score=lexical,
                        semantic_score=semantic,
                        reranker_score=None,
                    )
                )
            await session.flush()
            run_created = True
        for citation_id, candidate in zip(citation_ids, candidates, strict=True):
            resolved = await resolver.resolve(citation_id, run_id)
            if (
                resolved.document_id != candidate.identity.document_id
                or resolved.document_version_id != candidate.identity.document_version_id
                or resolved.source_id != candidate.identity.source_id.value
                or resolved.source_provenance_record_id != candidate.identity.provenance_record_id
            ):
                raise EvaluationError(
                    "resolved citation provenance differs from candidate identity"
                )
    except Exception as error:
        failure = type(error).__name__
    finally:
        cleanup = await cleanup_run(session_factory, run_id)
    after = await _global_counts(session_factory)
    if before != after:
        failure = failure or "GLOBAL_COUNT_MISMATCH"
    return {
        "case_id": case_id,
        "set_name": set_name,
        "pool_size": pool_size,
        "run_created": run_created,
        "citation_count": len(citation_ids),
        "resolvable": not failure,
        "failure": failure,
        "cleanup": cleanup,
        "global_counts_match": before == after,
    }


def pool_matrix(
    a_rows: Iterable[dict[str, Any]],
    c_rows: Iterable[dict[str, Any]],
    results: Iterable[QueryResult],
    semantic_reference: PoolReferenceSummary,
) -> tuple[tuple[PoolMeasurementSummary, ...], Any]:
    matrix = []
    a_rows, c_rows, results = tuple(a_rows), tuple(c_rows), tuple(results)
    for size in POOL_SIZES:
        rows = [row for row in a_rows if row["pool_size"] == size]
        expected_hits = sum(row["candidate_hits"] for row in rows)
        nonexpected = sum(row["nonexpected_candidate_count"] for row in rows)
        candidates = sum(row["candidate_identity_count"] for row in rows)
        c_failures = sum(
            len(row["invariant_failures"]) for row in c_rows if row["pool_size"] == size
        )
        matrix.append(
            PoolMeasurementSummary(
                pool_size=size,
                candidate_identity_count=expected_hits,
                nonexpected_candidate_rate=nonexpected / candidates if candidates else 0.0,
                p95_latency_ms=_percentile(
                    item.retrieval_eval_ms for item in results if item.set_name == "A"
                ),
                query_count=max((item.data_query_count for item in results), default=0),
                set_c_failure_count=c_failures,
            )
        )
    measurements = tuple(matrix)
    return measurements, select_pareto_pool(semantic_reference, measurements)


def gate_b(
    selection: Any,
    measurements: Iterable[PoolMeasurementSummary],
    a_rows: Iterable[dict[str, Any]],
    citation_rows: Iterable[dict[str, Any]],
) -> dict[str, Any]:
    measurement_by_size = {item.pool_size: item for item in measurements}
    if selection.status is not PoolSelectionStatus.SELECTED:
        return {
            "status": "NO_GO_PHASE_B",
            "pass": False,
            "selected_pool": None,
            "reason": "NO_SELECTION",
        }
    size = selection.selected_pool_size
    assert size is not None
    selected = measurement_by_size[size]
    contributions = [
        row["unique_expected_contribution"] for row in a_rows if row["pool_size"] == size
    ]
    retained = all(
        sum(values.get(lane.value, 0) for values in contributions) >= 1 for lane in RetrievalLane
    )
    citation_ok = all(
        row["resolvable"] and row["cleanup"] == "COMPLETED" and row["global_counts_match"]
        for row in citation_rows
    )
    passed = (
        selected.candidate_identity_count >= 27
        and retained
        and selected.set_c_failure_count == 0
        and selected.p95_latency_ms <= 2450
        and selected.query_count <= 12
        and citation_ok
    )
    return {
        "status": "PASS" if passed else "NO_GO_PHASE_B",
        "pass": passed,
        "selected_pool": size,
        "broad_expected_hits": selected.candidate_identity_count,
        "retained_lane_unique_contribution": retained,
        "citation_invariants": citation_ok,
    }


def _table(sheet: Any, header: tuple[str, ...], rows: Iterable[tuple[Any, ...]]) -> None:
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_reports(
    *,
    xlsx_path: Path,
    json_path: Path,
    markdown_path: Path,
    a_rows: tuple[dict[str, Any], ...],
    b_rows: tuple[dict[str, Any], ...],
    c_rows: tuple[dict[str, Any], ...],
    measurements: tuple[PoolMeasurementSummary, ...],
    selection: Any,
    gate: dict[str, Any],
    citation_rows: tuple[dict[str, Any], ...],
    semantic_reference: PoolReferenceSummary,
    partial_failure: str = "",
    fts_metadata: tuple[dict[str, Any], ...] = (),
    comparison_contract: dict[str, Any] | None = None,
    state_invariants: dict[str, object] | None = None,
) -> None:
    """Write all required report forms without questions, text, titles, URLs, or UUIDs."""
    for path in (xlsx_path, json_path, markdown_path):
        path.parent.mkdir(parents=True, exist_ok=True)
    b_with_evidence = [row for row in b_rows if row["jaccard"] is not None]
    b_semantic_with_evidence = [
        row for row in b_rows if row["semantic_reference_jaccard"] is not None
    ]
    first_pool_rows = [row for row in a_rows if row["pool_size"] == 8]
    recall = {
        "fused_diagnostic_top50_expected_hits": sum(
            row["fused_diagnostic_hits_at_50"] for row in first_pool_rows
        ),
        "fused_candidate_expected_hits": {
            str(size): sum(row["candidate_hits"] for row in a_rows if row["pool_size"] == size)
            for size in POOL_SIZES
        },
        "final_top3_expected_hits": {
            str(size): sum(row["final_hits"] for row in a_rows if row["pool_size"] == size)
            for size in POOL_SIZES
        },
    }
    collapse_lift = {
        str(size): {
            "raw_chunk_budget_hits": sum(
                row["semantic_raw_chunk_hits"][size] for row in first_pool_rows
            ),
            "document_collapse_hits": sum(
                row["semantic_document_collapse_hits"][size] for row in first_pool_rows
            ),
            "lift": sum(row["semantic_collapse_lift"][size] for row in first_pool_rows),
        }
        for size in (8, 20, 50)
    }
    rescue = {
        "direct_title_hit_total": sum(row["direct_title_hit_total"] for row in first_pool_rows),
        "lexical_expected_hit_total": sum(
            row["lexical_expected_hit_total"] for row in first_pool_rows
        ),
        "title_unique_rescue_count": sum(
            len(row["title_unique_rescue_numbers"]) for row in first_pool_rows
        ),
        "lexical_unique_rescue_count": sum(
            len(row["lexical_unique_rescue_numbers"]) for row in first_pool_rows
        ),
    }
    payload = {
        "report_schema_version": "QUALITY-RETRIEVAL-PHASE-B-1",
        "evaluation_scope": list(SOURCE_IDS),
        "production_registry_scope_mismatch": "EVALUATION_ONLY_SCOPE_VBQPPL_VNU_UEB; production registry rollout differs",
        "partial_failure": partial_failure,
        "semantic_top50_broad_reference": semantic_reference.model_dump(mode="json"),
        "semantic_reference_comparison": {
            "frozen_reference": "24/29",
            "current_expected_hit_count": semantic_reference.candidate_identity_count,
            "corpus_drift": semantic_reference.candidate_identity_count != 24,
        },
        "fused_diagnostic_top50": "reported separately from release pools",
        "recall_metrics": recall,
        "document_collapse_lift": collapse_lift,
        "title_lexical_rescue": rescue,
        "pool_measurements": [item.model_dump(mode="json") for item in measurements],
        "selection": selection.to_public_dict(),
        "gate_b": gate,
        "set_a": list(a_rows),
        "set_b": list(b_rows),
        "set_b_summary": {
            "mean_jaccard": round(
                sum(row["jaccard"] for row in b_with_evidence) / len(b_with_evidence), 6
            )
            if b_with_evidence
            else None,
            "evidence_consistency_rate": round(
                sum(row["parent_evidence"] == row["paraphrase_evidence"] for row in b_rows)
                / len(b_rows),
                6,
            )
            if b_rows
            else 0.0,
            "semantic_reference_mean_jaccard": round(
                sum(row["semantic_reference_jaccard"] for row in b_semantic_with_evidence)
                / len(b_semantic_with_evidence),
                6,
            )
            if b_semantic_with_evidence
            else None,
            "semantic_reference_decision_consistency_rate": round(
                sum(row["semantic_decision_consistent"] for row in b_rows) / len(b_rows), 6
            )
            if b_rows
            else 0.0,
        },
        "set_c": list(c_rows),
        "citation_invariants": list(citation_rows),
        "fts_read_metadata": list(fts_metadata),
        "phase_b2a_comparison_contract": comparison_contract,
        "phase_b2a_state_invariants": state_invariants,
        "later_milestones": "NOT_MEASURED_PHASE_B",
    }
    with json_path.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    book = openpyxl.Workbook()
    summary = book.active
    assert summary is not None
    summary.title = "Summary"
    _table(
        summary,
        ("Metric", "Value"),
        (
            ("Gate B", gate["status"]),
            ("Selected pool", gate["selected_pool"]),
            ("Scope", ",".join(SOURCE_IDS)),
            ("Semantic top50 broad reference", semantic_reference.candidate_identity_count),
            ("Fused diagnostic top50", "separate diagnostic"),
            ("Pool N fused candidate recall", "see Pool Matrix"),
            ("Final top3", "see Set A"),
            ("Semantic collapse lift @8", collapse_lift["8"]["lift"]),
            ("Title unique rescue", rescue["title_unique_rescue_count"]),
            ("Lexical unique rescue", rescue["lexical_unique_rescue_count"]),
            (
                "FTS query mode",
                comparison_contract["fts_query_mode"] if comparison_contract else "NATURAL",
            ),
            ("Later milestones", "NOT_MEASURED_PHASE_B"),
        ),
    )
    _table(
        book.create_sheet("Set A"),
        (
            "Case",
            "Pool",
            "Expected indexed",
            "Candidate hits",
            "Final top3 hits",
            "Final case hit",
            "Nonexpected candidate",
            "Duplicate final versions",
            "End-to-end ms",
            "Query count",
        ),
        (
            (
                row["case_id"],
                row["pool_size"],
                "; ".join(row["expected_indexed_numbers"]),
                row["candidate_hits"],
                row["final_hits"],
                row["final_case_hit"],
                row["nonexpected_candidate_count"],
                row["duplicate_final_versions"],
                row["end_to_end_ms"],
                row["query_count"],
            )
            for row in a_rows
        ),
    )
    _table(
        book.create_sheet("Pool Matrix"),
        (
            "Pool",
            "Expected candidate hits",
            "Nonexpected rate",
            "P95 ms",
            "Query count",
            "Set C failures",
        ),
        (
            (
                item.pool_size,
                item.candidate_identity_count,
                item.nonexpected_candidate_rate,
                item.p95_latency_ms,
                item.query_count,
                item.set_c_failure_count,
            )
            for item in measurements
        ),
    )
    _table(
        book.create_sheet("Lane Contributions"),
        (
            "Case",
            "Pool",
            "Generic semantic",
            "Generic content FTS",
            "Generic title FTS",
            "Expected semantic",
            "Expected content FTS",
            "Expected title FTS",
            "Title rescue",
            "Lexical rescue",
        ),
        (
            (
                row["case_id"],
                row["pool_size"],
                row["generic_lane_unique"].get("SEMANTIC", 0),
                row["generic_lane_unique"].get("CONTENT_FTS", 0),
                row["generic_lane_unique"].get("TITLE_FTS", 0),
                row["unique_expected_contribution"].get("SEMANTIC", 0),
                row["unique_expected_contribution"].get("CONTENT_FTS", 0),
                row["unique_expected_contribution"].get("TITLE_FTS", 0),
                "; ".join(row["title_unique_rescue_numbers"]),
                "; ".join(row["lexical_unique_rescue_numbers"]),
            )
            for row in a_rows
        ),
    )
    _table(
        book.create_sheet("Set B"),
        (
            "Case",
            "Parent",
            "Pool",
            "Pool Jaccard",
            "Semantic Jaccard",
            "Evidence consistent",
            "Decision consistent",
            "Semantic decision consistent",
        ),
        (
            (
                row["case_id"],
                row["parent_case_id"],
                row["pool_size"],
                row["jaccard"],
                row["semantic_reference_jaccard"],
                row["parent_evidence"] == row["paraphrase_evidence"],
                row["decision_consistent"],
                row["semantic_decision_consistent"],
            )
            for row in b_rows
        ),
    )
    _table(
        book.create_sheet("Set C"),
        (
            "Case",
            "Category",
            "Pool",
            "Evidence selected",
            "Data queries",
            "EXPLAIN queries",
            "Wall queries",
            "Invariant failures",
        ),
        (
            (
                row["case_id"],
                row["category"],
                row["pool_size"],
                row["evidence_selected"],
                row["data_query_count"],
                row["explain_query_count"],
                row["query_count"],
                ";".join(row["invariant_failures"]),
            )
            for row in c_rows
        ),
    )
    _table(
        book.create_sheet("Cost"),
        (
            "Case",
            "Embedding ms",
            "Reader ms",
            "Ranking ms",
            "Transaction ms",
            "Retrieval evaluation ms",
            "Wall end-to-end ms",
            "Data queries",
            "EXPLAIN queries",
            "Wall queries",
            "Lane rows/buffers",
        ),
        (
            (
                row["case_id"],
                row["embedding_ms"],
                row["reader_ms"],
                row["ranking_ms"],
                row["transaction_elapsed_ms"],
                row["retrieval_eval_ms"],
                row["end_to_end_ms"],
                row["data_query_count"],
                row["explain_query_count"],
                row["query_count"],
                "; ".join(
                    f"{metric['lane']}:{metric['rows_returned']} rows/{metric['buffers']}"
                    for metric in row["lane_metrics"]
                ),
            )
            for row in a_rows
            if row["pool_size"] == 8
        ),
    )
    _table(
        book.create_sheet("Citation Invariants"),
        (
            "Case",
            "Set",
            "Pool",
            "Run created",
            "Citation count",
            "Resolvable",
            "Cleanup",
            "Global counts match",
            "Failure",
        ),
        (
            (
                row["case_id"],
                row["set_name"],
                row["pool_size"],
                row["run_created"],
                row["citation_count"],
                row["resolvable"],
                row["cleanup"],
                row["global_counts_match"],
                row["failure"],
            )
            for row in citation_rows
        ),
    )
    book.save(xlsx_path)
    book.close()
    markdown_path.write_text(
        f"# Quality Retrieval Phase B\n\nGate B: **{gate['status']}**. Selected pool: `{gate['selected_pool']}`.\n\n| Pool | Expected candidate hits | Nonexpected identity rate | P95 retrieval-eval ms | Data queries | Set C failures |\n|---:|---:|---:|---:|---:|---:|\n"
        + "\n".join(
            f"| {item.pool_size} | {item.candidate_identity_count} | {item.nonexpected_candidate_rate:.3f} | {item.p95_latency_ms:.1f} | {item.query_count} | {item.set_c_failure_count} |"
            for item in measurements
        )
        + f"\n\nFTS query mode: `{comparison_contract['fts_query_mode'] if comparison_contract else 'NATURAL'}`. Collapse lift @8: `{collapse_lift['8']['lift']}`. Title rescue: `{rescue['title_unique_rescue_count']}`. Lexical rescue: `{rescue['lexical_unique_rescue_count']}`. Recall: `{recall}`. Semantic reference latency is embedding plus measured semantic-lane data elapsed; isolated semantic collapse timing is not separately measured. Analyzer/dynamic/rerank/insufficiency/full-text: **NOT_MEASURED_PHASE_B**. This is evaluation-only, not production behavior or a legal correctness claim.\n",
        encoding="utf-8",
    )


async def run(args: argparse.Namespace) -> int:
    engine: Any | None = None
    partial = ""
    all_results: tuple[QueryResult, ...] = ()
    citations: tuple[dict[str, Any], ...] = ()
    session_factory: Any | None = None
    state_before: dict[str, int] | None = None
    state_after: dict[str, int] | None = None
    state_flags: dict[str, object] | None = None
    try:
        parents, controls = parse_expert_workbook(args.expert_workbook), parse_m2_set(args.m2_set)
        engine = create_engine(Settings())  # type: ignore[call-arg]
        session_factory = create_session_factory(engine)
        state_before = await _state_counts(session_factory)
        state_flags = _runtime_state_flags()
        semantic = FastEmbedSemanticAdapter(
            SemanticSettings(model_path=args.model_path) if args.model_path else SemanticSettings()
        )
        all_results = await evaluate_sequentially(
            (*parents, *controls),
            reader=PostgresQualityCandidateReader(session_factory),
            embedder=semantic,
            explain_a=not args.no_explain,
            fts_query_mode=args.fts_query_mode,
        )
        oracle = await inventory_oracle(
            session_factory, (number for parent in parents for number in parent.expected_numbers)
        )
        denominator = sum(
            sum(number in oracle for number in parent.expected_numbers) for parent in parents
        )
        if denominator != EXPECTED_DENOMINATOR:
            raise EvaluationError(
                f"frozen inventory denominator {denominator} != {EXPECTED_DENOMINATOR}"
            )
        comparison_contract = phase_b2a_comparison_contract(
            parents, controls, oracle, denominator, args.fts_query_mode
        )
        a_rows = set_a_rows(parents, all_results, oracle)
        b_rows = set_b_rows(all_results)
        c_rows = set_c_rows(all_results)
        reference = semantic_reference_summary(parents, all_results, oracle)
        citation_values: list[dict[str, Any]] = []
        resolver = PostgresCitationResolver(session_factory)
        for result in all_results:
            if result.set_name not in {"A", "C"}:
                continue
            for size in POOL_SIZES:
                citation_values.append(
                    await citation_smoke(
                        session_factory,
                        resolver,
                        result.finals[size],
                        strategy_version=f"evaluation_only_phase_b_{result.set_name.lower()}_pool_{size}",
                        case_id=result.case_id,
                        set_name=result.set_name,
                        pool_size=size,
                    )
                )
        citations = tuple(citation_values)
        c_rows = attach_citation_invariants(c_rows, citations)
        state_after = await _state_counts(session_factory)
        state_invariants = _state_invariants(state_before, state_after, state_flags)
        measurements, selection = pool_matrix(a_rows, c_rows, all_results, reference)
        gate = gate_b(selection, measurements, a_rows, citations)
        write_reports(
            xlsx_path=args.output,
            json_path=args.json_output,
            markdown_path=args.markdown_output,
            a_rows=a_rows,
            b_rows=b_rows,
            c_rows=c_rows,
            measurements=measurements,
            selection=selection,
            gate=gate,
            citation_rows=citations,
            semantic_reference=reference,
            fts_metadata=fts_read_metadata(all_results),
            comparison_contract=comparison_contract,
            state_invariants=state_invariants,
        )
        print(
            json.dumps(
                {
                    "status": "ok",
                    "cases": len(all_results),
                    "pools": list(POOL_SIZES),
                    "gate": gate["status"],
                }
            )
        )
        return 0
    except Exception as error:
        partial = type(error).__name__
        if session_factory is not None and state_before is not None:
            try:
                state_after = await _state_counts(session_factory)
                state_flags = state_flags or _runtime_state_flags()
            except Exception:
                pass
        # A safe partial report is intentionally minimal and contains no retained query/candidate material.
        try:
            reference = PoolReferenceSummary(
                candidate_identity_count=0,
                nonexpected_candidate_rate=0,
                p95_latency_ms=0,
                query_count=0,
                set_c_failure_count=0,
            )
            selection = select_pareto_pool(
                reference,
                tuple(
                    PoolMeasurementSummary(
                        pool_size=size,
                        candidate_identity_count=0,
                        nonexpected_candidate_rate=0,
                        p95_latency_ms=0,
                        query_count=0,
                        set_c_failure_count=1,
                    )
                    for size in POOL_SIZES
                ),
            )
            write_reports(
                xlsx_path=args.output,
                json_path=args.json_output,
                markdown_path=args.markdown_output,
                a_rows=(),
                b_rows=(),
                c_rows=(),
                measurements=(),
                selection=selection,
                gate={
                    "status": "NO_GO_PHASE_B",
                    "pass": False,
                    "selected_pool": None,
                    "reason": partial,
                },
                citation_rows=citations,
                semantic_reference=reference,
                partial_failure=partial,
                state_invariants=(
                    _state_invariants(state_before, state_after, state_flags)
                    if state_before is not None and state_after is not None and state_flags is not None
                    else None
                ),
            )
        except Exception:
            pass
        print(
            json.dumps(
                {
                    "status": "failed",
                    "cases": len(all_results),
                    "pools": list(POOL_SIZES),
                    "gate": "NO_GO_PHASE_B",
                }
            )
        )
        return 2
    finally:
        if engine is not None:
            await engine.dispose()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run offline Phase B quality-retrieval candidate evaluation."
    )
    parser.add_argument("--expert-workbook", type=Path, default=DEFAULT_EXPERT)
    parser.add_argument("--m2-set", type=Path, default=DEFAULT_SET)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--json-output", type=Path)
    parser.add_argument("--markdown-output", type=Path)
    parser.add_argument("--model-path", type=Path)
    parser.add_argument("--no-explain", action="store_true")
    parser.add_argument(
        "--fts-query-mode",
        type=FTSQueryMode,
        choices=tuple(FTSQueryMode),
        default=FTSQueryMode.NATURAL,
        help="Evaluation-only FTS query construction mode (default: NATURAL).",
    )
    parser.add_argument(
        "--output-stem",
        type=Path,
        help="Write .xlsx/.json/.md artifacts from this explicit stem.",
    )
    return parser


def resolve_output_paths(args: argparse.Namespace) -> None:
    """Resolve optional stems without allowing a BOUNDED_OR run to replace frozen output."""

    if args.output_stem:
        if args.output_stem.suffix:
            raise ValueError("--output-stem must not include a suffix")
        args.output = args.output_stem.with_suffix(".xlsx")
        if args.json_output is None:
            args.json_output = args.output_stem.with_suffix(".json")
        if args.markdown_output is None:
            args.markdown_output = args.output_stem.with_suffix(".md")
    args.json_output = args.json_output or args.output.with_suffix(".json")
    args.markdown_output = args.markdown_output or args.output.with_suffix(".md")
    frozen_paths = {
        DEFAULT_OUTPUT,
        DEFAULT_OUTPUT.with_suffix(".json"),
        DEFAULT_OUTPUT.with_suffix(".md"),
    }
    paths = {args.output, args.json_output, args.markdown_output}
    legacy_default = paths == frozen_paths
    if args.fts_query_mode is FTSQueryMode.BOUNDED_OR and legacy_default:
        raise ValueError("BOUNDED_OR requires --output-stem or a mode-specific --output path")
    if not legacy_default and paths & frozen_paths:
        raise ValueError("mode-specific artifacts must not overwrite frozen Phase-B reports")


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        resolve_output_paths(args)
    except ValueError as error:
        parser.error(str(error))
    return asyncio.run(run(args))


if __name__ == "__main__":
    raise SystemExit(main())
