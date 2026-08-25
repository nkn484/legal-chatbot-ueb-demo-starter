"""Bounded, direct stress runner for the UEB legal-chatbot demo.

This runner deliberately bypasses all channel/runtime adapters.  Workbook grading
metadata is parsed only for post-result scoring and is never supplied to chat or a
provider request.  It emits no question, answer, prompt, or credential text.
"""
# ruff: noqa: E501

from __future__ import annotations

import argparse
import asyncio
import re
import sys
from collections import Counter
from collections.abc import Callable, Iterable
from dataclasses import dataclass, field
from pathlib import Path
from time import perf_counter
from typing import Any

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import delete, select

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "src") not in sys.path:
    sys.path.insert(0, str(ROOT / "src"))

from legal_chatbot.chat import ChatRequest, ChatSettings, GroundedChatService  # noqa: E402
from legal_chatbot.chat.errors import ChatError, ProviderOutputFailureClass  # noqa: E402
from legal_chatbot.chat.parser import StrictProviderJsonParser  # noqa: E402
from legal_chatbot.chat.planner_models import QueryPlannerOutcome  # noqa: E402
from legal_chatbot.chat.planner_service import LLMQueryPlanner  # noqa: E402
from legal_chatbot.core.config import Settings  # noqa: E402
from legal_chatbot.db.session import create_engine, create_session_factory  # noqa: E402
from legal_chatbot.documents.canonical_anchor_resolver import (  # noqa: E402
    PostgresCanonicalAnchorResolver,
)
from legal_chatbot.documents.citation_resolver import PostgresCitationResolver  # noqa: E402
from legal_chatbot.documents.grounding_evidence import (  # noqa: E402
    PostgresGroundingEvidenceAdapter,
)
from legal_chatbot.documents.hybrid_retrieval_repository import (  # noqa: E402
    PostgresHybridRetrievalRepository,
)
from legal_chatbot.documents.metadata_repair_repository import (  # noqa: E402
    MetadataRepairDiagnostics,
    PostgresMetadataRepairRetrievalRepository,
)
from legal_chatbot.documents.orm import (  # noqa: E402
    CitationRecord,
    CorpusCatalogEntry,
    DocumentChunk,
    DocumentVersion,
    LegalDocument,
    RetrievalRun,
    SourceProvenanceRecord,
)
from legal_chatbot.documents.quality_candidate_reader import (  # noqa: E402
    PostgresQualityCandidateReader,
)
from legal_chatbot.documents.quality_retrieval_pipeline import (  # noqa: E402
    LegalQualityCandidatePipeline,
)
from legal_chatbot.documents.quality_retrieval_repository import (  # noqa: E402
    PostgresQualityRetrievalRepository,
)
from legal_chatbot.documents.reranked_semantic_repository import (  # noqa: E402
    PostgresRerankedSemanticRepository,
    RerankedRetrievalDiagnostics,
)
from legal_chatbot.documents.retrieval_repository import (  # noqa: E402
    PostgresLexicalRetrievalRepository,
)
from legal_chatbot.providers.config import ProviderSettings  # noqa: E402
from legal_chatbot.providers.models import (  # noqa: E402
    GenerationRequest,
    GenerationResult,
    ProviderHealth,
    ProviderHealthStatus,
)
from legal_chatbot.providers.registry import create_provider  # noqa: E402
from legal_chatbot.reranking.config import RerankerSettings  # noqa: E402
from legal_chatbot.reranking.fastembed_adapter import FastEmbedRerankerAdapter  # noqa: E402
from legal_chatbot.retrieval.models import (  # noqa: E402
    coerce_transport_trust_mode,
    evidence_trust_label_for,
)
from legal_chatbot.retrieval.quality_repair.analyzer import LegalQuestionAnalyzer  # noqa: E402
from legal_chatbot.retrieval.quality_repair.models import SourceId  # noqa: E402
from legal_chatbot.retrieval.quality_repair.strategy import materialize_strategy  # noqa: E402
from legal_chatbot.retrieval.service import RetrievalService  # noqa: E402
from legal_chatbot.semantic.config import SemanticSettings  # noqa: E402
from legal_chatbot.semantic.constants import SEMANTIC_PROFILE_ID  # noqa: E402
from legal_chatbot.semantic.fastembed_adapter import FastEmbedSemanticAdapter  # noqa: E402
from legal_chatbot.sources.models import ProvenanceType  # noqa: E402

DEFAULT_INPUT = Path("docs/Stress_test_Legal_Chatbot_UEB_10_cau.xlsx")
DEFAULT_OUTPUT = Path("docs/Stress_test_Legal_Chatbot_UEB_10_cau_Ket_qua.xlsx")
SOURCES = ("VBQPPL", "VNU", "UEB")
CASE_IDS = tuple(f"Q{number:02d}" for number in range(1, 11))
QUALITY_EVALUATION_PROFILES = frozenset(
    {
        "quality_retrieval_document_collapse_v1",
        "quality_retrieval_hybrid_v1",
        "quality_retrieval_analyzer_protected_v1",
        "quality_retrieval_dynamic_evidence_v1",
        "quality_retrieval_evidence_repair_v1",
    }
)


@dataclass(frozen=True)
class StressCase:
    """Question input plus separately retained, post-result-only grading metadata."""

    case_id: str
    question: str
    topic: str = ""
    expected_sources: tuple[str, ...] = ()
    expected_documents: tuple[str, ...] = ()
    minimum: str = ""
    pass_condition: str = ""


@dataclass(frozen=True)
class CitationView:
    source_id: str
    document_number: str | None
    title: str | None
    url: str | None
    provenance_label: str


@dataclass
class CallRow:
    lane: str
    case_id: str
    concurrency: int
    round_number: int
    latency_ms: float
    outcome: str
    reason: str
    citation_count: int
    source_count: int
    error_code: str
    answer: str = ""
    citations: tuple[CitationView, ...] = ()
    provider: str = ""
    model: str = ""
    request_id_present: bool = False
    retrieval_run_id: Any | None = field(default=None, repr=False)
    retrieval_run_present: bool = False
    retrieval_decision: str = ""
    retrieval_reason: str = ""
    retrieval_strategy: str = ""
    provider_output_class: str = "NOT_APPLICABLE"


@dataclass(frozen=True)
class ApiRow:
    endpoint: str
    latency_ms: float
    error_code: str


@dataclass(frozen=True)
class CoverageRow:
    case_id: str
    expected_sources: tuple[str, ...]
    retrieved_sources: tuple[str, ...]
    source_coverage_percent: float
    expected_documents: tuple[str, ...]
    expected_document_inventory: tuple[str, ...]
    retrieved_documents: tuple[str, ...]
    indexed_expected_documents: tuple[str, ...]
    expected_document_hits: tuple[str, ...]
    corpus_blockers: tuple[str, ...]
    legal_correctness: str = "NOT_MEASURED_REQUIRES_HUMAN_REVIEW"


@dataclass(frozen=True)
class RetrievalDiagnostic:
    """Content-free persisted retrieval evidence for one captured run."""

    run_present: bool
    decision: str = ""
    reason: str = ""
    strategy: str = ""
    citations: tuple[CitationView, ...] = ()


@dataclass(frozen=True)
class PlannerRecord:
    """The only per-case planner data retained by the runner."""

    case_id: str
    outcome: QueryPlannerOutcome
    duration_ms: float


@dataclass(frozen=True)
class ProviderCounterSnapshot:
    calls: int
    successes: int
    failures: int


@dataclass
class ProviderCallCounters:
    """Content-free provider invocation counters and durations."""

    calls: int = 0
    successes: int = 0
    failures: int = 0
    durations_ms: list[float] = field(default_factory=list)

    def snapshot(self) -> ProviderCounterSnapshot:
        return ProviderCounterSnapshot(self.calls, self.successes, self.failures)


@dataclass(frozen=True)
class SemanticQueryRecord:
    """Per-real-case embedding timing without query text or vector retention."""

    case_id: str
    calls: int
    successes: int
    failures: int
    duration_ms: float


@dataclass
class SemanticEmbeddingCounters:
    """Aggregate label-only semantic query counters."""

    calls: int = 0
    successes: int = 0
    failures: int = 0
    durations_ms: list[float] = field(default_factory=list)


class LabelOnlySemanticEmbeddingPort:
    """Time query embedding without inspecting, logging, or storing text/vectors."""

    def __init__(
        self,
        delegate: Any,
        counters: SemanticEmbeddingCounters,
        *,
        case_id: str | None = None,
        records: dict[str, SemanticQueryRecord] | None = None,
    ) -> None:
        self._delegate = delegate
        self._counters = counters
        self._case_id = case_id
        self._records = records

    async def embed_query(self, text: str) -> Any:
        started = perf_counter()
        self._counters.calls += 1
        succeeded = False
        try:
            result = await self._delegate.embed_query(text)
            succeeded = True
            self._counters.successes += 1
            return result
        except BaseException:
            self._counters.failures += 1
            raise
        finally:
            duration_ms = (perf_counter() - started) * 1000
            self._counters.durations_ms.append(duration_ms)
            if self._case_id is not None and self._records is not None:
                self._records[self._case_id] = SemanticQueryRecord(
                    case_id=self._case_id,
                    calls=1,
                    successes=int(succeeded),
                    failures=int(not succeeded),
                    duration_ms=duration_ms,
                )

    async def embed_documents(self, texts: Any) -> Any:
        return await self._delegate.embed_documents(texts)


@dataclass
class RerankerCounters:
    calls: int = 0
    successes: int = 0
    failures: int = 0
    durations_ms: list[float] = field(default_factory=list)


@dataclass(frozen=True)
class RerankerRecord:
    case_id: str
    outcome: str
    duration_ms: float
    strategy_version: str = ""
    pre_chunk_count: int = 0
    pre_document_version_count: int = 0
    post_collapse_document_version_count: int = 0
    final_document_version_count: int = 0
    fallback: bool = False


@dataclass(frozen=True)
class MetadataRepairRecord:
    case_id: str
    semantic_candidate_count: int
    identity_candidate_count: int
    title_candidate_count: int
    ambiguous_identity_count: int
    metadata_no_support_count: int
    final_count: int
    fallback: bool
    strategy_version: str


class LabelOnlyRerankerPort:
    """Measure reranking without retaining query, hydrated text, candidate IDs, or logits."""

    def __init__(
        self,
        delegate: Any,
        counters: RerankerCounters,
        case_id: str | None,
        records: dict[str, RerankerRecord],
    ) -> None:
        self._delegate = delegate
        self._counters = counters
        self._case_id = case_id
        self._records = records

    async def rerank(self, request: Any) -> Any:
        started = perf_counter()
        self._counters.calls += 1
        outcome = "SUCCESS"
        try:
            result = await self._delegate.rerank(request)
            self._counters.successes += 1
            return result
        except BaseException:
            outcome = "FALLBACK"
            self._counters.failures += 1
            raise
        finally:
            duration = (perf_counter() - started) * 1000
            self._counters.durations_ms.append(duration)
            if self._case_id is not None:
                self._records[self._case_id] = RerankerRecord(self._case_id, outcome, duration)


def _reranked_observer(
    case_id: str | None, records: dict[str, RerankerRecord]
) -> Callable[[RerankedRetrievalDiagnostics], None]:
    """Associate content-free repository diagnostics with the runner's external case label."""

    def observe(diagnostics: RerankedRetrievalDiagnostics) -> None:
        if case_id is None:
            return
        previous = records.get(case_id, RerankerRecord(case_id, "NOT_CALLED", 0.0))
        records[case_id] = RerankerRecord(
            case_id=case_id,
            outcome="FALLBACK" if diagnostics.reranker_fallback else previous.outcome,
            duration_ms=previous.duration_ms,
            strategy_version=diagnostics.strategy_version,
            pre_chunk_count=diagnostics.pre_rerank_chunk_candidate_count,
            pre_document_version_count=diagnostics.pre_rerank_document_version_count,
            post_collapse_document_version_count=diagnostics.post_collapse_document_version_count,
            final_document_version_count=diagnostics.final_citation_document_version_count,
            fallback=diagnostics.reranker_fallback,
        )

    return observe


def _metadata_repair_observer(
    case_id: str | None, records: dict[str, MetadataRepairRecord]
) -> Callable[[MetadataRepairDiagnostics], None]:
    def observe(diagnostics: MetadataRepairDiagnostics) -> None:
        if case_id is not None:
            records[case_id] = MetadataRepairRecord(
                case_id,
                diagnostics.semantic_candidate_count,
                diagnostics.exact_identity_candidate_count,
                diagnostics.title_candidate_count,
                diagnostics.ambiguous_identity_count,
                diagnostics.metadata_no_support_count,
                diagnostics.final_count,
                diagnostics.reranker_fallback,
                diagnostics.strategy_version,
            )

    return observe


class LabelOnlyProviderProxy:
    """Count calls without inspecting or retaining provider request/result content."""

    def __init__(self, label: str, delegate: Any, counters: ProviderCallCounters) -> None:
        self.label = label
        self._delegate = delegate
        self.counters = counters

    async def generate(self, request: GenerationRequest) -> GenerationResult:
        started = perf_counter()
        self.counters.calls += 1
        try:
            result = await self._delegate.generate(request)
        except asyncio.CancelledError:
            self.counters.failures += 1
            raise
        except Exception:
            self.counters.failures += 1
            raise
        else:
            self.counters.successes += 1
            return result
        finally:
            self.counters.durations_ms.append((perf_counter() - started) * 1000)

    async def health_check(self) -> ProviderHealth:
        return await self._delegate.health_check()

    async def aclose(self) -> None:
        """The runner owns and closes the shared delegate exactly once."""

        return None


class LabelOnlyProviderOutputParser:
    """Record only a parser's safe failure class for one externally labelled case."""

    def __init__(self, delegate: Any, case_id: str, records: dict[str, str]) -> None:
        self._delegate = delegate
        self._case_id = case_id
        self._records = records

    def parse(self, output: str) -> Any:
        try:
            answer = self._delegate.parse(output)
        except ChatError as error:
            self._records[self._case_id] = (
                error.provider_output_class.value
                if error.provider_output_class is not None
                else ProviderOutputFailureClass.UNKNOWN.value
            )
            raise
        except Exception:
            self._records[self._case_id] = ProviderOutputFailureClass.UNKNOWN.value
            raise
        self._records[self._case_id] = "NONE"
        return answer


class LabelOnlyProviderOutputProvider:
    """Record safe pre-parser classes without retaining a provider response."""

    def __init__(
        self,
        delegate: Any,
        case_id: str,
        records: dict[str, str],
        max_response_bytes: int,
    ) -> None:
        self._delegate = delegate
        self._case_id = case_id
        self._records = records
        self._max_response_bytes = max_response_bytes

    async def generate(self, request: GenerationRequest) -> Any:
        result = await self._delegate.generate(request)
        if not isinstance(result, GenerationResult):
            self._records[self._case_id] = ProviderOutputFailureClass.PORT_RESULT_TYPE.value
        elif len(result.text.encode("utf-8")) > self._max_response_bytes:
            self._records[self._case_id] = ProviderOutputFailureClass.RESPONSE_BYTES.value
        return result

    async def health_check(self) -> ProviderHealth:
        return await self._delegate.health_check()

    async def aclose(self) -> None:
        return None


class RecordingPlanner:
    """Associate a planner outcome with a fixed case without retaining its plan or input."""

    def __init__(
        self, delegate: LLMQueryPlanner, case_id: str, records: dict[str, PlannerRecord]
    ) -> None:
        self._delegate = delegate
        self._case_id = case_id
        self._records = records

    async def plan(self, question: str) -> Any:
        started = perf_counter()
        outcome = QueryPlannerOutcome.INVALID_OUTPUT
        try:
            result = await self._delegate.plan(question)
            outcome = result.outcome
            return result
        finally:
            self._records[self._case_id] = PlannerRecord(
                case_id=self._case_id,
                outcome=outcome,
                duration_ms=(perf_counter() - started) * 1000,
            )


class MechanicalProvider:
    """Deterministic provider fake; it neither logs nor retains request content."""

    async def generate(self, request: GenerationRequest) -> GenerationResult:
        del request
        return GenerationResult(
            text='{"answer":"Dạ, em đã tổng hợp các căn cứ được truy hồi để thầy/cô tham khảo."}',
            provider="mechanical-fake",
            model="deterministic-json",
            request_id=None,
            duration_ms=0,
        )

    async def health_check(self) -> ProviderHealth:
        return ProviderHealth(
            status=ProviderHealthStatus.HEALTHY,
            provider="mechanical-fake",
            model="deterministic-json",
            duration_ms=0,
        )

    async def aclose(self) -> None:
        return None


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _metadata_sources(value: object) -> tuple[str, ...]:
    upper = _text(value).upper()
    return tuple(source for source in SOURCES if re.search(rf"\b{source}\b", upper))


def _metadata_documents(value: object) -> tuple[str, ...]:
    """Extract document-number-like tokens without treating prose as an expected document."""

    value_text = _text(value)
    values = re.findall(
        r"\d+(?:\s*/\s*[\wÀ-ỹĐđ.-]+){1,3}", value_text, flags=re.UNICODE
    )
    return tuple(dict.fromkeys(re.sub(r"\s+", "", item) for item in values))


def parse_stress_workbook(path: Path) -> tuple[StressCase, ...]:
    """Read precisely Q01--Q10; no grading field is part of a question input."""

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "Câu hỏi" not in workbook.sheetnames or "Khung chấm" not in workbook.sheetnames:
            raise ValueError("required stress workbook sheets are missing")
        question_sheet = workbook["Câu hỏi"]
        grading_sheet = workbook["Khung chấm"]
        questions = {
            _text(row[0]): _text(row[1])
            for row in question_sheet.iter_rows(min_row=5, values_only=True)
            if len(row) >= 2 and _text(row[0]) in CASE_IDS and _text(row[1])
        }
        grading: dict[str, tuple[str, tuple[str, ...], tuple[str, ...], str, str]] = {}
        for row in grading_sheet.iter_rows(min_row=5, values_only=True):
            if not row or _text(row[0]) not in CASE_IDS:
                continue
            padded = tuple(row) + (None,) * 6
            grading[_text(padded[0])] = (
                _text(padded[1]),
                _metadata_sources(padded[2]),
                _metadata_documents(padded[3]),
                _text(padded[4]),
                _text(padded[5]),
            )
        if tuple(sorted(questions)) != CASE_IDS or tuple(sorted(grading)) != CASE_IDS:
            raise ValueError("workbook must contain exactly Q01 through Q10 in both sheets")
        return tuple(
            StressCase(
                case_id=case_id,
                question=questions[case_id],
                topic=grading[case_id][0],
                expected_sources=grading[case_id][1],
                expected_documents=grading[case_id][2],
                minimum=grading[case_id][3],
                pass_condition=grading[case_id][4],
            )
            for case_id in CASE_IDS
        )
    finally:
        workbook.close()


def percentile(values: Iterable[float], percent: int) -> float:
    """Nearest-rank percentile, deterministic for small bounded stress samples."""

    ordered = sorted(values)
    if not ordered:
        return 0.0
    index = max(0, min(len(ordered) - 1, (len(ordered) * percent + 99) // 100 - 1))
    return float(ordered[index])


def latency_metrics(rows: Iterable[CallRow | ApiRow]) -> dict[str, float | int]:
    values = [row.latency_ms for row in rows]
    return {
        "calls": len(values),
        "p50_ms": percentile(values, 50),
        "p95_ms": percentile(values, 95),
        "p99_ms": percentile(values, 99),
    }


def score_case(
    case: StressCase, citations: Iterable[CitationView], catalog_statuses: dict[str, str]
) -> CoverageRow:
    """Score only source/document retrieval coverage, never legal correctness."""

    citation_list = tuple(citations)
    retrieved_sources = tuple(sorted({item.source_id for item in citation_list}))
    retrieved_documents = tuple(
        sorted({item.document_number for item in citation_list if item.document_number})
    )
    expected_sources = case.expected_sources
    source_hits = set(expected_sources) & set(retrieved_sources)
    source_percent = 100.0 * len(source_hits) / len(expected_sources) if expected_sources else 0.0
    expected_documents = case.expected_documents
    inventory = tuple(
        f"{item}:{catalog_statuses.get(item, 'NOT_IN_CATALOG')}" for item in expected_documents
    )
    indexed = tuple(
        item for item in expected_documents if catalog_statuses.get(item) == "INDEXED"
    )
    document_hits = tuple(item for item in expected_documents if item in set(retrieved_documents))
    blockers = tuple(
        f"{item}:{catalog_statuses.get(item, 'NOT_IN_CATALOG')}"
        for item in expected_documents
        if catalog_statuses.get(item) != "INDEXED"
    )
    return CoverageRow(
        case_id=case.case_id,
        expected_sources=expected_sources,
        retrieved_sources=retrieved_sources,
        source_coverage_percent=source_percent,
        expected_documents=expected_documents,
        expected_document_inventory=inventory,
        retrieved_documents=retrieved_documents,
        indexed_expected_documents=indexed,
        expected_document_hits=document_hits,
        corpus_blockers=blockers,
    )


def _provenance_label(transport_trust_mode: str, provenance_type: str) -> str:
    """Expose only the derived trust label from persisted evidence metadata."""

    try:
        return evidence_trust_label_for(
            coerce_transport_trust_mode(transport_trust_mode), ProvenanceType(provenance_type)
        ).value
    except ValueError:
        return "UNRESOLVED"


async def load_retrieval_diagnostics(
    session_factory: Any, rows: Iterable[CallRow]
) -> dict[Any, RetrievalDiagnostic]:
    """Load the persisted run/citation view before the runner deletes captured runs."""

    run_ids = {row.retrieval_run_id for row in rows if row.retrieval_run_id is not None}
    if not run_ids:
        return {}
    diagnostics: dict[Any, RetrievalDiagnostic] = {}
    async with session_factory() as session:
        run_rows = tuple(
            await session.execute(
                select(
                    RetrievalRun.id,
                    RetrievalRun.evidence_decision,
                    RetrievalRun.evidence_reason,
                    RetrievalRun.strategy_version,
                ).where(RetrievalRun.id.in_(run_ids))
            )
        )
        for run_id, decision, reason, strategy in run_rows:
            diagnostics[run_id] = RetrievalDiagnostic(True, decision, reason, strategy)
        citation_rows = tuple(
            await session.execute(
                select(
                    CitationRecord.retrieval_run_id,
                    SourceProvenanceRecord.source_id,
                    DocumentVersion.document_number,
                    DocumentVersion.title,
                    DocumentVersion.canonical_url,
                    SourceProvenanceRecord.transport_trust_mode,
                    SourceProvenanceRecord.provenance_type,
                )
                .select_from(CitationRecord)
                .join(DocumentChunk, CitationRecord.document_chunk_id == DocumentChunk.id)
                .join(DocumentVersion, DocumentChunk.document_version_id == DocumentVersion.id)
                .join(LegalDocument, DocumentVersion.document_id == LegalDocument.id)
                .join(
                    SourceProvenanceRecord,
                    CitationRecord.source_provenance_record_id == SourceProvenanceRecord.id,
                )
                .where(CitationRecord.retrieval_run_id.in_(run_ids))
                .order_by(CitationRecord.retrieval_run_id, CitationRecord.rank)
            )
        )
    grouped: dict[Any, list[CitationView]] = {}
    for (
        run_id,
        source_id,
        document_number,
        title,
        url,
        trust_mode,
        provenance_type,
    ) in citation_rows:
        grouped.setdefault(run_id, []).append(
            CitationView(
                source_id=source_id,
                document_number=document_number,
                title=title,
                url=url,
                provenance_label=_provenance_label(trust_mode, provenance_type),
            )
        )
    return {
        run_id: RetrievalDiagnostic(
            run_present=diagnostic.run_present,
            decision=diagnostic.decision,
            reason=diagnostic.reason,
            strategy=diagnostic.strategy,
            citations=tuple(grouped.get(run_id, ())),
        )
        for run_id, diagnostic in diagnostics.items()
    }


def apply_retrieval_diagnostics(
    rows: Iterable[CallRow], diagnostics: dict[Any, RetrievalDiagnostic]
) -> None:
    """Populate report-only evidence fields independently of the final provider result."""

    for row in rows:
        diagnostic = diagnostics.get(row.retrieval_run_id)
        if diagnostic is None:
            continue
        row.retrieval_run_present = diagnostic.run_present
        row.retrieval_decision = diagnostic.decision
        row.retrieval_reason = diagnostic.reason
        row.retrieval_strategy = diagnostic.strategy
        row.citations = diagnostic.citations
        row.citation_count = len(diagnostic.citations)
        row.source_count = len({citation.source_id for citation in diagnostic.citations})


def score_real_cases(
    cases: tuple[StressCase, ...], real_rows: Iterable[CallRow], catalog_statuses: dict[str, str]
) -> list[CoverageRow]:
    """Score persisted retrieval citations, never the final chat result citation tuple."""

    citations_by_case = {row.case_id: row.citations for row in real_rows}
    return [
        score_case(case, citations_by_case.get(case.case_id, ()), catalog_statuses)
        for case in cases
    ]


async def _call_chat(
    service: GroundedChatService,
    case: StressCase,
    *,
    lane: str,
    concurrency: int,
    round_number: int,
    captured_run_ids: set[Any],
    provider_output_classes: dict[str, str] | None = None,
) -> CallRow:
    started = perf_counter()
    try:
        # Only this field is passed to the chat seam; grading remains post-result metadata.
        result = await service.respond(ChatRequest(question=case.question))
        if result.retrieval_run_id is not None:
            captured_run_ids.add(result.retrieval_run_id)
        return CallRow(
            lane=lane,
            case_id=case.case_id,
            concurrency=concurrency,
            round_number=round_number,
            latency_ms=(perf_counter() - started) * 1000,
            outcome=result.outcome.value,
            reason=result.reason.value,
            citation_count=0,
            source_count=0,
            error_code="" if result.outcome.value != "REFUSAL" else result.reason.value,
            answer=result.answer if lane == "REAL_SHINE" else "",
            provider=result.provider or "",
            model=result.model or "",
            request_id_present=result.provider_request_id is not None,
            retrieval_run_id=result.retrieval_run_id,
            provider_output_class=(
                provider_output_classes.get(case.case_id, "NOT_APPLICABLE")
                if provider_output_classes is not None
                else "NOT_APPLICABLE"
            ),
        )
    except Exception:
        return CallRow(
            lane=lane,
            case_id=case.case_id,
            concurrency=concurrency,
            round_number=round_number,
            latency_ms=(perf_counter() - started) * 1000,
            outcome="SAFE_EXCEPTION",
            reason="SAFE_EXCEPTION",
            citation_count=0,
            source_count=0,
            error_code="SAFE_EXCEPTION",
        )


async def _bounded_calls(
    service: GroundedChatService,
    cases: tuple[StressCase, ...],
    *,
    lane: str,
    concurrency: int,
    rounds: int,
    captured_run_ids: set[Any],
) -> list[CallRow]:
    semaphore = asyncio.Semaphore(concurrency)

    async def one(case: StressCase, round_number: int) -> CallRow:
        async with semaphore:
            return await _call_chat(
                service,
                case,
                lane=lane,
                concurrency=concurrency,
                round_number=round_number,
                captured_run_ids=captured_run_ids,
            )

    return list(
        await asyncio.gather(
            *(one(case, round_number) for round_number in range(1, rounds + 1) for case in cases)
        )
    )


def _counter_delta(
    before: ProviderCounterSnapshot, after: ProviderCounterSnapshot
) -> ProviderCounterSnapshot:
    return ProviderCounterSnapshot(
        calls=after.calls - before.calls,
        successes=after.successes - before.successes,
        failures=after.failures - before.failures,
    )


async def _sequential_real_calls(
    cases: tuple[StressCase, ...],
    *,
    service_for_case: Any,
    captured_run_ids: set[Any],
    planner_counters: ProviderCallCounters,
    answer_counters: ProviderCallCounters,
    provider_output_classes: dict[str, str],
) -> tuple[list[CallRow], dict[str, ProviderCounterSnapshot], dict[str, ProviderCounterSnapshot]]:
    """Execute Q01--Q10 in order and attach provider counts without request-text state."""

    rows: list[CallRow] = []
    planner_by_case: dict[str, ProviderCounterSnapshot] = {}
    answer_by_case: dict[str, ProviderCounterSnapshot] = {}
    for case in cases:
        planner_before = planner_counters.snapshot()
        answer_before = answer_counters.snapshot()
        rows.append(
            await _call_chat(
                service_for_case(case),
                case,
                lane="REAL_SHINE",
                concurrency=1,
                round_number=1,
                captured_run_ids=captured_run_ids,
                provider_output_classes=provider_output_classes,
            )
        )
        planner_by_case[case.case_id] = _counter_delta(planner_before, planner_counters.snapshot())
        answer_by_case[case.case_id] = _counter_delta(answer_before, answer_counters.snapshot())
    return rows, planner_by_case, answer_by_case


async def api_probe(
    base_url: str, requests_per_endpoint: int = 50, concurrency: int = 10
) -> list[ApiRow]:
    """Probe health endpoints only; response bodies and network exceptions are never retained."""

    base_url = base_url.rstrip("/")
    semaphore = asyncio.Semaphore(concurrency)
    timeout = httpx.Timeout(10.0)
    async with httpx.AsyncClient(timeout=timeout) as client:
        async def one(endpoint: str) -> ApiRow:
            async with semaphore:
                started = perf_counter()
                try:
                    response = await client.get(f"{base_url}{endpoint}")
                    error_code = (
                        "" if 200 <= response.status_code < 400 else f"HTTP_{response.status_code}"
                    )
                except httpx.HTTPError:
                    error_code = "NETWORK_ERROR"
                return ApiRow(endpoint, (perf_counter() - started) * 1000, error_code)

        return list(
            await asyncio.gather(
                *(
                    one(endpoint)
                    for endpoint in ("/live", "/ready")
                    for _ in range(requests_per_endpoint)
                )
            )
        )


async def _catalog_statuses(session_factory: Any, cases: tuple[StressCase, ...]) -> dict[str, str]:
    expected = {document for case in cases for document in case.expected_documents}
    if not expected:
        return {}
    async with session_factory() as session:
        rows = tuple(
            await session.execute(
                select(CorpusCatalogEntry.document_number, CorpusCatalogEntry.processing_status)
            )
        )
    statuses: dict[str, str] = {}
    for document_number, status in rows:
        normalized = _metadata_documents(document_number)
        for number in normalized:
            if number in expected:
                statuses[number] = str(status)
    return statuses


async def _cleanup(session_factory: Any, run_ids: set[Any]) -> str:
    if not run_ids:
        return "NOT_NEEDED"
    try:
        async with session_factory.begin() as session:
            await session.execute(
                delete(CitationRecord).where(CitationRecord.retrieval_run_id.in_(run_ids))
            )
            await session.execute(delete(RetrievalRun).where(RetrievalRun.id.in_(run_ids)))
        return "COMPLETED"
    except Exception:
        return "FAILED_SAFE"


def _join(values: Iterable[str]) -> str:
    return "; ".join(value for value in values if value)


def _autosize(sheet: Any) -> None:
    for column in sheet.columns:
        letter = column[0].column_letter
        content_width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[letter].width = min(45, max(12, content_width))


def _real_result_values(row: CallRow | None) -> tuple[Any, ...]:
    if row is None:
        return ("", "NOT_RUN", "NOT_RUN", "", 0, 0, "NOT_RUN", "", "", "", "", "", "", "", False)
    return (
        round(row.latency_ms, 3),
        row.outcome,
        row.reason,
        row.answer,
        row.citation_count,
        row.source_count,
        row.error_code,
        _join(citation.source_id for citation in row.citations),
        _join(citation.document_number or "" for citation in row.citations),
        _join(citation.title or "" for citation in row.citations),
        _join(citation.url or "" for citation in row.citations),
        _join(citation.provenance_label for citation in row.citations),
        row.provider,
        row.model,
        row.request_id_present,
    )


def _planner_diagnostic_values(
    case: StressCase,
    *,
    planner_enabled: bool,
    planner_records: dict[str, PlannerRecord],
    planner_by_case: dict[str, ProviderCounterSnapshot],
    answer_by_case: dict[str, ProviderCounterSnapshot],
    real_by_case: dict[str, CallRow],
    coverage_by_case: dict[str, CoverageRow],
) -> tuple[Any, ...]:
    record = planner_records.get(case.case_id)
    zero_counters = ProviderCounterSnapshot(0, 0, 0)
    planner_stats = planner_by_case.get(case.case_id, zero_counters)
    answer_stats = answer_by_case.get(case.case_id, zero_counters)
    chat_row = real_by_case.get(case.case_id)
    coverage_row = coverage_by_case.get(case.case_id)
    indexed_hit_count = 0
    if coverage_row is not None:
        indexed_hit_count = len(
            set(coverage_row.indexed_expected_documents) & set(coverage_row.retrieved_documents)
        )
    return (
        case.case_id,
        planner_enabled,
        record is not None,
        record.outcome.value if record is not None else "NOT_CALLED",
        round(record.duration_ms, 3) if record is not None else "",
        planner_stats.calls,
        planner_stats.successes,
        planner_stats.failures,
        answer_stats.calls,
        answer_stats.successes,
        answer_stats.failures,
        chat_row.outcome if chat_row is not None else "NOT_RUN",
        chat_row.reason if chat_row is not None else "NOT_RUN",
        chat_row.retrieval_run_present if chat_row is not None else False,
        chat_row.retrieval_decision if chat_row is not None else "",
        chat_row.retrieval_reason if chat_row is not None else "",
        chat_row.retrieval_strategy if chat_row is not None else "",
        chat_row.citation_count if chat_row is not None else 0,
        chat_row.source_count if chat_row is not None else 0,
        indexed_hit_count,
    )


def _semantic_diagnostic_values(
    case: StressCase,
    *,
    semantic_mode: str,
    real_by_case: dict[str, CallRow],
    coverage_by_case: dict[str, CoverageRow],
    semantic_records: dict[str, SemanticQueryRecord],
) -> tuple[object, ...]:
    """Build a content-free semantic diagnostics row without retaining query embeddings."""

    row = real_by_case.get(case.case_id)
    coverage = coverage_by_case.get(case.case_id)
    embedding = semantic_records.get(case.case_id)
    indexed_hits = 0
    if coverage is not None:
        indexed_hits = len(
            set(coverage.indexed_expected_documents) & set(coverage.retrieved_documents)
        )
    return (
        case.case_id,
        semantic_mode,
        row.retrieval_decision if row is not None else "NOT_RUN",
        row.retrieval_strategy if row is not None else "NOT_RUN",
        row.citation_count if row is not None else 0,
        row.source_count if row is not None else 0,
        indexed_hits,
        round(embedding.duration_ms, 3) if embedding is not None else "",
    )


def _reranker_diagnostic_values(
    case: StressCase,
    *,
    enabled: bool,
    records: dict[str, RerankerRecord],
    real_by_case: dict[str, CallRow],
    coverage_by_case: dict[str, CoverageRow],
) -> tuple[object, ...]:
    record = records.get(case.case_id)
    row = real_by_case.get(case.case_id)
    coverage = coverage_by_case.get(case.case_id)
    expected_hits = 0
    if coverage is not None:
        expected_hits = len(
            set(coverage.indexed_expected_documents) & set(coverage.retrieved_documents)
        )
    cited_documents = (
        {item.document_number for item in row.citations if item.document_number} if row else set()
    )

    indexed_expected = set(coverage.indexed_expected_documents) if coverage is not None else set()
    non_expected_rate: float | str = ""
    if cited_documents:
        non_expected_rate = len(cited_documents - indexed_expected) / len(cited_documents)
    return (
        case.case_id,
        enabled,
        record is not None,
        record.outcome if record is not None else "NOT_CALLED",
        round(record.duration_ms, 3) if record is not None else "",
        row.retrieval_decision if row is not None else "NOT_RUN",
        row.retrieval_strategy if row is not None else "NOT_RUN",
        record.pre_chunk_count if record is not None else "",
        record.final_document_version_count if record is not None else "",
        row.citation_count if row is not None else 0,
        row.source_count if row is not None else 0,
        expected_hits,
        non_expected_rate,
    )


def _provider_output_diagnostic_values(
    case: StressCase,
    *,
    real_by_case: dict[str, CallRow],
    mechanical_by_case: dict[str, CallRow],
    answer_by_case: dict[str, ProviderCounterSnapshot],
) -> tuple[object, ...]:
    """Return the workbook's content-free provider-output view for one real case."""

    row = real_by_case.get(case.case_id)
    if row is None:
        mechanical = mechanical_by_case.get(case.case_id)
        if mechanical is None:
            return (case.case_id, "NOT_RUN", "NOT_RUN", False, 0, 0, "NOT_RUN")
        return (
            case.case_id,
            mechanical.outcome,
            mechanical.reason,
            mechanical.retrieval_run_present,
            mechanical.citation_count,
            0,
            "NOT_APPLICABLE",
        )
    provider_calls = answer_by_case.get(case.case_id, ProviderCounterSnapshot(0, 0, 0)).calls
    return (
        case.case_id,
        row.outcome,
        row.reason,
        row.retrieval_run_present,
        row.citation_count,
        provider_calls,
        row.provider_output_class,
    )


def _table_sheet(sheet: Any, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(list(row))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    _autosize(sheet)


def write_report(
    output: Path,
    cases: tuple[StressCase, ...],
    mechanical: list[CallRow],
    real: list[CallRow],
    api_rows: list[ApiRow],
    coverage: list[CoverageRow],
    *,
    cleanup_status: str,
    infrastructure_status: str,
    mechanical_wall_seconds: float = 0.0,
    real_wall_seconds: float = 0.0,
    api_wall_seconds: float = 0.0,
    lexical_repair_enabled: bool = False,
    planner_enabled: bool = False,
    planner_records: dict[str, PlannerRecord] | None = None,
    planner_by_case: dict[str, ProviderCounterSnapshot] | None = None,
    answer_by_case: dict[str, ProviderCounterSnapshot] | None = None,
    planner_counters: ProviderCallCounters | None = None,
    answer_counters: ProviderCallCounters | None = None,
    semantic_mode: str = "off",
    semantic_profile_id: str = "",
    semantic_counters: SemanticEmbeddingCounters | None = None,
    semantic_records: dict[str, SemanticQueryRecord] | None = None,
    rerank_enabled: bool = False,
    reranker_counters: RerankerCounters | None = None,
    reranker_records: dict[str, RerankerRecord] | None = None,
    metadata_repair_enabled: bool = False,
    metadata_repair_records: dict[str, MetadataRepairRecord] | None = None,
    quality_strategy: str = "disabled",
    quality_selected_pool: int | None = None,
) -> None:
    """Write the fixed report sheets; credentials are intentionally not accepted."""

    planner_records = planner_records or {}
    planner_by_case = planner_by_case or {}
    answer_by_case = answer_by_case or {}
    planner_counters = planner_counters or ProviderCallCounters()
    answer_counters = answer_counters or ProviderCallCounters()
    semantic_counters = semantic_counters or SemanticEmbeddingCounters()
    semantic_records = semantic_records or {}
    reranker_counters = reranker_counters or RerankerCounters()
    reranker_records = reranker_records or {}
    metadata_repair_records = metadata_repair_records or {}
    reranker_fallback_count = sum(record.fallback for record in reranker_records.values())
    real_by_case = {row.case_id: row for row in real}
    mechanical_by_case = {row.case_id: row for row in mechanical}
    provider_output_classes = Counter(
        _provider_output_diagnostic_values(
            case,
            real_by_case=real_by_case,
            mechanical_by_case=mechanical_by_case,
            answer_by_case=answer_by_case,
        )[-1]
        for case in cases
    )
    workbook = openpyxl.Workbook()
    summary = workbook.active
    assert summary is not None
    summary.title = "Tổng hợp"
    all_calls = mechanical + real
    metrics = latency_metrics(all_calls)
    mechanical_metrics = latency_metrics(mechanical)
    real_metrics = latency_metrics(real)
    api_metrics = latency_metrics(api_rows)
    api_error_count = sum(bool(row.error_code) for row in api_rows)
    planner_outcomes = Counter(record.outcome.value for record in planner_records.values())
    planner_durations = [record.duration_ms for record in planner_records.values()]
    semantic_durations = semantic_counters.durations_ms
    chat_percentiles = (
        f"{metrics['p50_ms']:.2f} / {metrics['p95_ms']:.2f} / {metrics['p99_ms']:.2f}"
    )
    api_percentiles = " / ".join(
        f"{api_metrics[key]:.2f}" for key in ("p50_ms", "p95_ms", "p99_ms")
    )
    mechanical_throughput = (
        round(len(mechanical) / mechanical_wall_seconds, 3) if mechanical_wall_seconds else 0
    )
    real_throughput = round(len(real) / real_wall_seconds, 3) if real_wall_seconds else 0
    api_throughput = round(len(api_rows) / api_wall_seconds, 3) if api_wall_seconds else 0
    mechanical_provider_calls = sum(bool(row.provider) for row in mechanical)
    _table_sheet(
        summary,
        ["Mục", "Giá trị"],
        [
            ("Methodology", "Direct bounded chat composition; grading is post-result only."),
            ("Limits", "No live Zalo load; no legal-correctness claim; human review required."),
            (
                "Mechanical lane",
                "Mechanical retrieval load only; planner disabled; deterministic fake provider.",
            ),
            ("Mechanical calls", len(mechanical)),
            ("Mechanical provider generation calls actual", mechanical_provider_calls),
            ("Real SHINE cases attempted", len(real)),
            ("Real SHINE answer-provider calls actual", answer_counters.calls),
            (
                "Mechanical p50/p95/p99 ms",
                f"{mechanical_metrics['p50_ms']:.2f} / {mechanical_metrics['p95_ms']:.2f} / "
                f"{mechanical_metrics['p99_ms']:.2f}",
            ),
            (
                "Real chat route p50/p95/p99 ms",
                f"{real_metrics['p50_ms']:.2f} / {real_metrics['p95_ms']:.2f} / "
                f"{real_metrics['p99_ms']:.2f}",
            ),
            ("Chat p50/p95/p99 ms", chat_percentiles),
            ("Mechanical wall seconds", round(mechanical_wall_seconds, 3)),
            ("Mechanical throughput calls/s", mechanical_throughput),
            ("Real lane wall seconds", round(real_wall_seconds, 3)),
            ("Real lane throughput cases/s", real_throughput),
            ("API wall seconds", round(api_wall_seconds, 3)),
            ("API throughput calls/s", api_throughput),
            ("API p50/p95/p99 ms", api_percentiles),
            ("API HTTP/network errors", api_error_count),
            ("Cleanup", cleanup_status),
            ("Runner infrastructure", infrastructure_status),
            ("Lexical repair", "enabled" if lexical_repair_enabled else "disabled"),
            ("Quality strategy", quality_strategy),
            (
                "Quality selected pool",
                quality_selected_pool if quality_selected_pool is not None else "NOT_USED",
            ),
            ("Planner enabled", "enabled" if planner_enabled else "disabled"),
            ("Semantic mode", semantic_mode),
            ("Semantic profile", semantic_profile_id if semantic_mode != "off" else "NOT_USED"),
            ("Reranker enabled", rerank_enabled),
            ("Reranker calls", reranker_counters.calls),
            ("Reranker fallback count", reranker_fallback_count),
            (
                "Reranker success/failure",
                f"{reranker_counters.successes}/{reranker_counters.failures}",
            ),
            (
                "Reranker p50/p95 ms",
                f"{percentile(reranker_counters.durations_ms, 50):.2f} / "
                f"{percentile(reranker_counters.durations_ms, 95):.2f}",
            ),
            ("Semantic query embed calls", semantic_counters.calls),
            (
                "Semantic query embed success/failure",
                f"{semantic_counters.successes}/{semantic_counters.failures}",
            ),
            (
                "Semantic query embed p50/p95 ms",
                f"{percentile(semantic_durations, 50):.2f} / "
                f"{percentile(semantic_durations, 95):.2f}",
            ),
            ("Planner calls", planner_counters.calls),
            (
                "Planner calls success/failure",
                f"{planner_counters.successes}/{planner_counters.failures}",
            ),
            (
                "Planner p50/p95 ms",
                f"{percentile(planner_durations, 50):.2f} / "
                f"{percentile(planner_durations, 95):.2f}",
            ),
            ("Answer provider calls", answer_counters.calls),
            (
                "Answer calls success/failure",
                f"{answer_counters.successes}/{answer_counters.failures}",
            ),
            ("Actual provider failures", planner_counters.failures + answer_counters.failures),
            (
                "Retrieval versus final output",
                "Persisted retrieval evidence is independent of final chat/provider output.",
            ),
            *(
                (f"Planner outcome {outcome}", count)
                for outcome, count in sorted(planner_outcomes.items())
            ),
            *(
                (f"Provider output class {output_class}", count)
                for output_class, count in sorted(provider_output_classes.items())
            ),
        ],
    )
    results = workbook.create_sheet("Câu hỏi & Kết quả")
    _table_sheet(
        results,
        [
            "Case ID", "Question", "Latency ms", "Outcome", "Reason", "Answer",
            "Citation count", "Source count", "Error code", "Source IDs", "Document numbers",
            "Titles", "URLs", "Provenance labels", "Provider", "Model", "Request ID present",
        ],
        [
            (case.case_id, case.question, *_real_result_values(real_by_case.get(case.case_id)))
            for case in cases
        ],
    )
    mechanical_sheet = workbook.create_sheet("Mechanical Load")
    _table_sheet(
        mechanical_sheet,
        [
            "Lane", "Case ID", "Concurrency", "Round", "Latency ms", "Outcome", "Reason",
            "Citation count", "Source count", "Error code",
        ],
        (
            (
                row.lane, row.case_id, row.concurrency, row.round_number, round(row.latency_ms, 3),
                row.outcome, row.reason, row.citation_count, row.source_count, row.error_code,
            )
            for row in mechanical
        ),
    )
    api_sheet = workbook.create_sheet("API Probe")
    _table_sheet(
        api_sheet,
        ["Endpoint", "Latency ms", "HTTP error"],
        ((row.endpoint, round(row.latency_ms, 3), row.error_code) for row in api_rows),
    )
    coverage_sheet = workbook.create_sheet("Corpus Coverage")
    _table_sheet(
        coverage_sheet,
        [
            "Case ID", "Expected sources", "Retrieved sources", "Source coverage %",
            "Expected documents", "Expected document inventory", "Retrieved documents",
            "Indexed expected documents", "Expected-document hits", "Corpus blockers",
            "Narrative legal correctness",
        ],
        (
            (
                row.case_id, _join(row.expected_sources), _join(row.retrieved_sources),
                row.source_coverage_percent, _join(row.expected_documents),
                _join(row.expected_document_inventory), _join(row.retrieved_documents),
                _join(row.indexed_expected_documents), _join(row.expected_document_hits),
                _join(row.corpus_blockers), row.legal_correctness,
            )
            for row in coverage
        ),
    )
    diagnostics_sheet = workbook.create_sheet("Planner Diagnostics")
    real_by_case = {row.case_id: row for row in real}
    coverage_by_case = {row.case_id: row for row in coverage}
    _table_sheet(
        diagnostics_sheet,
        [
            "Case ID", "Planner enabled", "Planner called", "Planner outcome",
            "Planner duration ms", "Planner provider calls", "Planner provider success",
            "Planner provider failure", "Answer provider calls", "Answer provider success",
            "Answer provider failure", "Chat outcome", "Chat reason", "Retrieval run present",
            "Retrieval decision", "Retrieval reason", "Retrieval strategy", "Citation count",
            "Source count", "Expected indexed-doc hit count",
        ],
        (
            _planner_diagnostic_values(
                case,
                planner_enabled=planner_enabled,
                planner_records=planner_records,
                planner_by_case=planner_by_case,
                answer_by_case=answer_by_case,
                real_by_case=real_by_case,
                coverage_by_case=coverage_by_case,
            )
            for case in cases
        ),
    )
    semantic_sheet = workbook.create_sheet("Semantic Diagnostics")
    _table_sheet(
        semantic_sheet,
        [
            "Case ID", "Mode", "Retrieval decision", "Retrieval strategy", "Citation count",
            "Source count", "Expected indexed-doc hit count", "Query embedding duration ms",
        ],
        (
            _semantic_diagnostic_values(
                case,
                semantic_mode=semantic_mode,
                real_by_case=real_by_case,
                coverage_by_case=coverage_by_case,
                semantic_records=semantic_records,
            )
            for case in cases
        ),
    )
    reranker_sheet = workbook.create_sheet("Reranker Diagnostics")
    _table_sheet(
        reranker_sheet,
        [
            "Case ID", "Enabled", "Called", "Outcome", "Duration ms", "Retrieval decision",
            "Retrieval strategy", "Pre-rerank chunk candidate count",
            "Final document-version diversity", "Citation count",
            "Source count", "Expected indexed-doc hit count", "Non-expected candidate rate",
        ],
        (
            _reranker_diagnostic_values(
                case,
                enabled=rerank_enabled,
                records=reranker_records,
                real_by_case=real_by_case,
                coverage_by_case=coverage_by_case,
            )
            for case in cases
        ),
    )
    metadata_sheet = workbook.create_sheet("Metadata Repair Diagnostics")
    _table_sheet(
        metadata_sheet,
        [
            "Case ID", "Enabled", "Semantic candidates", "Identity candidates", "Title candidates",
            "Ambiguous identities", "Metadata no-support", "Final count", "Fallback", "Strategy",
        ],
        (
            (
                case.case_id,
                metadata_repair_enabled,
                metadata_repair_records[case.case_id].semantic_candidate_count
                if case.case_id in metadata_repair_records else 0,
                metadata_repair_records[case.case_id].identity_candidate_count
                if case.case_id in metadata_repair_records else 0,
                metadata_repair_records[case.case_id].title_candidate_count
                if case.case_id in metadata_repair_records else 0,
                metadata_repair_records[case.case_id].ambiguous_identity_count
                if case.case_id in metadata_repair_records else 0,
                metadata_repair_records[case.case_id].metadata_no_support_count
                if case.case_id in metadata_repair_records else 0,
                metadata_repair_records[case.case_id].final_count
                if case.case_id in metadata_repair_records else 0,
                metadata_repair_records[case.case_id].fallback
                if case.case_id in metadata_repair_records else False,
                metadata_repair_records[case.case_id].strategy_version
                if case.case_id in metadata_repair_records else "NOT_CALLED",
            )
            for case in cases
        ),
    )
    provider_output_sheet = workbook.create_sheet("Provider Output Diagnostics")
    _table_sheet(
        provider_output_sheet,
        [
            "Case ID",
            "Outcome",
            "Reason",
            "Retrieval run present",
            "Citation count",
            "Provider call count",
            "Output class",
        ],
        (
            _provider_output_diagnostic_values(
                case,
                real_by_case=real_by_case,
                mechanical_by_case=mechanical_by_case,
                answer_by_case=answer_by_case,
            )
            for case in cases
        ),
    )
    regression = workbook.create_sheet("Regression")
    distributions = Counter((row.lane, row.outcome, row.reason or "NONE") for row in all_calls)
    api_distributions = Counter(row.error_code or "SUCCESS" for row in api_rows)
    _table_sheet(
        regression,
        ["Category", "Value", "Count"],
        [
            *(
                (lane, f"{outcome} / {reason}", count)
                for (lane, outcome, reason), count in sorted(distributions.items())
            ),
            *(
                ("API Probe", error_code, count)
                for error_code, count in sorted(api_distributions.items())
            ),
        ],
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()


async def run(args: argparse.Namespace) -> int:
    if args.semantic_mode != "off" and (
        args.lexical_repair_enabled or args.planner_enabled
    ):
        return 2
    if args.rerank_enabled and args.semantic_mode != "semantic":
        return 2
    if args.metadata_repair_enabled and not args.rerank_enabled:
        return 2
    if args.quality_strategy != "disabled" and (
        args.semantic_mode == "off"
        or args.rerank_enabled
        or args.metadata_repair_enabled
        or args.lexical_repair_enabled
        or args.planner_enabled
        or args.quality_selected_pool is None
        or args.quality_strategy not in QUALITY_EVALUATION_PROFILES
    ):
        return 2
    cases = parse_stress_workbook(args.input)
    mechanical: list[CallRow] = []
    real: list[CallRow] = []
    api_rows: list[ApiRow] = []
    coverage: list[CoverageRow] = []
    cleanup_status = "NOT_NEEDED"
    infrastructure_status = "COMPLETED"
    mechanical_wall_seconds = 0.0
    real_wall_seconds = 0.0
    api_wall_seconds = 0.0
    engine: Any = None
    provider: Any = None
    session_factory: Any = None
    run_ids: set[Any] = set()
    planner_records: dict[str, PlannerRecord] = {}
    planner_by_case: dict[str, ProviderCounterSnapshot] = {}
    answer_by_case: dict[str, ProviderCounterSnapshot] = {}
    provider_output_classes: dict[str, str] = {}
    planner_counters = ProviderCallCounters()
    answer_counters = ProviderCallCounters()
    semantic_counters = SemanticEmbeddingCounters()
    semantic_records: dict[str, SemanticQueryRecord] = {}
    reranker_counters = RerankerCounters()
    reranker_records: dict[str, RerankerRecord] = {}
    metadata_repair_records: dict[str, MetadataRepairRecord] = {}
    quality_strategy: Any | None = None
    quality_reader: Any | None = None
    try:
        provider_settings = ProviderSettings()
        mechanical_settings = ChatSettings(retrieval_planner_enabled=False)
        real_settings = ChatSettings(retrieval_planner_enabled=args.planner_enabled)
        engine = create_engine(Settings())
        session_factory = create_session_factory(engine)
        reranker: Any | None = None
        if args.semantic_mode == "off":
            repository = PostgresLexicalRetrievalRepository(
                session_factory, SOURCES, lexical_repair_enabled=args.lexical_repair_enabled
            )
            semantic_embedder: Any | None = None
        else:
            semantic_settings = SemanticSettings()
            semantic_embedder = FastEmbedSemanticAdapter(semantic_settings)
            reranker = FastEmbedRerankerAdapter(RerankerSettings()) if args.rerank_enabled else None
            repository = PostgresHybridRetrievalRepository(
                session_factory,
                SOURCES,
                semantic_embedder,
                mode=args.semantic_mode,
            )
            if not await repository.coverage_complete():
                raise RuntimeError("SEMANTIC_COVERAGE_INCOMPLETE")
        if args.quality_strategy != "disabled":
            assert semantic_embedder is not None
            quality_strategy = materialize_strategy(
                args.quality_strategy, args.quality_selected_pool
            )
            if quality_strategy.family.reranker_enabled:
                raise RuntimeError("QUALITY_RERANK_NOT_APPROVED")
            quality_reader = PostgresQualityCandidateReader(session_factory)
            if quality_strategy.family.dynamic_evidence_enabled:
                mechanical_settings = mechanical_settings.model_copy(update={"max_citations": 6})
                real_settings = real_settings.model_copy(update={"max_citations": 6})
        retrieval = RetrievalService(repository)
        resolver = PostgresCitationResolver(session_factory)

        def retrieval_for_case(case_id: str | None) -> Any:
            if semantic_embedder is None:
                return retrieval
            embedder = LabelOnlySemanticEmbeddingPort(
                semantic_embedder,
                semantic_counters,
                case_id=case_id,
                records=semantic_records if case_id is not None else None,
            )
            if quality_strategy is not None:
                quality_pipeline = LegalQualityCandidatePipeline(
                    quality_reader,
                    embedder,
                    quality_strategy,
                    tuple(SourceId(source_id) for source_id in SOURCES),
                )
                return RetrievalService(
                    PostgresQualityRetrievalRepository(
                        session_factory, LegalQuestionAnalyzer(), quality_pipeline
                    )
                )
            if reranker is not None and args.metadata_repair_enabled:
                return RetrievalService(
                    PostgresMetadataRepairRetrievalRepository(
                        session_factory,
                        SOURCES,
                        embedder,
                        LabelOnlyRerankerPort(
                            reranker, reranker_counters, case_id, reranker_records
                        ),
                        timeout_seconds=getattr(RerankerSettings(), "timeout_seconds", 5.0),
                        observer=_metadata_repair_observer(case_id, metadata_repair_records),
                    )
                )
            if reranker is not None:
                return RetrievalService(
                    PostgresRerankedSemanticRepository(
                        session_factory,
                        SOURCES,
                        embedder,
                        LabelOnlyRerankerPort(
                            reranker, reranker_counters, case_id, reranker_records
                        ),
                        timeout_seconds=getattr(RerankerSettings(), "timeout_seconds", 5.0),
                        observer=_reranked_observer(case_id, reranker_records),
                    )
                )
            return RetrievalService(
                PostgresHybridRetrievalRepository(
                    session_factory, SOURCES, embedder, mode=args.semantic_mode
                )
            )

        def service_for(
            chat_provider: Any,
            chat_settings: ChatSettings,
            retrieval_port: Any,
            query_planner: Any = None,
            canonical_anchor_resolver: Any = None,
            parser: Any = None,
        ) -> GroundedChatService:
            return GroundedChatService(
                retrieval_port,
                PostgresGroundingEvidenceAdapter(session_factory, chat_settings),
                resolver,
                chat_provider,
                parser or StrictProviderJsonParser(),
                chat_settings,
                provider_settings,
                query_planner,
                canonical_anchor_resolver,
            )

        for level in args.mechanical_concurrency:
            lane_started = perf_counter()
            mechanical.extend(
                await _bounded_calls(
                    service_for(
                        MechanicalProvider(), mechanical_settings, retrieval_for_case(None)
                    ),
                    cases,
                    lane="MECHANICAL",
                    concurrency=level,
                    rounds=args.mechanical_rounds, captured_run_ids=run_ids,
                )
            )
            mechanical_wall_seconds += perf_counter() - lane_started
        if not args.skip_real_shine:
            provider = create_provider(provider_settings)
            planner_provider = LabelOnlyProviderProxy("planner", provider, planner_counters)
            answer_provider = LabelOnlyProviderProxy("answer", provider, answer_counters)
            planner_delegate = (
                LLMQueryPlanner(planner_provider, real_settings, provider_settings)
                if args.planner_enabled
                else None
            )
            anchor_resolver = (
                PostgresCanonicalAnchorResolver(session_factory, SOURCES)
                if args.planner_enabled
                else None
            )

            def real_service_for_case(case: StressCase) -> GroundedChatService:
                recording_planner = (
                    RecordingPlanner(planner_delegate, case.case_id, planner_records)
                    if planner_delegate is not None
                    else None
                )
                return service_for(
                    LabelOnlyProviderOutputProvider(
                        answer_provider,
                        case.case_id,
                        provider_output_classes,
                        provider_settings.max_response_bytes,
                    ),
                    real_settings,
                    retrieval_for_case(case.case_id),
                    recording_planner,
                    anchor_resolver,
                    LabelOnlyProviderOutputParser(
                        StrictProviderJsonParser(), case.case_id, provider_output_classes
                    ),
                )

            real_started = perf_counter()
            real, planner_by_case, answer_by_case = await _sequential_real_calls(
                cases,
                service_for_case=real_service_for_case,
                captured_run_ids=run_ids,
                planner_counters=planner_counters,
                answer_counters=answer_counters,
                provider_output_classes=provider_output_classes,
            )
            real_wall_seconds = perf_counter() - real_started
        api_started = perf_counter()
        api_rows = await api_probe(args.api_base_url)
        api_wall_seconds = perf_counter() - api_started
        diagnostics = await load_retrieval_diagnostics(session_factory, (*mechanical, *real))
        apply_retrieval_diagnostics((*mechanical, *real), diagnostics)
        statuses = await _catalog_statuses(session_factory, cases)
        coverage = score_real_cases(cases, real, statuses)
    except Exception:
        infrastructure_status = "FAILED_SAFE"
    finally:
        if engine is not None and session_factory is not None:
            cleanup_status = await _cleanup(session_factory, run_ids)
        if provider is not None:
            try:
                await provider.aclose()
            except Exception:
                infrastructure_status = "FAILED_SAFE"
        if engine is not None:
            await engine.dispose()
        write_report(
            args.output, cases, mechanical, real, api_rows, coverage,
            cleanup_status=cleanup_status, infrastructure_status=infrastructure_status,
            mechanical_wall_seconds=mechanical_wall_seconds,
            real_wall_seconds=real_wall_seconds,
            api_wall_seconds=api_wall_seconds,
            lexical_repair_enabled=args.lexical_repair_enabled,
            planner_enabled=args.planner_enabled,
            planner_records=planner_records,
            planner_by_case=planner_by_case,
            answer_by_case=answer_by_case,
            planner_counters=planner_counters,
            answer_counters=answer_counters,
            semantic_mode=args.semantic_mode,
            semantic_profile_id=SEMANTIC_PROFILE_ID if args.semantic_mode != "off" else "",
            semantic_counters=semantic_counters,
            semantic_records=semantic_records,
            rerank_enabled=args.rerank_enabled,
            reranker_counters=reranker_counters,
            reranker_records=reranker_records,
            metadata_repair_enabled=args.metadata_repair_enabled,
            metadata_repair_records=metadata_repair_records,
            quality_strategy=args.quality_strategy,
            quality_selected_pool=args.quality_selected_pool,
        )
    return 0 if infrastructure_status == "COMPLETED" else 2


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Bounded legal-chatbot stress runner")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--mechanical-concurrency", type=int, nargs="+", default=[1, 2, 5])
    parser.add_argument("--mechanical-rounds", type=int, default=2)
    parser.add_argument("--skip-real-shine", action="store_true")
    parser.add_argument("--lexical-repair-enabled", action="store_true")
    parser.add_argument("--planner-enabled", action="store_true")
    parser.add_argument("--semantic-mode", choices=("off", "semantic", "hybrid"), default="off")
    parser.add_argument("--rerank-enabled", action="store_true")
    parser.add_argument("--metadata-repair-enabled", action="store_true")
    parser.add_argument(
        "--quality-strategy",
        choices=("disabled", *sorted(QUALITY_EVALUATION_PROFILES)),
        default="disabled",
    )
    parser.add_argument("--quality-selected-pool", type=int, choices=(8, 12, 16, 20))
    parser.add_argument("--api-base-url", default="http://127.0.0.1:8000")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    if args.mechanical_rounds < 1 or any(level < 1 for level in args.mechanical_concurrency):
        return 2
    if args.semantic_mode != "off" and (
        args.lexical_repair_enabled or args.planner_enabled
    ):
        return 2
    if args.rerank_enabled and args.semantic_mode != "semantic":
        return 2
    if args.metadata_repair_enabled and (not args.rerank_enabled or args.semantic_mode != "semantic"):
        return 2
    try:
        return asyncio.run(run(args))
    except Exception:
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
